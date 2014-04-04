# -*- coding: utf-8 -*-
import sys
import uuid

# Create Test DB
from django.conf import settings
from django.contrib.auth.models import User, Group, Permission, ContentType
from django.contrib.auth.hashers import make_password
from repanier.models import Customer
from repanier.models import Producer
from repanier.models import Staff
from repanier.models import Product
from repanier.models import LUT_ProductionMode
from repanier.models import LUT_DepartmentForCustomer

from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from openpyxl import load_workbook

def get_header(worksheet):
  header = []
  if worksheet:
    row_num = 0
    col_num = 0
    c = worksheet.cell(row=row_num, column=col_num)
    while (c.value!=None ) and (col_num < 50):
      header.append(c.value)
      col_num+=1
      c = worksheet.cell(row=row_num, column=col_num)
  return header

def get_row(worksheet, header, row_num):
  row = {}
  if worksheet:
    # last_row is a row with all cells empty
    last_row = True
    for col_num, col_header in enumerate(header):
      c = worksheet.cell(row=row_num, column=col_num)
      if c.value:
        last_row = False
      row[col_header]=c.value
  if last_row:
    row = {}
  return row

def add_all_customers(worksheet):
  print("---------- Customers ----------")
  header = get_header(worksheet)
  if header:
    row_num = 1
    row = get_row(worksheet, header, row_num)
    while row:

      print row

      user = None
      user_set = User.objects.filter(username = row['short_basket_name'])[:1]
      if user_set:
        for user in user_set:
          user.password = make_password("=test=")
          user.is_superuser=False
          user.is_staff=False
          user.is_active=True
          user.email=row['email'] if row['email'] else row['short_basket_name'] + u"@no-spam.ws"
          user.first_name="N/A"
          user.last_name="N/A"
          user.save()
      else:
        user = User.objects.create(username = row['short_basket_name'], 
          password = make_password(uuid.uuid1().hex),
          is_superuser=False, is_staff=False, is_active=True,
          email=row['email'] if row['email'] else row['short_basket_name'] + u"@no-spam.ws",
          first_name="N/A",
          last_name="N/A")

      customer = None
      customer_set = Customer.objects.filter(user_id = user.id)[:1]
      if customer_set:
        for customer in customer_set:
          customer.short_basket_name = row['short_basket_name']
          customer.long_basket_name = row['long_basket_name']
          customer.phone1 = row['phone1']
          customer.phone2 = row['phone2']
          customer.address = row['address']
          customer.vat_id = row['vat_id']
          customer.balance = row['balance']
          customer.is_active = True
          customer.save()
      else:
        Customer.objects.create(
          user_id = user.id,
          short_basket_name = row['short_basket_name'],
          long_basket_name = row['long_basket_name'],
          phone1 = row['phone1'],
          phone2 = row['phone2'],
          address = row['address'],
          date_balance = row['date_balance'],
          balance = row['balance'],
          vat_id = row['vat_id'],
          is_active = True)

      row_num += 1
      row = get_row(worksheet, header, row_num)


def add_all_staffs(worksheet):
  print("---------- Staffs ----------")
  header = get_header(worksheet)
  if header:
    row_num = 1
    row = get_row(worksheet, header, row_num)
    while row:

      print row

      user = None
      user_set = User.objects.filter(username = row['username'])[:1]
      if user_set:
        for user in user_set:
          user.password = make_password("=test=")
          user.is_superuser=False
          user.is_staff=False
          user.is_active=True
          user.email=row['email'] if row['email'] else row['username'] + u"@no-spam.ws"
          user.first_name="N/A"
          user.last_name="N/A"
          user.save()
      else:
        user = User.objects.create(username = row['username'], 
          password = make_password(uuid.uuid1().hex),
          is_superuser=False, is_staff=False, is_active=True,
          email=row['email'] if row['email'] else row['username'] + u"@no-spam.ws",
          first_name="N/A",
          last_name="N/A")

      customer_responsible_id = None
      if row['customer_responsible']:
        customer_responsible = Customer.objects.get(
          short_basket_name = row['customer_responsible'])

      staff = None
      staff_set = Staff.objects.filter(user_id = user.id)[:1]
      if staff_set:
        for staff in staff_set:
          staff.long_name = row['long_name']
          staff.customer_responsible_id = customer_responsible.id
          staff.is_reply_to_order_email = ( row['is_reply_to_order_email'] != None )
          staff.is_reply_to_invoice_email = ( row['is_reply_to_invoice_email'] != None )
          staff.is_active = True
          staff.save()
      else:
        Staff.objects.create(
          user_id = user.id,
          long_name = row['long_name'],
          customer_responsible_id = customer_responsible.id,
          is_reply_to_order_email = ( row['is_reply_to_order_email'] != None ),
          is_reply_to_invoice_email = ( row['is_reply_to_invoice_email'] != None ),
          is_active = True)

      row_num += 1
      row = get_row(worksheet, header, row_num)

def add_all_producers(worksheet):
  print("---------- Producers ----------")
  header = get_header(worksheet)
  if header:
    row_num = 1
    row = get_row(worksheet, header, row_num)
    while row:

      print row

      producer = None
      producer_set = Producer.objects.filter(short_profile_name = row['short_profile_name'])[:1]
      if producer_set:
        for producer in producer_set:
          producer.long_profile_name = row['long_profile_name']
          producer.email = row['email'] if row['email'] else row['short_profile_name'] + u"@no-spam.ws"
          producer.phone1 = row['phone1']
          producer.phone2 = row['phone2']
          producer.fax = row['fax']
          producer.bank_account = row['bank_account']
          producer.vat_id = row['vat_id']
          producer.address = row['address']
          producer.price_list_multiplier = row['price_list_multiplier']
          producer.date_balance = row['date_balance']
          producer.balance = row['balance']
          producer.is_active = True
          producer.save()
      else:
        Producer.objects.create(
          short_profile_name = row['short_profile_name'],
          long_profile_name = row['long_profile_name'],
          email = row['email'] if row['email'] else row['short_profile_name'] + u"@no-spam.ws",
          phone1 = row['phone1'],
          phone2 = row['phone2'],
          fax = row['fax'],
          address = row['address'],
          price_list_multiplier = row['price_list_multiplier'],
          date_balance = row['date_balance'],
          balance = row['balance'],
          is_active = True)

      row_num += 1
      row = get_row(worksheet, header, row_num)


def add_all_products(worksheet):
  print("---------- Products ----------")
  header = get_header(worksheet)
  if header:
    row_num = 1
    row = get_row(worksheet, header, row_num)
    while row:

      print row

      producer = None
      if row['producer']:
        producer = Producer.objects.get(
          short_profile_name = row['producer'])

      production_mode_id = None
      if row['production_mode']:
        production_mode_set = LUT_ProductionMode.objects.filter(
          short_name = row['production_mode'])[:1]
        if production_mode_set:
          production_mode_id = production_mode_set[0].id
        else:
          production_mode = LUT_ProductionMode.objects.create(
            short_name = row['production_mode'],
            is_active = True)
          if production_mode!= None:
            production_mode_id = production_mode.id 

      department_for_customer_id = None
      if row['department_for_customer']:
        department_for_customer_set = LUT_DepartmentForCustomer.objects.filter(
          short_name = row['department_for_customer'])[:1]
        if department_for_customer_set:
          department_for_customer_id = department_for_customer_set[0].id
        else:
          department_for_customer = LUT_DepartmentForCustomer.objects.create(
            short_name = row['department_for_customer'],
            is_active = True)
          if department_for_customer!= None:
            department_for_customer_id = department_for_customer.id 

      product = None
      product_set = Product.objects.filter(
        producer_id = producer.id,
        long_name = row['long_name']
        )[:1]
      if product_set:
        for product in product_set:
          product.producer_id = producer.id
          product.long_name = row['long_name']
          product.production_mode_id = production_mode_id
          product.department_for_customer_id = department_for_customer_id
          product.order_by_kg_pay_by_kg = ( row['order_by_kg_pay_by_kg'] != None )
          product.order_by_piece_pay_by_kg = ( row['order_by_piece_pay_by_kg'] != None )
          product.order_average_weight = row['order_average_weight']
          product.order_by_piece_pay_by_piece = ( row['order_by_piece_pay_by_piece'] != None )
          product.producer_must_give_order_detail_per_customer = ( row['producer_must_give_order_detail_per_customer'] != None )
          product.original_unit_price = row['producer_unit_price']
          product.customer_minimum_order_quantity = row['customer_minimum_order_quantity']
          product.customer_increment_order_quantity = row['customer_increment_order_quantity']
          product.customer_alert_order_quantity = row['customer_alert_order_quantity']
          product.is_active = True
          product.save()
      else:
        product = Product.objects.create(
          producer_id = producer.id,
          long_name = row['long_name'],
          production_mode_id = production_mode_id,
          department_for_customer_id = department_for_customer_id,
          order_by_kg_pay_by_kg = ( row['order_by_kg_pay_by_kg'] != None ),
          order_by_piece_pay_by_kg = ( row['order_by_piece_pay_by_kg'] != None ),
          order_average_weight = row['order_average_weight'],
          order_by_piece_pay_by_piece = ( row['order_by_piece_pay_by_piece'] != None ),
          producer_must_give_order_detail_per_customer = ( row['producer_must_give_order_detail_per_customer'] != None ),
          original_unit_price = row['producer_unit_price'],
          customer_minimum_order_quantity = row['customer_minimum_order_quantity'],
          customer_increment_order_quantity = row['customer_increment_order_quantity'],
          customer_alert_order_quantity = row['customer_alert_order_quantity'],
          is_into_offer = False,
          is_active = True
        )

      row_num += 1
      row = get_row(worksheet, header, row_num)

  order = 0
  for obj in Product.objects.all().order_by(
      "producer__short_profile_name", "long_name"):
    order += 1
    obj.product_order = order
    obj.save()

#  producer.followed_by.add(1,3,4,5)

def main():
  wb = load_workbook("intial_db.xlsx")
  add_all_customers(wb.get_sheet_by_name("Customers"))
  add_all_staffs(wb.get_sheet_by_name("Staffs"))
  add_all_producers(wb.get_sheet_by_name("Producers"))
  add_all_products(wb.get_sheet_by_name("Products"))

if __name__ == '__main__':
        if '--version' in sys.argv:
                print(__version__)
        else:
                main();