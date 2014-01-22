# -*- coding: utf-8 -*-
import re

from const import *
from django.conf import settings
from django.conf.urls import patterns, url
from django.contrib import admin
from django.contrib.auth import get_user_model
from django.contrib.auth.models import User
from django.core import urlresolvers

from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.utils import timezone
from django.core import validators
from django.core.exceptions import ValidationError
from django.db import models
from django.db.models import Q, F
from django import forms

from adminsortable.admin import SortableAdminMixin
# from repanier.adminsortable import SortableAdminMixin

from repanier.models import LUT_ProductionMode
from repanier.models import LUT_DepartmentForCustomer
from repanier.models import LUT_DepartmentForProducer
from repanier.models import LUT_PermanenceRole

from repanier.models import Producer
from repanier.models import Permanence
from repanier.models import SiteProducer
from repanier.models import Customer
from repanier.models import SiteCustomer
from repanier.models import SiteStaff
from repanier.models import Product
from repanier.models import PermanenceBoard
from repanier.models import OfferItem
from repanier.models import PermanenceInPreparation
from repanier.models import PermanenceDone
from repanier.models import Purchase
from repanier.models import BankAccount


# import django
# class LocalizedModelForm(forms.ModelForm):
#     def __new__(cls, *args, **kwargs):
#         new_class = super(LocalizedModelForm, cls).__new__(cls, *args, **kwargs)
#         for field in new_class.base_fields.values():
#             if isinstance(field, django.forms.DecimalField):
#                 field.localize = True
#                 field.widget.is_localized = True
#         return new_class

# Filters in the right sidebar of the change list page of the admin
from django.contrib.admin import SimpleListFilter


class ProductFilterBySiteProducer(SimpleListFilter):
	# Human-readable title which will be displayed in the
	# right admin sidebar.
	title = _("producers")
	# Parameter for the filter that will be used in the URL query.
	parameter_name = 'site_producer'

	def lookups(self, request, model_admin):
		"""
		Returns a list of tuples. The first element in each
		tuple is the coded value for the option that will
		appear in the URL query. The second element is the
		human-readable name for the option that will appear
		in the right sidebar.
		"""
		# This list is a collection of site producer.id, .name
		return [(c.id, c.short_profile_name) for c in 
			SiteProducer.objects.all().active().with_login()
			]

	def queryset(self, request, queryset):
		"""
		Returns the filtered queryset based on the value
		provided in the query string and retrievable via
		`self.value()`.
		"""
		# This query set is a collection of products
		if self.value():
			return queryset.site_producer_is(self.value())
		else:
			return queryset


class ProductFilterByDepartmentForProducer(SimpleListFilter):
	title = _("departments for producer")
	parameter_name = 'department_for_producer'

	def lookups(self, request, model_admin):
		# This list is a collection of site department.id, .name
		return [(c.id, c.short_name) for c in 
			LUT_DepartmentForProducer.objects.all().active()
			]

	def queryset(self, request, queryset):
		# This query set is a collection of products
		if self.value():
			return queryset.department_for_producer_is(self.value())
		else:
			return queryset

class ProductFilterByDepartmentForThisProducer(SimpleListFilter):
	title = _("departments for producer")
	parameter_name = 'department_for_this_producer'

	def lookups(self, request, model_admin):
		site_producer = request.GET.get('site_producer')
		inner_qs = Product.objects.all().active().site_producer_is(
			site_producer).order_by(
			'department_for_producer__id').distinct(
			'department_for_producer__id')
		# return []
		return [('all', _('All'))] + [(c.id, c.short_name) for c in 
			LUT_DepartmentForProducer.objects.all().active().filter(product__in=inner_qs)
			]

	def choices(self, cl):
		for lookup, title in self.lookup_choices:
			yield {
				'selected': self.value() == lookup,
				'query_string': cl.get_query_string({
					self.parameter_name: lookup,
				}, []),
				'display': title,
			}

	def queryset(self, request, queryset):
		# This query set is a collection of products
		if self.value()=='all':
			site_producer = request.GET.get('site_producer')
			return queryset.site_producer_is(site_producer)
		else:
			return queryset.department_for_producer_is(self.value())

# LUT
class LUT_ProductionModeAdmin(admin.ModelAdmin):
	list_display = ('short_name', 'is_active')
	list_display_links = ('short_name',)
	exclude = ['site',]
	list_max_show_all = True

	# def queryset(self, request):
	# 	qs = super(LUT_ProductionModeAdmin, self).queryset(request)
	# 	return qs.filter(site=settings.SITE_ID)

admin.site.register(LUT_ProductionMode, LUT_ProductionModeAdmin)

class LUT_DepartmentForCustomerAdmin(admin.ModelAdmin):
	list_display = ('short_name', 'is_active')
	list_display_links = ('short_name',)
	exclude = ['site',]
	list_max_show_all = True

admin.site.register(LUT_DepartmentForCustomer, LUT_DepartmentForCustomerAdmin)

class LUT_DepartmentForProducerAdmin(admin.ModelAdmin):
	list_display = ('short_name', 'is_active')
	list_display_links = ('short_name',)
	exclude = ['site',]
	list_max_show_all = True

admin.site.register(LUT_DepartmentForProducer, LUT_DepartmentForProducerAdmin)

class LUT_PermanenceRoleAdmin(admin.ModelAdmin):
	list_display = ('short_name', 'is_active')
	list_display_links = ('short_name',)
	exclude = ['site',]
	list_max_show_all = True

admin.site.register(LUT_PermanenceRole, LUT_PermanenceRoleAdmin)

# Custom User
class UserDataForm(forms.ModelForm):

	username = forms.CharField(label=_('Username'), max_length=30, 
			help_text=_(
				'Required. 30 characters or fewer. Letters, numbers and @/./+/-/_ characters'
			),
			validators=[
				validators.RegexValidator(re.compile('^[\w.@+-]+$'), _('Enter a valid username.'), 'invalid')
			])
	password1 = forms.CharField(label=_('Password1'), max_length=128, required=False)
	password2 = forms.CharField(label=_('Password2'), max_length=128, required=False)
	email = forms.EmailField(label=_('Email'))
	first_name = forms.CharField(label=_('First_name'), max_length=30)
	last_name = forms.CharField(label=_('Last_name'), max_length=30)

	def __init__(self, *args, **kwargs):
		super(UserDataForm, self).__init__(*args, **kwargs)
		self.user = None

	def error(self,field, msg):
		if field not in self._errors:
			self._errors[field]= self.error_class([msg])

	def clean(self, *args, **kwargs):
		# The SiteStaff has no first_name or last_name because it's a function with login/pwd.
		# A SiteCustomer with a first_name and last_name is responsible of this funcition. 
		if any(self.errors):
			if not self['first_name'].html_name in self.data:
				if 'first_name' in self._errors:
					del self._errors['first_name']
				self.data['first_name'] = self.fields['first_name'].initial
			if not self['last_name'].html_name in self.data:
				if 'last_name' in self._errors:
					del self._errors['last_name']
				self.data['last_name'] = self.fields['last_name'].initial
		# Check that the password is set when it's a new user. 
		password1 = self.cleaned_data.get("password1")
		password2 = self.cleaned_data.get("password2")
		initial_username = self.fields['username'].initial
		if initial_username=='':
			if not password1:
				self.error('password1',_('The given password must be set'))
		# Check that the two password entries match
		if password1 and password2 and password1 != password2:
			self.error('password1',_("Passwords must match"))
		# Check that the user name is set. The required attribute is not a waranty
		# that the browser hasn't removed it.
		username = self.cleaned_data.get("username")
		if not username:
			self.error('username',_('The given username must be set'))
		# Check that the email is set
		email = self.cleaned_data.get("email")
		if not email:
			self.error('email',_('The given email must be set'))
		# Check that the email is not already used
		user=None
		email = User.objects.normalize_email(email)
		try:
			user = User.objects.get(email=email)
		except User.DoesNotExist:
			pass
		if user != None and initial_username!=user.username:
			self.error('email',_('The given email is used by another user'))
		# Check that the username is not already used
		if self.fields['username'].initial!=username:
			user=None
			try:
				user = User.objects.get(username=username)
			except User.DoesNotExist:
				pass
			if user != None:
				self.error('username',_('The given username is used by another user'))
		super(UserDataForm, self).clean(*args, **kwargs)
		return self.cleaned_data

	def save(self, *args, **kwargs):
		super(UserDataForm, self).save(*args, **kwargs)
		change = (self.instance.id != None)
		username = self.data['username']
		email = self.data['email']
		password = self.data['password1']
		first_name = self.data['first_name']
		last_name = self.data['last_name']
		user = None
		if change:
			user=User.objects.get(id=self.instance.user_id)
			user.username = username
			user.email = email
			user.first_name = first_name
			user.last_name = last_name
			if password:
				user.set_password(password)
			user.save()
		else:
			user = User.objects.create_user(
				username=username, email=email, password=password,
				first_name=first_name, last_name=last_name)
		self.user = user
		return self.instance

# Producer
class ProducerWithUserDataForm(UserDataForm):

	class Meta:
		model = Producer

class ProducerWithUserDataAdmin(admin.ModelAdmin):
	form = ProducerWithUserDataForm
	fields = ['username', 'password1', 'password2', 'email', 'first_name', 
		'last_name', 'phone1', 'phone2', 'fax', 'address',
		'bank_account', 'is_active']
	search_fields = ('user__username',)
	list_display = ('__unicode__', 'phone1', 'address', 'is_active')
	list_max_show_all = True

	def get_form(self,request, obj=None, **kwargs):
		form = super(ProducerWithUserDataAdmin,self).get_form(request, obj, **kwargs)
		username = form.base_fields['username']
		password1 = form.base_fields['password1']
		password1.initial = ''
		password2 = form.base_fields['password2']
		password2.initial = ''
		email = form.base_fields['email']
		first_name= form.base_fields['first_name']
		last_name= form.base_fields['last_name']
		if obj:
			user_model = get_user_model()
			user = user_model.objects.get(id=obj.user_id)
			username.initial = getattr(user, user_model.USERNAME_FIELD)
			# username.widget.attrs['readonly'] = True
			email.initial = user.email
			first_name.initial = user.first_name
			last_name.initial = user.last_name
		else:
			# Clean data displayed
			username.initial = ''
			# username.widget.attrs['readonly'] = False
			password1.initial = ''
			password2.initial = ''
			email.initial = ''
			first_name.initial = ''
			last_name.initial = ''
		return form

	def save_model(self, request, obj, form, change):
		obj.user = form.user
		super(ProducerWithUserDataAdmin,self).save_model(
			request, obj, form, change)

admin.site.register(Producer, ProducerWithUserDataAdmin)


class SiteProducerAdmin(admin.ModelAdmin):
	fields = ['producer', 'short_profile_name', 'long_profile_name', 'memo',
	 'date_previous_balance', 'previous_balance', 'amount_in', 'amount_out',
	 'represent_this_buyinggroup', 'is_active']
	exclude = ['site',]
	readonly_fields = ('date_previous_balance', 
		'previous_balance', 
		'amount_in', 
		'amount_out')
	search_fields = ('short_profile_name',)
	list_display = ('short_profile_name', 'get_products', 'get_producer_phone1', 'get_producer_address', 'represent_this_buyinggroup',
		'is_active')
	list_max_show_all = True

	def get_producer_phone1(self, obj):
		if obj.producer:
			return '%s'%(obj.producer.phone1)
		else:
			return ''
	get_producer_phone1.short_description = _("phone1") 

	def get_producer_address(self, obj):
		if obj.producer:
			return '%s'%(obj.producer.address)
		else:
			return ''
	get_producer_address.short_description = _("address") 

	def get_form(self,request, obj=None, **kwargs):
		form = super(SiteProducerAdmin,self).get_form(request, obj, **kwargs)
		producer = form.base_fields["producer"]
		# producer.widget.can_add_related = False

		if obj:
			# Don't allow to change the producer/login
			producer.empty_label = None
			producer.initial = obj.producer
			if obj.producer:
				producer.queryset = Producer.objects.id(obj.producer_id)
			else:
				producer.queryset = Producer.objects.none()
		else:
			# Don't allow to add the same producer twice
			producer.queryset = Producer.objects.not_producer_of_the_buyinggroup()
		return form

	# def queryset(self, request):
	# 	qs = super(SiteProducerAdmin, self).queryset(request)
	# 	return qs.filter(site=settings.SITE_ID)
admin.site.register(SiteProducer, SiteProducerAdmin)

# Customer
class CustomerWithUserDataForm(UserDataForm):

	class Meta:
		model = Customer

class CustomerWithUserDataAdmin(admin.ModelAdmin):
	form = CustomerWithUserDataForm
	fields = ['username', 'password1', 'password2', 'email', 'first_name', 
		'last_name', 'phone1', 'phone2', 'address',
		'is_active']
	search_fields = ('user__username',)
	list_display = ('__unicode__', 'phone1', 'phone2', 'address', 'is_active')
	list_max_show_all = True

	def get_form(self,request, obj=None, **kwargs):
		form = super(CustomerWithUserDataAdmin,self).get_form(request, obj, **kwargs)
		# https://docs.djangoproject.com/en/1.5/topics/auth/customizing/#django.contrib.auth.models.AbstractBaseUser
		username = form.base_fields['username']
		password1 = form.base_fields['password1']
		password1.initial = ''
		password2 = form.base_fields['password2']
		password2.initial = ''
		email = form.base_fields['email']
		first_name= form.base_fields['first_name']
		last_name= form.base_fields['last_name']
		if obj:
			print("obj")
			user_model = get_user_model()
			user = user_model.objects.get(id=obj.user_id)
			username.initial = getattr(user, user_model.USERNAME_FIELD)
			# username.widget.attrs['readonly'] = False
			email.initial = user.email
			first_name.initial = user.first_name
			last_name.initial = user.last_name
		else:
			# Clean data displayed
			username.initial = ''
			# username.widget.attrs['readonly'] = False
			password1.initial = ''
			password2.initial = ''
			email.initial = ''
			first_name.initial = ''
			last_name.initial = ''
		return form

	def save_model(self, request, obj, form, change):
		obj.user = form.user
		super(CustomerWithUserDataAdmin,self).save_model(
			request, obj, form, change)

admin.site.register(Customer, CustomerWithUserDataAdmin)


class SiteCustomerAdmin(admin.ModelAdmin):
	fields = ('customer', 'short_basket_name', 'long_basket_name',
	 'date_previous_balance', 'previous_balance', 
	 'amount_in', 'amount_out',
	 'represent_this_buyinggroup', 'is_active')
	exclude = ['site',]
	readonly_fields = ('date_previous_balance', 'previous_balance',
	 'amount_in', 'amount_out')
	search_fields = ('short_basket_name',)
	list_display = ('__unicode__', 'represent_this_buyinggroup',
		'is_active')
	list_max_show_all = True

	def get_form(self,request, obj=None, **kwargs):
		form = super(SiteCustomerAdmin,self).get_form(request, obj, **kwargs)
		customer = form.base_fields["customer"]
		# customer.widget.can_add_related = False

		if obj:
			# Don't allow to change the customer/login
			customer.empty_label = None
			customer.initial = obj.customer
			if obj.customer:
				customer.queryset = Customer.objects.id(obj.customer_id)
			else:
				customer.queryset = Customer.objects.none()
		else:
			# Don't allow to add the same customer twice
			customer.queryset = Customer.objects.not_customer_of_the_buyinggroup()
		return form

	# def queryset(self, request):
	# 	qs = super(SiteCustomerAdmin, self).queryset(request)
	# 	return qs.filter(site=settings.SITE_ID)
admin.site.register(SiteCustomer, SiteCustomerAdmin)

# SiteStaff
class SiteStaffWithUserDataForm(UserDataForm):

	class Meta:
		model = SiteStaff

class SiteStaffWithUserDataAdmin(admin.ModelAdmin):
	form = SiteStaffWithUserDataForm
	fields = ['username', 'password1', 'password2', 'email', 
		'customer_responsible','long_name', 'memo', 'is_active']
	exclude = ['site',]
	list_display = ('__unicode__', 'customer_responsible', 'get_sitecustomer_phone1', 'is_active')
	list_max_show_all = True

	def get_form(self,request, obj=None, **kwargs):
		form = super(SiteStaffWithUserDataAdmin,self).get_form(request, obj, **kwargs)
		username = form.base_fields['username']
		password1 = form.base_fields['password1']
		password1.initial = ''
		password2 = form.base_fields['password2']
		password2.initial = ''
		email = form.base_fields['email']
		first_name= form.base_fields['first_name']
		last_name= form.base_fields['last_name']
		customer_responsible = form.base_fields["customer_responsible"]
		customer_responsible.widget.can_add_related = False

		if obj:
			user_model = get_user_model()
			user = user_model.objects.get(id=obj.user_id)
			username.initial = getattr(user, user_model.USERNAME_FIELD)
			# username.widget.attrs['readonly'] = True
			email.initial = user.email
			first_name.initial = user.first_name
			last_name.initial = user.last_name
			customer_responsible.empty_label = None
			customer_responsible.initial = obj.customer_responsible
		else:
			# Clean data displayed
			username.initial = ''
			# username.widget.attrs['readonly'] = False
			password1.initial = ''
			password2.initial = ''
			email.initial = ''
			first_name.initial = 'N/A'
			last_name.initial = 'N/A'
		customer_responsible.queryset = SiteCustomer.objects.all(
			).active().with_login().order_by(
			"short_basket_name")
		return form

	def save_model(self, request, obj, form, change):
		obj.user = form.user
		form.user.is_staff = True
		form.user.save()
		super(SiteStaffWithUserDataAdmin,self).save_model(
			request, obj, form, change)

admin.site.register(SiteStaff, SiteStaffWithUserDataAdmin)

class ProductAdmin(SortableAdminMixin, admin.ModelAdmin):
	list_display = ('site_producer',
		'department_for_producer',
		'long_name',
		'is_into_offer',
		'producer_unit_price',
		'order_average_weight',
		'order_by_piece_pay_by_kg',
		'producer_must_give_order_detail_per_customer',
		'customer_minimum_order_quantity',
		'customer_increment_order_quantity',
		'customer_alert_order_quantity',
		'is_active')
	list_display_links = ('long_name',)
	# list_editable = ('is_active', 
	# 	'producer_unit_price',
	# 	'order_average_weight')
	readonly_fields = ('is_created_on', 
		'is_updated_on')
	exclude = ['site',]
	list_max_show_all = True
	# ordering = ('site_producer', 
	# 	'department_for_producer',
	# 	'long_name',)
	search_fields = ('long_name',)
	list_filter = ('is_active',
		ProductFilterBySiteProducer, 
		ProductFilterByDepartmentForProducer,)
	actions = ['flip_flop_select_for_offer_status', 'duplicate_product'	]

	def flip_flop_select_for_offer_status(self, request, queryset):
		queryset.active().is_not_selected_for_offer().update(is_into_offer=None)
		queryset.is_selected_for_offer().update(is_into_offer=False)
		queryset.is_waiting_to_be_selected_for_offer().update(is_into_offer=True)

	flip_flop_select_for_offer_status.short_description = _(
		'flip_flop_select_for_offer_status for offer')

	def duplicate_product(self, request, queryset):
		for product in queryset:
			super(ProductAdmin,self).move_for_duplicate(product)
			long_name_prefix = "COPY_OF_"
			length_long_name_prefix = len(long_name_prefix)
			max_length = Product._meta.get_field('long_name').max_length
			if len(product.long_name) + length_long_name_prefix > max_length:
				product.long_name = long_name_prefix + \
					product.long_name[:max_length-length_long_name_prefix ]
			else:
				product.long_name = long_name_prefix + product.long_name
			product.id = None
			product.save()

	duplicate_product.short_description = _('duplicate product')

	def get_form(self,request, obj=None, **kwargs):
		form = super(ProductAdmin,self).get_form(request, obj, **kwargs)
		site_producer = form.base_fields["site_producer"]
		department_for_producer = form.base_fields["department_for_producer"]
		department_for_customer = form.base_fields["department_for_customer"]
		production_mode = form.base_fields["production_mode"]
		site_producer.widget.can_add_related = False
		department_for_producer.widget.can_add_related = False
		department_for_customer.widget.can_add_related = False
		production_mode.widget.can_add_related = False

		if obj:
			site_producer.empty_label = None
			department_for_producer.empty_label = None
			department_for_customer.empty_label = None
			production_mode.empty_label = None
		site_producer.queryset = SiteProducer.objects.all(
			).active().with_login()
		department_for_producer.queryset = LUT_DepartmentForProducer.objects.all(
			).active()
		department_for_customer.queryset = LUT_DepartmentForCustomer.objects.all(
			).active()
		production_mode.queryset = LUT_ProductionMode.objects.all(
			).active()
		return form

	def get_list_filter(self, request):
		site_producer = request.GET.get('department_for_this_producer')
		if site_producer:
			return ('is_active', ProductFilterByDepartmentForThisProducer,)
		else:
			return self.list_filter

	# def queryset(self, request):
	# 	qs = super(ProductAdmin, self).queryset(request)
	#  	return qs.order_by('order',)

admin.site.register(Product, ProductAdmin)

# Permanence
class PermanenceBoardInline(admin.TabularInline):
	model = PermanenceBoard
	fields = ['permanence_role', 'site_customer']
	extra = 1

	def formfield_for_foreignkey(self, db_field, request, **kwargs):
		if db_field.name == "site_customer":
			kwargs["queryset"] = SiteCustomer.objects.all(
				).with_login().active()
		if db_field.name == "permanence_role":
			kwargs["queryset"] = LUT_PermanenceRole.objects.all(
				).active()
		return super(PermanenceBoardInline, self).formfield_for_foreignkey(db_field, request, **kwargs)

class OfferItemInline(admin.TabularInline):
	model = OfferItem
	fields = ['product', 'producer_unit_price', 'get_total_order_quantity']
	readonly_fields = ('product', 'get_total_order_quantity')
	extra = 0
	max_num = 0

class PermanenceInPreparationAdmin(admin.ModelAdmin):
	fieldsets = [
		('Permanance', 
			{'fields': ['distribution_date', 'short_name', 'status', 
			'memo', 'producers']}
		),
	]
	# readonly_fields = ('status', 'is_created_on', 'is_updated_on')
	exclude = ['site']
	list_max_show_all = True
	filter_horizontal = ('producers',)
	inlines = [PermanenceBoardInline, OfferItemInline]
	date_hierarchy = 'distribution_date'
	list_display = ('__unicode__', 'get_siteproducers', 'get_sitecustomers', 'get_board', 'status')
	ordering = ('distribution_date',)
	actions = ['planify', 'open_orders','close_orders', 
		'send_orders_to_producers', 'back_to_previous_status', 
		'export_xlsx', 'import_xlsx', 'send_email', 'export_docx'
	]

	def export_docx(self, request, queryset):
		from docx import *
		import cStringIO
		response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
		response['Content-Disposition'] = 'attachment; filename=mymodel.docx'
		# Create our properties, contenttypes, and other support files
		title    = 'Python docx demo'
		subject  = 'A practical example of making docx from Python'
		creator  = 'Mike MacCana'
		keywords = ['python', 'Office Open XML', 'Word']

		relationships = relationshiplist()
		document = newdocument()
		body = document.xpath('/w:document/w:body', namespaces=nsprefixes)[0]
		body.append(heading(u"Titre àéè§ùµ", 1))
		body.append(heading("Sous-titre 1", 2))
		body.append(paragraph("Texte"))

		points = [ 
			'Remarque 1',
			'Remarque 2',
			'Remarque 3'
		]
		for point in points:
			body.append(paragraph(point, style='ListNumber')) 

		body.append(heading("Sous-titre 2", 2))
		body.append(paragraph("Table"))
		tbl_rows = [['Id', 'Date de distribution 2', 'Status']]
		for obj in queryset:
			row = [
				str(obj.pk),
				obj.distribution_date.strftime('%Y/%m/%d'),
				obj.status,
			]
			tbl_rows.append(row)
		body.append(table(tbl_rows))
		coreprops = coreproperties(title=title, subject=subject, creator=creator,
			keywords=keywords)
		appprops = appproperties()
		contenttypes = contenttypes()
		websettings = websettings()
		wordrelationships = wordrelationships(relationships)

		# Save our document
		fobj = cStringIO.StringIO()
		savedocx(document, coreprops, appprops, contenttypes, websettings,
			wordrelationships, fobj)
		a = fobj.getvalue() 
		fobj.close() 
		response.write(a) 
		return response

	export_docx.short_description = _("Export DOCX")


	def send_email(self, request, queryset):
		from django.core.mail import send_mail, BadHeaderError
		from django.http import HttpResponseRedirect

		try:
			send_mail('Test sujet', 'Test msg', 'ask.it@repanier.be', ['pcolmant@gmail.com'])
		except BadHeaderError:
			self.message_user(request, _("Invalid header found."))

		# return HttpResponseRedirect(request.get_full_path())

	send_email.short_description = _("Send e-mail")

	class ImportXlsxForm(forms.Form):
		_selected_action = forms.CharField(widget=forms.MultipleHiddenInput)
		file_to_import = forms.FileField(
			label=_('File to import'), allow_empty_file=False
		)

	def handle_uploaded_file(self, request, queryset, file_to_import):
		from openpyxl import load_workbook

		wb = load_workbook(file_to_import)
		ws = wb.get_active_sheet()
		i = 0
		c = ws.cell(row=i, column=0)
		while ( c.value!=None ) and (i < 10000):
			d = ws.cell(row=i, column=1)
			print("%s - %s") % (c.value.encode('utf-8'), d.value.encode('utf-8'))
			i = i + 1
			c = ws.cell(row=i, column=0)
		c = ws.cell(row=0, column=1)
		# for permanence in queryset:
		# 	permanence.short_name = c.value.encode('utf-8')
		# 	permanence.save()
		return

	def import_xlsx(self, request, queryset):
		# http://www.jpichon.net/blog/2010/08/django-admin-actions-and-intermediate-pages/
		from views import render_response
		from django.http import HttpResponseRedirect

		form = None
		if 'apply' in request.POST:
			form = self.ImportXlsxForm(request.POST, request.FILES)
			if form.is_valid():
				file_to_import = request.FILES['file_to_import']
				if('.xlsx' in file_to_import.name) and (file_to_import.size <= 1000000):
					self.handle_uploaded_file(request, queryset, file_to_import)
					self.message_user(request, _("Successfully imported %s.") % (file_to_import.name))
				else:
					self.message_user(request, 
						_("Error when importing %s : File size must be <= 1 Mb and extension must be .xlsx") % (file_to_import.name)
					)
			else:
				self.message_user(request, _("Error when importing %s.") % (file_to_import.name))
			return HttpResponseRedirect(request.get_full_path())
		if not form:
			form = self.ImportXlsxForm(
				initial={'_selected_action': request.POST.getlist(admin.ACTION_CHECKBOX_NAME)}
			)
		return render_response(request, 'import_xlsx.html', {
				'permanences': queryset,
				'import_xlsx_form': form,
		})

	import_xlsx.short_description = _("Import XLSX")

	def export_xlsx(self, request, queryset):
		import openpyxl
		from openpyxl.cell import get_column_letter
		response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
		response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
		wb = openpyxl.Workbook()
		ws = wb.get_active_sheet()
		ws.title = "MyModel"

		row_num = 0

		columns = [
			(u"ID", 15),
			(u"distribution_date", 70),
			(u"status", 70),
		]

		for col_num in xrange(len(columns)):
			c = ws.cell(row=row_num, column=col_num)
			c.value = columns[col_num][0]
			c.style.font.bold = True
			# set column width
			ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

		for obj in queryset:
			row_num += 1
			row = [
				obj.pk,
				obj.distribution_date,
				obj.status,
			]
			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.value = row[col_num]
				c.style.alignment.wrap_text = True

		wb.save(response)
		return response

	export_xlsx.short_description = _("Export XLSX")

	def get_readonly_fields(self, request, obj=None):
		if obj:
			status = obj.status
			if status>PERMANENCE_PLANIFIED:
				return('status', 'is_created_on', 'is_updated_on','producers')
		return ('status', 'is_created_on', 'is_updated_on')

	def planify(self, request, queryset):
		queryset.filter(status=PERMANENCE_DISABLED).update(status=PERMANENCE_PLANIFIED)

	planify.short_description = _('planify permanence')

	def open_orders(self, request, queryset):
		for permanence in queryset.filter(status=PERMANENCE_PLANIFIED):
			permanence.status=PERMANENCE_OPEN
			site_producers_in_this_permanence = SiteProducer.objects.filter(
				permanence=permanence).active()
			for product in Product.objects.filter(
				site_producer__in = site_producers_in_this_permanence
				).active().is_selected_for_offer():
				offer_item_set= list(OfferItem.objects.all().product(
					product).permanence(permanence).order_by()[:1])
				if offer_item_set:
					offer_item = offer_item_set[0]
					offeritem.is_active=True
					offeritem.producer_unit_price = product.producer_unit_price
					offeritem.save()
				else:
					OfferItem.objects.create(
						permanence = permanence,
						product = product,
						producer_unit_price = product.producer_unit_price)
			permanence.save()

	open_orders.short_description = _('open orders')

	def close_orders(self, request, queryset):
		permanence_set = queryset.filter(status=PERMANENCE_OPEN)
		if permanence_set:
			for permanence in permanence_set:
				Purchase.objects.all(
					).permanence(permanence).update(
						validated_quantity=F('order_quantity')
					)
				permanence.status = PERMANENCE_CLOSED
				permanence.save()

	close_orders.short_description = _('close orders')

	def send_orders_to_producers(self, request, queryset):
		queryset.filter(status=PERMANENCE_CLOSED).update(status=PERMANENCE_SEND)

	send_orders_to_producers.short_description = _('send orders to producers')

	def back_to_previous_status(self, request, queryset):
		for permanence in queryset:
			if permanence.status==PERMANENCE_PLANIFIED:
				permanence.status=PERMANENCE_DISABLED
				permanence.save()
			if permanence.status==PERMANENCE_OPEN:
				for offeritem in OfferItem.objects.filter(
					permanence=permanence):
					offeritem.is_active=False
					offeritem.save()
				permanence.status=PERMANENCE_PLANIFIED
				permanence.save()
			if permanence.status==PERMANENCE_CLOSED:
				permanence.status=PERMANENCE_OPEN
				permanence.save()
			if permanence.status==PERMANENCE_SEND:
				permanence.status=PERMANENCE_CLOSED
				permanence.save()

	back_to_previous_status.short_description = _('back to previous status')

	def get_actions(self, request):
		actions = super(PermanenceInPreparationAdmin, self).get_actions(request)
		if 'delete_selected' in actions:
			del actions['delete_selected']
		if not actions:
			try:
				self.list_display.remove('action_checkbox')
			except ValueError:
				pass
		return actions

	def queryset(self, request):
		qs = super(PermanenceInPreparationAdmin, self).queryset(request)
		# return qs.filter(Q(site=settings.SITE_ID),
		# 	Q(status = '010') |  Q(status = '020'))
		return qs.filter(site=settings.SITE_ID,
			status__lte=PERMANENCE_SEND)
admin.site.register(PermanenceInPreparation, PermanenceInPreparationAdmin)

class PermanenceDoneAdmin(admin.ModelAdmin):
	fieldsets = [
		('Permanance', 
			{'fields': ['distribution_date', 'short_name', 'status', 
			'memo']}
		),
	]
	readonly_fields = ('status', 'is_created_on', 'is_updated_on')
	exclude = ['site']
	list_max_show_all = True
	# inlines = [PermanenceBoardInline, DeliveryBoardInline]
	inlines = [PermanenceBoardInline, OfferItemInline]
	date_hierarchy = 'distribution_date'
	list_display = ('__unicode__', 'get_siteproducers', 'get_sitecustomers', 'get_board', 'status')
	ordering = ('distribution_date',)
	actions = ['orders_prepared', 'done','back_to_previous_status']

	# def get_urls(self):
	#         urls = super(PermanenceDoneAdmin, self).get_urls()
	#         my_urls = patterns('',
	#             (r'\d+/purchase/$', self.admin_site.admin_view(self.purchase)),
	#         )
	#         return my_urls + urls

	# def purchase(self, request, id):
	# 	print(id)
	# 	print(request.GET['site_customer'])

	def orders_prepared(self, request, queryset):
		permanence_set = queryset.filter(status=PERMANENCE_SEND)
		if permanence_set:
			for permanence in permanence_set:
				purchase_set = Purchase.objects.all().permanence(permanence)
				if purchase_set:
					for purchase in purchase_set:
						offer_item_set = list(OfferItem.objects.all().product(
							purchase.product).permanence(permanence)[:1])
						if offer_item_set:
							offer_item = offer_item_set[0]
							purchase.preparator_recorded_quantity = purchase.validated_quantity
							purchase.effective_balance = purchase.validated_quantity * offer_item.producer_unit_price
							purchase.save()
						else:
							product_set = list(Product.objects.id(purchase.product_id)[:1])
							if product_set:
								product = product_set[0]
								purchase.preparator_recorded_quantity = purchase.validated_quantity
								purchase.effective_balance = purchase.validated_quantity * product.producer_unit_price
								purchase.save()
				permanence.status = PERMANENCE_PREPARED
				permanence.save()

	orders_prepared.short_description = _('orders prepared')

	def done(self, request, queryset):
		permanence_set = queryset.filter(status=PERMANENCE_PREPARED)
		if permanence_set:
			for permanence in permanence_set:
				purchase_set = Purchase.objects.all().permanence(permanence)
				if purchase_set:
					for purchase in purchase_set:
						if not(purchase.is_recorded_on_site_customer):
							bank_account_set = BankAccount.objects.filter(
								site_customer=purchase.site_customer, 
								is_recorded_on_site_customer=False)
							if bank_account_set:
								for bank_account in bank_account_set:
									if bank_account.bank_amount_in:
										purchase.site_customer.amount_in += bank_account.bank_amount_in
									if bank_account.bank_amount_out:
										purchase.site_customer.amount_out += bank_account.bank_amount_out
									bank_account.is_recorded_on_site_customer=True
									bank_account.save()
							purchase.site_customer.amount_out += purchase.effective_balance
							purchase.site_customer.save()
							purchase.is_recorded_on_site_customer = True
						if not(purchase.is_recorded_on_site_producer):
							bank_account_set = BankAccount.objects.filter(
								site_producer=purchase.site_producer, 
								is_recorded_on_site_producer=False)
							if bank_account_set:
								for bank_account in bank_account_set:
									if bank_account.bank_amount_in:
										purchase.site_producer.amount_in += bank_account.bank_amount_in
									if bank_account.bank_amount_out:
										purchase.site_producer.amount_out += bank_account.bank_amount_out
									bank_account.is_recorded_on_site_producer=True
									bank_account.save()
							purchase.site_producer.amount_in += purchase.effective_balance
							purchase.site_producer.save()
							purchase.is_recorded_on_site_producer = True
						purchase.save()
				permanence.status = PERMANENCE_DONE
				permanence.save()

	done.short_description = _('done')

	def back_to_previous_status(self, request, queryset):
		for permanence in queryset:
			if permanence.status==PERMANENCE_PREPARED:
				permanence.status=PERMANENCE_SEND
				permanence.save()
			if permanence.status==PERMANENCE_DONE:
				permanence.status=PERMANENCE_PREPARED
				permanence.save()

	back_to_previous_status.short_description = _('back to previous status')

	def get_actions(self, request):
		actions = super(PermanenceDoneAdmin, self).get_actions(request)
		if 'delete_selected' in actions:
			del actions['delete_selected']
		if not actions:
			try:
				self.list_display.remove('action_checkbox')
			except ValueError:
				pass
		return actions

	def queryset(self, request):
		qs = super(PermanenceDoneAdmin, self).queryset(request)
		# return qs.filter(Q(site=settings.SITE_ID),
		# 	Q(status = '010') |  Q(status = '020'))
		return qs.filter(site=settings.SITE_ID,
			status__gte=PERMANENCE_SEND)
admin.site.register(PermanenceDone, PermanenceDoneAdmin)

class PurchaseAdmin(admin.ModelAdmin):
	list_max_show_all = True
	exclude = ['site', 
		'is_recorded_on_previous_site_customer',
		'is_recorded_on_previous_siteproducer'
	]	
	list_display = ['distribution_date', 'product','site_customer', 
		'order_quantity', 'validated_quantity', 'preparator_recorded_quantity',
		'effective_balance']
	date_hierarchy = 'distribution_date'
	list_filter = ('distribution_date',)
	ordering = ('product', 'site_customer')
	fields = ('permanence',
		'site_customer',
		'product',
		'order_quantity',
		'validated_quantity',
		'preparator_recorded_quantity',
		'comment',
		'effective_balance',
		'is_recorded_on_site_customer',
		'is_recorded_on_site_producer')
	actions = []

	def get_readonly_fields(self, request, obj=None):
		if obj:
			status = obj.permanence.status
			if status<PERMANENCE_CLOSED:
				return('validated_quantity',
					'preparator_recorded_quantity',
					'comment',
					'effective_balance',
					'is_recorded_on_site_customer',
					'is_recorded_on_site_producer')
			if PERMANENCE_CLOSED<=status<PERMANENCE_SEND:
				return('order_quantity',
					'preparator_recorded_quantity',
					'comment',
					'effective_balance',
					'is_recorded_on_site_customer',
					'is_recorded_on_site_producer')
			if PERMANENCE_SEND<=status<PERMANENCE_PREPARED:
				return('order_quantity',
					'validated_quantity',
					'effective_balance',
					'is_recorded_on_site_customer',
					'is_recorded_on_site_producer')
			if PERMANENCE_PREPARED<=status<PERMANENCE_DONE:
				return('order_quantity',
					'validated_quantity',
					'preparator_recorded_quantity',
					'is_recorded_on_site_customer',
					'is_recorded_on_site_producer')
			return ('order_quantity',
				'validated_quantity',
				'preparator_recorded_quantity',
				'comment',
				'effective_balance',
				'is_recorded_on_site_customer',
				'is_recorded_on_site_producer')
		else:
			return ('validated_quantity',
				'preparator_recorded_quantity',
				'comment',
				'effective_balance',
				'is_recorded_on_site_customer',
				'is_recorded_on_site_producer')

	def get_form(self,request, obj=None, **kwargs):
		form = super(PurchaseAdmin,self).get_form(request, obj, **kwargs)
		permanence = form.base_fields["permanence"]
		site_customer = form.base_fields["site_customer"]
		product = form.base_fields["product"]
		permanence.widget.can_add_related = False
		site_customer.widget.can_add_related = False
		product.widget.can_add_related = False

		if obj:
			permanence.empty_label = None
			permanence.queryset = Permanence.objects.filter(
				id = obj.permanence_id)
			site_customer.empty_label = None
			site_customer.queryset = SiteCustomer.objects.filter(
				id = obj.site_customer_id)
			product.empty_label = None
			product.queryset = Product.objects.filter(
				id = obj.product_id)
		else:
			permanence.queryset = Permanence.objects.all().is_open()
			site_customer.queryset = SiteCustomer.objects.all().active()
			product.queryset = Product.objects.all().active()
		return form

	def save_model(self, request, purchase, form, change):
		# obj.preformed_by = request.user
		# obj.ip_address = utils.get_client_ip(request)
		purchase.site_producer = purchase.product.site_producer
		purchase.distribution_date = purchase.permanence.distribution_date
		result_set = list(OfferItem.objects.all().permanence(
			purchase.permanence).product(
			purchase.product).order_by()[:1])
		if not(result_set):
			# Add the product to the OfferItem 
			# to be managed with the other OfferItem's
			OfferItem.objects.create(
				permanence = purchase.permanence,
				product = purchase.product,
				producer_unit_price = purchase.product.producer_unit_price)
		purchase.save()

	def get_actions(self, request):
		actions = super(PurchaseAdmin, self).get_actions(request)
		if 'delete_selected' in actions:
			del actions['delete_selected']
		if not actions:
			try:
				self.list_display.remove('action_checkbox')
			except ValueError:
				pass
		return actions

admin.site.register(Purchase, PurchaseAdmin)

# Accounting
class BankAccountAdmin(admin.ModelAdmin):
	list_max_show_all = True
	exclude = ('site',)	
	list_display = ['operation_date', 'site_producer' , 'site_customer',
	 'bank_amount_in', 'bank_amount_out'] 
	date_hierarchy = 'operation_date'
	ordering = ('operation_date',)
	fields=('operation_date', 
		('site_producer', 'site_customer'), 'operation_comment', 'bank_amount_in',
		 'bank_amount_out', 
		 ('is_recorded_on_site_customer', 'is_recorded_on_site_producer'), 
		 ('is_recorded_on_previous_site_customer', 'is_recorded_on_previous_site_producer'),
		 ('is_created_on', 'is_updated_on') )
	actions = []

	def get_readonly_fields(self, request, obj=None):
		if obj:
			readonly = ['operation_date',
				'bank_amount_in',
				'bank_amount_out',
				'is_created_on', 'is_updated_on',
				'is_recorded_on_site_customer',
				'is_recorded_on_site_producer',
				'is_recorded_on_previous_site_customer',
				'is_recorded_on_previous_site_producer']
			if obj.site_customer==None:
				readonly.append('site_customer')
			if obj.site_producer==None:
				readonly.append('site_producer')
			return readonly
		return ('is_created_on', 'is_updated_on',
			'is_recorded_on_site_customer',
			'is_recorded_on_site_producer',
			'is_recorded_on_previous_site_customer',
			'is_recorded_on_previous_site_producer')

	def get_form(self,request, obj=None, **kwargs):
		form = super(BankAccountAdmin,self).get_form(request, obj, **kwargs)
		if obj:
			if obj.site_customer:
				site_customer = form.base_fields["site_customer"]
				site_customer.widget.can_add_related = False
				site_customer.empty_label = None
				site_customer.queryset = SiteCustomer.objects.id(
					obj.site_customer_id)
			if obj.site_producer:
				site_producer = form.base_fields["site_producer"]
				site_producer.widget.can_add_related = False
				site_producer.empty_label = None
				site_producer.queryset = SiteProducer.objects.id(
					obj.site_producer_id)
		else:
			site_producer = form.base_fields["site_producer"]
			site_customer = form.base_fields["site_customer"]
			site_producer.widget.can_add_related = False
			site_customer.widget.can_add_related = False
			site_producer.queryset = SiteProducer.objects.all(
				).not_the_buyinggroup().active().order_by(
				"short_profile_name")
			site_customer.queryset = SiteCustomer.objects.all(
				).not_the_buyinggroup().active().order_by(
				"short_basket_name")
		return form

	def get_actions(self, request):
		actions = super(BankAccountAdmin, self).get_actions(request)
		if 'delete_selected' in actions:
			del actions['delete_selected']
		if not actions:
			try:
				self.list_display.remove('action_checkbox')
			except ValueError:
				pass
		return actions

admin.site.register(BankAccount, BankAccountAdmin)