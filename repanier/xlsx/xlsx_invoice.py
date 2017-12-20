# -*- coding: utf-8

from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from django.utils.translation import ugettext_lazy as _
from openpyxl import load_workbook

import repanier.apps
from repanier.xlsx.export_tools import *
from repanier.const import *
from repanier.models import Configuration
from repanier.models import CustomerInvoice
from repanier.models.bankaccount import BankAccount
from repanier.models.customer import Customer
from repanier.models.invoice import ProducerInvoice
from repanier.models.permanence import Permanence
from repanier.models.producer import Producer
from repanier.models.product import Product
from repanier.models.purchase import Purchase
from repanier.tools import get_invoice_unit, get_reverse_invoice_unit, \
    create_or_update_one_purchase, reorder_offer_items, reorder_purchases
from repanier.xlsx.import_tools import get_customer_email_2_id_dict, \
    get_header, get_row


def export_bank(permanence, wb=None, sheet_name=EMPTY_STRING):
    # Detail of bank movements for a permanence
    wb, ws = new_landscape_a4_sheet(wb, "{} {}".format(_('Dashboard'), sheet_name),
                                    permanence)

    row_num = 0

    bank_account = BankAccount.objects.filter(
        permanence_id=permanence.id,
        customer__isnull=True,
        producer__isnull=True
    ).order_by('?').first()
    if bank_account is None:
        # Permanence not invoiced yet : Nothing to do
        return wb
    customer_set = Customer.objects.filter(
        Q(
            customerinvoice__isnull=False, is_anonymized=False
        ) | Q(
            customerinvoice__permanence_id=permanence.id
        )
    ).distinct()
    for customer in customer_set:
        bank_amount_in = bank_amount_out = DECIMAL_ZERO
        prepared = DECIMAL_ZERO
        customer_invoice = CustomerInvoice.objects.filter(
            customer_id=customer.id,
            permanence_id=permanence.id
        ).order_by('?').first()
        if customer_invoice is not None:
            balance_before = customer_invoice.previous_balance.amount
            bank_amount_in = customer_invoice.bank_amount_in.amount
            bank_amount_out = customer_invoice.bank_amount_out.amount
            if customer_invoice.customer_id == customer_invoice.customer_charged_id:
                prepared = customer_invoice.get_total_price_with_tax().amount
            balance_after = customer_invoice.balance.amount
        else:
            last_customer_invoice = CustomerInvoice.objects.filter(
                customer_id=customer.id,
                invoice_sort_order__lte=bank_account.id
            ).order_by('-invoice_sort_order').first()
            if last_customer_invoice is not None:
                balance_before = last_customer_invoice.balance.amount
                balance_after = last_customer_invoice.balance.amount
            else:
                # No invoice yet.
                balance_before = customer.initial_balance.amount
                balance_after = customer.initial_balance.amount

        row = [
            (_('Name'), 40, customer.long_basket_name, NumberFormat.FORMAT_TEXT),
            (_('Previous balance'), 15, balance_before, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Cash in'), 10, bank_amount_in, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Cash out'), 10, bank_amount_out, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Prepared'), 10, prepared, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Final balance'), 15, balance_after, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Name'), 20, customer.short_basket_name, NumberFormat.FORMAT_TEXT),
        ]

        if row_num == 0:
            worksheet_set_header(ws, row)
            row_num += 1

        for col_num in range(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.value = row[col_num][ROW_VALUE]
            c.style.number_format.format_code = row[col_num][ROW_FORMAT]

        row_num += 1

    row_break = row_num
    row_num += 1

    producer_set = Producer.objects.filter(producerinvoice__isnull=False).distinct()
    for producer in producer_set:
        bank_amount_in = bank_amount_out = DECIMAL_ZERO
        prepared = DECIMAL_ZERO
        producer_invoice = ProducerInvoice.objects.filter(
            producer_id=producer.id,
            permanence_id=permanence.id,
        ).order_by('?').first()
        if producer_invoice is not None:
            balance_before = -producer_invoice.previous_balance.amount
            bank_amount_in = producer_invoice.bank_amount_in.amount
            bank_amount_out = producer_invoice.bank_amount_out.amount
            prepared = producer_invoice.get_total_price_with_tax().amount
            balance_after = -producer_invoice.balance.amount
        else:
            last_producer_invoice = ProducerInvoice.objects.filter(
                producer_id=producer.id,
                invoice_sort_order__lte=bank_account.id
            ).order_by('-invoice_sort_order').first()
            if last_producer_invoice is not None:
                balance_before = -last_producer_invoice.balance.amount
                balance_after = -last_producer_invoice.balance.amount
            else:
                # No invoice yet.
                balance_before = -producer.initial_balance.amount
                balance_after = -producer.initial_balance.amount

        row = [
            (_('Name'), 40, producer.long_profile_name, NumberFormat.FORMAT_TEXT),
            (_('Previous balance'), 15, balance_before, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Cash in'), 10, bank_amount_in, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Cash out'), 10, bank_amount_out, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Prepared'), 10, prepared, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Final balance'), 15, balance_after, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX),
            (_('Name'), 20, producer.short_profile_name, NumberFormat.FORMAT_TEXT),
        ]

        if row_num == 0:
            worksheet_set_header(ws, row)
            row_num += 1

        for col_num in range(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.value = row[col_num][ROW_VALUE]
            c.style.number_format.format_code = row[col_num][ROW_FORMAT]

        row_num += 1

    final_bank_amount = bank_account.bank_amount_in.amount - bank_account.bank_amount_out.amount
    bank_account = BankAccount.objects.filter(
        id__lt=bank_account.id,
        customer__isnull=True,
        producer__isnull=True
    ).order_by("-id").first()
    if bank_account is not None:
        initial_bank_amount = bank_account.bank_amount_in.amount - bank_account.bank_amount_out.amount
    else:
        # This shouldn't occur because an initial balance is automatically generated
        # if not present
        # when invoicing the very first permanence.
        initial_bank_amount = DECIMAL_ZERO
    row_num += 1
    c = ws.cell(row=row_num, column=1)
    c.value = initial_bank_amount
    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
    c = ws.cell(row=row_num, column=4)
    formula = "B{}+SUM(C{}:C{})-SUM(D{}:D{})".format(row_num + 1, 2, row_num - 1, 2, row_num - 1)
    c.value = '=' + formula
    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX

    row_num += 1
    c = ws.cell(row=row_num, column=4)
    formula = "SUM(F{}:F{})-SUM(F{}:F{})".format(2, row_break, row_break + 2, row_num - 2)
    c.value = '=' + formula
    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX

    row_num += 1
    c = ws.cell(row=row_num, column=4)
    c.value = final_bank_amount
    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX

    return wb


def export_invoice(permanence=None, year=None, customer=None, producer=None, wb=None, sheet_name=EMPTY_STRING):
    # Detail of what has been prepared
    from repanier.apps import REPANIER_SETTINGS_CONFIG

    hide_producer_prices = False
    hide_customer_prices = False
    purchase_set = Purchase.objects.all()
    if producer is not None:
        purchase_set = purchase_set.filter(producer_id=producer.id)
        hide_customer_prices = True
    if permanence is not None:
        purchase_set = purchase_set.filter(permanence_id=permanence.id)
    if year is not None:
        purchase_set = purchase_set.filter(permanence__permanence_date__year=year)
    if customer is not None:
        purchase_set = purchase_set.filter(customer_invoice__customer_charged_id=customer.id)
        hide_producer_prices = True
    if purchase_set.exists():

        wb, ws = new_landscape_a4_sheet(wb, sheet_name, permanence)
        row = []
        row_num = 0
        hide_column_deposit = True

        for purchase in purchase_set:

            qty = purchase.quantity_invoiced

            if purchase.offer_item.unit_deposit.amount != DECIMAL_ZERO:
                hide_column_deposit = False

            unit = get_invoice_unit(order_unit=purchase.offer_item.order_unit, qty=qty)
            row = [
                (_("Permanence"), 20, purchase.permanence, NumberFormat.FORMAT_TEXT, False),
                (_("Producer"), 15, purchase.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
                (_("Basket"), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT, False),
                (_("Department"), 15,
                 purchase.offer_item.department_for_customer.short_name if purchase.offer_item.department_for_customer is not None else EMPTY_STRING,
                 NumberFormat.FORMAT_TEXT, False),
                (_("Product"), 60, purchase.get_long_name(), NumberFormat.FORMAT_TEXT, False),
                (_("Quantity"), 10, qty, '#,##0.????',
                 True if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG else False),
                (_("Unit"), 10, unit, NumberFormat.FORMAT_TEXT, False),
                (_("Deposit"), 10, purchase.offer_item.unit_deposit.amount, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX,
                 False)]
            if hide_producer_prices:
                row += [
                    (_("Empty"), 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                    (_("Empty"), 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                    (_("Empty"), 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False)
                ]
            else:
                row += [
                    (_("Producer unit price"), 10, purchase.get_producer_unit_price(),
                     repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
                    (_("Purchase price"), 10, purchase.purchase_price.amount, repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX,
                     False),
                    (_("VAT"), 10, purchase.producer_vat.amount,
                     repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False)
                ]

            if hide_customer_prices:
                row += [
                    (_("Empty"), 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                    (_("Empty"), 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False),
                    (_("Empty"), 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False)
                ]
            else:
                row += [
                    (_("Customer unit price"), 10, purchase.get_customer_unit_price(),
                     repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
                    (_("Selling price"), 10, purchase.selling_price.amount,
                     repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
                    (_("VAT"), 10, purchase.customer_vat.amount,
                     repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX, False),
                ]
            if hide_producer_prices and hide_customer_prices:
                row += [
                    (_("Empty"), 10, EMPTY_STRING, NumberFormat.FORMAT_TEXT, False)
                ]
            else:
                row += [
                    (_("VAT level"), 10, purchase.offer_item.get_vat_level_display(), NumberFormat.FORMAT_TEXT, False)
                ]
            row += [
                (_("Comment"), 30, EMPTY_STRING if purchase.comment is None else purchase.comment, NumberFormat.FORMAT_TEXT,
                 False),
                (_("Invoice status"), 10, purchase.get_status_display(), NumberFormat.FORMAT_TEXT, False),
                (_("Customer"), 10, purchase.customer.user.email, NumberFormat.FORMAT_TEXT, False),
                (_("Reference"), 10, purchase.offer_item.reference, NumberFormat.FORMAT_TEXT, False),
                (_("Wrapped"), 10, purchase.offer_item.wrapped, NumberFormat.FORMAT_TEXT, False),
            ]

            if row_num == 0:
                worksheet_set_header(ws, row)
                row_num += 1

            for col_num in range(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = "{}".format(row[col_num][ROW_VALUE])
                c.style.number_format.format_code = row[col_num][ROW_FORMAT]
                if row[col_num][ROW_BOX]:
                    c.style.borders.top.border_style = Border.BORDER_THIN
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c.style.borders.left.border_style = Border.BORDER_THIN
                    c.style.borders.right.border_style = Border.BORDER_THIN
                else:
                    c.style.borders.bottom.border_style = Border.BORDER_HAIR
                if col_num == 8:
                    c.style.font.bold = True

            row_num += 1

        if ws is not None:
            if hide_column_deposit:
                ws.column_dimensions[get_column_letter(8)].visible = False
            if hide_producer_prices:
                ws.column_dimensions[get_column_letter(9)].visible = False
                ws.column_dimensions[get_column_letter(10)].visible = False
                ws.column_dimensions[get_column_letter(11)].visible = False
            if hide_customer_prices:
                ws.column_dimensions[get_column_letter(12)].visible = False
                ws.column_dimensions[get_column_letter(13)].visible = False
                ws.column_dimensions[get_column_letter(14)].visible = False
            if hide_producer_prices and hide_customer_prices:
                ws.column_dimensions[get_column_letter(15)].visible = False
            ws.column_dimensions[get_column_letter(18)].visible = False
            ws.column_dimensions[get_column_letter(19)].visible = False
            ws.column_dimensions[get_column_letter(20)].visible = False
            for col_num in range(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.top.border_style = Border.BORDER_THIN
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if col_num == 2:
                    c.value = "{} {}".format(_("Total Price"), sheet_name)
                    c.style.font.bold = True
                if col_num == 9:
                    formula = "SUM(J{}:J{})".format(2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                if col_num == 10:
                    formula = "SUM(K{}:K{})".format(2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                if col_num == 12:
                    formula = "SUM(M{}:M{})".format(2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                if col_num == 13:
                    formula = "SUM(N{}:N{})".format(2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
            if customer is not None:
                config = REPANIER_SETTINGS_CONFIG
                group_label = config.group_label
                if group_label:
                    row_num += 1
                    for col_num in range(len(row)):
                        c = ws.cell(row=row_num, column=col_num)
                        c.style.borders.top.border_style = Border.BORDER_THIN
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        if col_num == 2:
                            c.value = group_label
                            c.style.font.bold = True
    return wb


@transaction.atomic
def import_invoice_sheet(worksheet, invoice_reference=None,
                          customer_2_id_dict=None,
                          producer=None
                          ):
    error = False
    error_msg = None
    header = get_header(worksheet)
    if header:
        now = timezone.now().date()
        lut_reverse_vat = dict(LUT_ALL_VAT_REVERSE)
        import_counter = 0
        row_num = 1
        sid = transaction.savepoint()
        try:

            permanence = Permanence.objects.create(
                permanence_date=now,
                short_name=invoice_reference,
                status=PERMANENCE_SEND,
                highest_status=PERMANENCE_SEND
            )
            permanence.producers.add(producer)
            row = get_row(worksheet, header, row_num)
            while row and not error:
                customer_name = row[_("Customer")]
                if customer_name:
                    if customer_name in customer_2_id_dict:
                        customer_id = customer_2_id_dict[customer_name]
                    else:
                        error = True
                        error_msg = _("Row %(row_num)d : No valid customer") % {'row_num': row_num + 1}
                        break
                    product_reference = row[_("Reference")]
                    unit = row[_("Unit")]
                    order_unit = get_reverse_invoice_unit(unit)
                    vat = row[_("VAT level")]
                    vat_level = lut_reverse_vat[vat]
                    product = Product.objects.filter(producer_id=producer.id, reference=product_reference).order_by('?').first()
                    if product is None:
                        product = Product.objects.create(
                            producer=producer,
                            reference=product_reference,
                        )
                    product.long_name = row[_("Product")]
                    # The producer unit price is the imported customer unit price
                    # If the group get a reduction, this one must be mentionned into the producer admin screen
                    # into the "price_list_multiplier" field
                    product.producer_unit_price = row[_("Customer unit price")]
                    product.unit_deposit = row[_("Deposit")]
                    product.order_unit = order_unit
                    product.vat_level = vat_level
                    product.wrapped = row[_("Wrapped")]
                    product.save()
                    qty_and_price_display = product.get_qty_and_price_display(customer_price=False)
                    if product.long_name.endswith(qty_and_price_display):
                        product.long_name = product.long_name[:-len(qty_and_price_display)]
                        product.save()
                    offer_item = product.get_or_create_offer_item(permanence, reset_add_2_stock=True)
                    create_or_update_one_purchase(
                        customer_id,
                        offer_item,
                        q_order=Decimal(row[_("Quantity")]),
                        status=PERMANENCE_SEND,
                        batch_job=True, is_box_content=False,
                        comment=row[_("Comment")]
                    )
                    import_counter += 1

                row_num += 1
                row = get_row(worksheet, header, row_num)
            reorder_offer_items(permanence.id)
            reorder_purchases(permanence.id)

        except KeyError as e:
            # Missing field
            error = True
            error_msg = _("Row %(row_num)d : A required column is missing %(error_msg)s.") % {
                'row_num': row_num + 1, 'error_msg': str(e)}
        except Exception as e:
            error = True
            error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
        if not error and import_counter == 0:
            error = True
            error_msg = "{}".format(_("Nothing to import."))
        if error:
            transaction.savepoint_rollback(sid)
        else:
            transaction.savepoint_commit(sid)
    return error, error_msg


def handle_uploaded_invoice(request, permanences, file_to_import, producer, invoice_reference):
    if producer is None:
        error = True
        error_msg = _("A producer must be given.")
    else:
        wb = load_workbook(file_to_import)
        # dict for performance optimisation purpose : read the DB only once
        customer_2_id_dict = get_customer_email_2_id_dict()
        ws = wb.worksheets[0]
        error, error_msg = import_invoice_sheet(
            ws, invoice_reference=invoice_reference,
            customer_2_id_dict=customer_2_id_dict,
            producer=producer
        )
    return error, error_msg
