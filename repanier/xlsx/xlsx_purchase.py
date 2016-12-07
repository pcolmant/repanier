# -*- coding: utf-8
from __future__ import unicode_literals

from django.db import transaction
from django.http import HttpResponse
from django.utils import translation
from django.utils.text import slugify
from django.utils.translation import ugettext_lazy as _
from openpyxl import load_workbook
from openpyxl.style import Fill
from openpyxl.styles import Color

from export_tools import *
from import_tools import *
from repanier.apps import REPANIER_SETTINGS_CURRENCY_XLSX
from repanier.const import *
from repanier.models import OfferItem
from repanier.models import Producer
from repanier.models import Purchase
from repanier.tools import cap, next_row, recalculate_prices, recalculate_order_amount
from views import import_xslx_view


def next_purchase(purchases):
    purchase = next_row(purchases)
    while purchase is not None \
            and purchase.quantity_invoiced <= DECIMAL_ZERO \
            and purchase.offer_item.order_unit < PRODUCT_ORDER_UNIT_DEPOSIT:
        purchase = next_row(purchases)
    return purchase


def export_purchase(permanence=None, year=None, producer=None, customer=None, wb=None):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Format"), 5),
        (_("Id"), 10),
        (_("Date"), 15),
        (_("producer"), 15),
        (_("product"), 60),
        (_("customer"), 15),
        (_("quantity invoiced"), 10),
        (_("producer unit price"), 10),
        (_("deposit"), 10),
        (_("purchase price"), 10),
        (_("tax"), 10),
        (_("rule of 3"), 10),
        (_("comment"), 30),
        (_("vat level"), 10),
    ]

    if producer is None:
        if permanence is not None:
            producers = Producer.objects.filter(
                producerinvoice__permanence_id=permanence.id
            ).distinct().iterator()
            title1 = "%s" % permanence
        else:
            producers = Producer.objects.filter(
                producerinvoice__permanence__permanence_date__year=year
            ).distinct().iterator()
            title1 = "%s-%d" % (customer.short_basket_name, year)
    else:
        producers = Producer.objects.filter(id=producer.id).iterator()
        title1 = "%s-%d" % (producer.short_profile_name, year)
    producer = next_row(producers)
    if producer is not None:
        wb, ws = new_landscape_a4_sheet(
            wb,
            title1,
            _('invoices'),
            header
        )
        row_num = 1
        count_all_purchase = 0
        purchase_price_all_purchase = []
        tax_all_purchase = []
        while producer is not None:
            producer_save = producer
            # count_producer_purchase = 0
            producer_price = DECIMAL_ZERO
            if producer.invoice_by_basket:
                if year is None:
                    purchases = Purchase.objects.filter(
                        permanence_id=permanence.id,
                        producer_id=producer.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).order_by(
                        "customer__short_basket_name",
                        "offer_item__translations__preparation_sort_order"
                    ).iterator()
                else:
                    if customer is None:
                        purchases = Purchase.objects.filter(
                            permanence__status__gte=PERMANENCE_DONE,
                            permanence_date__year=year,
                            producer_id=producer.id,
                            offer_item__translations__language_code=translation.get_language()
                        ).order_by(
                            "permanence_date",
                            "permanence_id",
                            "customer__short_basket_name",
                            "offer_item__translations__preparation_sort_order"
                        ).iterator()
                    else:
                        purchases = Purchase.objects.filter(
                            permanence__status__gte=PERMANENCE_DONE,
                            permanence_date__year=year,
                            customer_id=customer.id,
                            producer_id=producer.id,
                            offer_item__translations__language_code=translation.get_language()
                        ).order_by(
                            "permanence_date",
                            "permanence_id",
                            "offer_item__translations__preparation_sort_order"
                        ).iterator()
                purchase = next_purchase(purchases)
                while purchase is not None:
                    permanence_save = purchase.permanence
                    count_permanence_purchase = 0
                    row_start_permanence = 0
                    row_num += 1
                    while purchase is not None and permanence_save.id == purchase.permanence_id:
                        customer_save = purchase.customer
                        count_purchase = 0
                        row_start_purchase = 0
                        purchases_price = DECIMAL_ZERO
                        while purchase is not None and permanence_save.id == purchase.permanence_id \
                                and customer_save.id == purchase.customer_id:
                            offer_item_save = purchase.offer_item
                            department_for_customer_save = offer_item_save.department_for_customer
                            department_for_customer_save__short_name = department_for_customer_save.short_name \
                                if department_for_customer_save is not None else EMPTY_STRING

                            while purchase is not None and permanence_save.id == purchase.permanence_id \
                                    and customer_save.id == purchase.customer_id \
                                    and department_for_customer_save == purchase.offer_item.department_for_customer:
                                c = ws.cell(row=row_num, column=1)
                                c.value = purchase.id
                                c = ws.cell(row=row_num, column=2)
                                c.value = permanence_save.permanence_date
                                c.style.number_format.format_code = NumberFormat.FORMAT_DATE_DDMMYYYY
                                c = ws.cell(row=row_num, column=3)
                                c.value = producer_save.short_profile_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                if count_purchase == 0:
                                    row_start_purchase = row_num + 1
                                    if count_permanence_purchase == 0:
                                        c.style.font.bold = True
                                        row_start_permanence = row_start_purchase
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = "A"
                                else:
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = "B"
                                count_purchase += 1
                                c = ws.cell(row=row_num, column=4)
                                if department_for_customer_save__short_name is not None:
                                    c.value = "%s - %s" % (
                                        purchase.get_long_name(), department_for_customer_save__short_name)
                                else:
                                    c.value = "%s" % purchase.get_long_name()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=5)
                                c.value = customer_save.short_basket_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                if count_purchase == 0 and customer is None:
                                    c.style.font.bold = True
                                c = ws.cell(row=row_num, column=6)
                                c.value = purchase.quantity_invoiced
                                c.style.number_format.format_code = '#,##0.????'
                                if year is None:
                                    c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(7) + str(row_num + 1), 'notEqual',
                                        [str(purchase.quantity_invoiced)], True, wb,
                                        None, None, yellowFill
                                    )
                                c = ws.cell(row=row_num, column=7)
                                c.value = purchase.get_producer_unit_price()
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                c = ws.cell(row=row_num, column=8)
                                c.value = purchase.offer_item.unit_deposit.amount
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                c = ws.cell(row=row_num, column=9)
                                c.value = '=ROUND(G%s*(H%s+I%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                                if year is None:
                                    purchase_price = (purchase.quantity_invoiced *
                                                      (purchase.get_producer_unit_price() +
                                                       purchase.get_unit_deposit())).quantize(TWO_DECIMALS)
                                    purchases_price += purchase_price
                                    if offer_item_save.order_unit in [
                                        PRODUCT_ORDER_UNIT_KG, PRODUCT_ORDER_UNIT_PC_KG,
                                        PRODUCT_ORDER_UNIT_LT
                                    ]:
                                        c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(10) + str(row_num + 1), 'notEqual',
                                        [str(purchase_price)], True, wb,
                                        None, None, yellowFill
                                    )
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                c = ws.cell(row=row_num, column=10)
                                c.value = '=G%s*%s' % (row_num + 1, purchase.offer_item.customer_vat.amount)
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                c = ws.cell(row=row_num, column=12)
                                c.value = cap(purchase.comment, 100)
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=13)
                                c.value = purchase.offer_item.get_vat_level_display()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                delta = 5
                                for col_num in range(5):
                                    c = ws.cell(row=row_num, column=delta + col_num)
                                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                                row_num += 1
                                purchase = next_purchase(purchases)

                        count_permanence_purchase += count_purchase
                        if year is None and count_purchase > 1:
                            c = ws.cell(row=row_num, column=11)
                            c.value = '=SUM(J%s:J%s)' % (row_start_purchase, row_num)
                            c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.font.color = Color(Color.BLUE)
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(12) + str(row_num + 1), 'notEqual',
                                [str(purchases_price)], True, wb,
                                None, None, yellowFill
                            )
                            c = ws.cell(row=row_num, column=0)
                            c.value = "C"
                            row_num += 1

                        producer_price += purchases_price
                    if count_permanence_purchase > 0:
                        count_all_purchase += count_permanence_purchase
                        purchase_price_producer_purchase = 'ROUND(SUM(J%s:J%s),2)' % (row_start_permanence, row_num)
                        purchase_price_all_purchase.append(purchase_price_producer_purchase)
                        tax_producer_purchase = 'SUM(K%s:K%s)' % (row_start_permanence, row_num)
                        tax_all_purchase.append(tax_producer_purchase)

                        row_num += 1
                        c = ws.cell(row=row_num, column=8)
                        c.value = "%s : %s %s" % (_("Total"), producer_save.short_profile_name, permanence_save)
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.font.bold = True
                        c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
                        c = ws.cell(row=row_num, column=9)
                        c.value = '=%s' % purchase_price_producer_purchase
                        c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                        c.style.font.bold = True
                        if year is None:
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(10) + str(row_num + 1), 'notEqual',
                                [str(producer_price)], True, wb,
                                None, None, yellowFill
                            )
                        c = ws.cell(row=row_num, column=10)
                        c.value = '=%s' % tax_producer_purchase
                        c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                        row_num += 1
                        for col_num in range(14):
                            c = ws.cell(row=row_num, column=col_num)
                            c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
                        row_num += 1
            else:
                if year is None:
                    # Using quantity_for_preparation_sort_order the order is by customer__short_basket_name if the product
                    # is to be distributed by piece, otherwise by lower qty first.
                    purchases = Purchase.objects.filter(
                        permanence_id=permanence.id,
                        producer_id=producer.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).order_by(  # "product__placement",
                        "offer_item__translations__preparation_sort_order",
                        "quantity_for_preparation_sort_order",
                        "customer__short_basket_name"
                    ).iterator()
                else:
                    if customer is None:
                        purchases = Purchase.objects.filter(
                            permanence__status__gte=PERMANENCE_DONE,
                            permanence_date__year=year,
                            producer_id=producer.id,
                            offer_item__translations__language_code=translation.get_language()
                        ).order_by(
                            "permanence_date",
                            "permanence_id",
                            "offer_item__translations__preparation_sort_order",
                            "quantity_for_preparation_sort_order",
                            "customer__short_basket_name"
                        ).iterator()
                    else:
                        purchases = Purchase.objects.filter(
                            permanence__status__gte=PERMANENCE_DONE,
                            permanence_date__year=year,
                            customer_id=customer.id,
                            producer_id=producer.id,
                            offer_item__translations__language_code=translation.get_language()
                        ).order_by(
                            "permanence_date",
                            "permanence_id",
                            "offer_item__translations__preparation_sort_order",
                            "quantity_for_preparation_sort_order"
                        ).iterator()
                purchase = next_purchase(purchases)
                while purchase is not None:
                    permanence_save = purchase.permanence
                    count_permanence_purchase = 0
                    row_start_permanence = 0
                    row_num += 1
                    while purchase is not None and permanence_save.id == purchase.permanence_id:
                        producer_save = purchase.producer
                        department_for_customer_save = purchase.offer_item.department_for_customer
                        department_for_customer_save__short_name = department_for_customer_save.short_name \
                            if department_for_customer_save is not None else EMPTY_STRING
                        while purchase is not None and permanence_save.id == purchase.permanence_id \
                                and producer_save == purchase.producer \
                                and department_for_customer_save == purchase.offer_item.department_for_customer:
                            offer_item_save = purchase.offer_item
                            count_offer_item = 0
                            row_first_offer_item = row_num
                            offer_items_price = DECIMAL_ZERO
                            for col_num in range(14):
                                c = ws.cell(row=row_num, column=col_num)
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                            row_start_offer_item = 0
                            while purchase is not None and offer_item_save == purchase.offer_item:
                                c = ws.cell(row=row_num, column=1)
                                c.value = purchase.id
                                c = ws.cell(row=row_num, column=2)
                                c.value = permanence_save.permanence_date
                                c.style.number_format.format_code = NumberFormat.FORMAT_DATE_DDMMYYYY
                                c = ws.cell(row=row_num, column=3)
                                c.value = producer_save.short_profile_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                if count_offer_item == 0:
                                    row_start_offer_item = row_num + 1
                                    if count_permanence_purchase == 0:
                                        c.style.font.bold = True
                                        row_start_permanence = row_start_offer_item
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = "A"
                                else:
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = "B"
                                c = ws.cell(row=row_num, column=4)
                                if department_for_customer_save__short_name is not None:
                                    c.value = "%s - %s" % (
                                        purchase.get_long_name(), department_for_customer_save__short_name)
                                else:
                                    c.value = "%s" % purchase.get_long_name()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                if count_offer_item != 0:
                                    c.style.font.color.index = 'FF939393'
                                c = ws.cell(row=row_num, column=5)
                                c.value = purchase.customer.short_basket_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=6)
                                c.value = purchase.quantity_invoiced
                                c.style.number_format.format_code = '#,##0.????'
                                if year is None:
                                    c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(7) + str(row_num + 1), 'notEqual',
                                        [str(purchase.quantity_invoiced)], True, wb,
                                        None, None, yellowFill
                                    )
                                c = ws.cell(row=row_num, column=7)
                                if count_offer_item == 0:
                                    producer_unit_price = purchase.get_producer_unit_price()
                                    c.value = producer_unit_price
                                    c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(8) + str(row_num + 1), 'notEqual',
                                        [str(producer_unit_price)], True, wb,
                                        None, None, yellowFill
                                    )
                                else:
                                    c.value = '=H%s' % (row_first_offer_item + 1)
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                c = ws.cell(row=row_num, column=8)
                                c.value = purchase.offer_item.unit_deposit.amount
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                c = ws.cell(row=row_num, column=9)
                                c.value = '=ROUND(G%s*(H%s+I%s),2)' % (row_num + 1, row_first_offer_item + 1, row_num + 1)
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                if year is None:
                                    offer_item_price = (purchase.quantity_invoiced *
                                                        (purchase.get_producer_unit_price() +
                                                         purchase.get_unit_deposit())).quantize(
                                        TWO_DECIMALS)
                                    offer_items_price += offer_item_price
                                    if offer_item_save.order_unit in [
                                        PRODUCT_ORDER_UNIT_KG, PRODUCT_ORDER_UNIT_PC_KG,
                                        PRODUCT_ORDER_UNIT_LT
                                    ]:
                                        c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(10) + str(row_num + 1), 'notEqual',
                                        [str(offer_item_price)], True, wb,
                                        None, None, yellowFill
                                    )
                                c = ws.cell(row=row_num, column=10)
                                c.value = '=G%s*%s' % (row_num + 1, purchase.offer_item.customer_vat.amount)
                                c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                c = ws.cell(row=row_num, column=12)
                                c.value = cap(purchase.comment, 100)
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=13)
                                c.value = purchase.offer_item.get_vat_level_display()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT

                                delta = 5
                                for col_num in range(5):
                                    c = ws.cell(row=row_num, column=delta + col_num)
                                    c.style.borders.bottom.border_style = Border.BORDER_THIN

                                purchase = next_purchase(purchases)
                                row_num += 1
                                count_offer_item += 1

                            count_permanence_purchase += count_offer_item
                            if year is None and count_offer_item > 1:
                                if not offer_item_save.wrapped and offer_item_save.order_unit in [
                                    PRODUCT_ORDER_UNIT_KG, PRODUCT_ORDER_UNIT_PC_KG,
                                    PRODUCT_ORDER_UNIT_LT
                                ]:
                                    c = ws.cell(row=row_num, column=11)
                                    c.value = '=SUM(J%s:J%s)' % (row_start_offer_item, row_num)
                                    c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                                    c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(12) + str(row_num + 1), 'notEqual',
                                        [str(offer_items_price)], True, wb,
                                        None, None, yellowFill
                                    )
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = "D"

                            producer_price += offer_items_price

                    if count_permanence_purchase > 0:
                        count_all_purchase += count_permanence_purchase
                        purchase_price_producer_purchase = 'ROUND(SUM(J%s:J%s),2)' % (row_start_permanence, row_num)
                        purchase_price_all_purchase.append(purchase_price_producer_purchase)
                        tax_producer_purchase = 'SUM(K%s:K%s)' % (row_start_permanence, row_num)
                        tax_all_purchase.append(tax_producer_purchase)

                        row_num += 1
                        c = ws.cell(row=row_num, column=8)
                        c.value = "%s : %s %s" % (_("Total"), producer_save.short_profile_name, permanence_save)
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.font.bold = True
                        c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
                        c = ws.cell(row=row_num, column=9)
                        c.value = '=%s' % purchase_price_producer_purchase
                        c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                        c.style.font.bold = True
                        if year is None:
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(10) + str(row_num + 1), 'notEqual',
                                [str(producer_price)], True, wb,
                                None, None, yellowFill
                            )
                        c = ws.cell(row=row_num, column=10)
                        c.value = '=%s' % tax_producer_purchase
                        c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
                        row_num += 1
                        for col_num in range(14):
                            c = ws.cell(row=row_num, column=col_num)
                            c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
                        row_num += 1
            producer = next_row(producers)

        if count_all_purchase > 0:
            row_num += 1
            c = ws.cell(row=row_num, column=8)
            c.value = "%s : %s" % (_('Total'), title1)
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            c.style.font.bold = True
            c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
            c = ws.cell(row=row_num, column=9)
            c.value = '=%s' % "+".join(purchase_price_all_purchase)
            c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
            c.style.font.bold = True
            c = ws.cell(row=row_num, column=10)
            c.value = '=%s' % "+".join(tax_all_purchase)
            c.style.number_format.format_code = REPANIER_SETTINGS_CURRENCY_XLSX
            row_num += 1
            for col_num in range(14):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED

        if year is None:
            ws.column_dimensions[get_column_letter(3)].visible = False
            ws.column_dimensions[get_column_letter(11)].visible = False
        else:
            ws.column_dimensions[get_column_letter(12)].visible = False
        ws.column_dimensions[get_column_letter(1)].visible = False

    return wb


def admin_export_permanence_by_producer(request, permanence):
    wb = export_purchase(permanence=permanence, wb=None)
    if wb is not None:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = "attachment; filename={0}-{1}.xlsx".format(
            slugify(_("Invoices")),
            slugify(permanence)
        )
        wb.save(response)
        return response
    else:
        return None


def admin_export_year_by_producer(year, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename={0}-{1}.xlsx".format(
        slugify(_("Invoices")),
        str(year)
    )
    wb = None
    for producer in queryset.order_by('short_profile_name'):
        wb = export_purchase(producer=producer, year=year, wb=wb)
    if wb is not None:
        wb.save(response)
    return response


def admin_export_year_by_customer(year, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = "attachment; filename={0}-{1}.xlsx".format(
        slugify(_("Invoices")),
        str(year)
    )
    wb = None
    for customer in queryset:
        wb = export_purchase(customer=customer, year=year, wb=wb)
    if wb is not None:
        wb.save(response)
    return response


@transaction.atomic
def import_purchase_sheet(worksheet, permanence=None,
                          customer_2_id_dict=None,
                          producer_2_id_dict=None
                          ):
    error = False
    error_msg = None
    import_counter = 0
    header = get_header(worksheet)
    if header:
        row_num = 1
        array_purchase = []
        rule_of_3_source = DECIMAL_ZERO
        row = get_row(worksheet, header, row_num)
        while row and not error:
            try:
                row_format = row[_('Format')]
                if row_format in ["A", "B"]:
                    import_counter += 1
                    if row[_('Id')] is None:
                        error = True
                        error_msg = _("Row %(row_num)d : No purchase id given.") % {'row_num': row_num + 1}
                        break
                    row_id = Decimal(row[_('Id')])
                    purchase = Purchase.objects.filter(id=row_id).order_by('?').first()
                    if purchase is None:
                        error = True
                        error_msg = _("Row %(row_num)d : No purchase corresponding to the given purchase id.") % {
                            'row_num': row_num + 1}
                        break
                    if purchase.permanence_id != permanence.id:
                        error = True
                        error_msg = _("Row %(row_num)d : The given permanence doesn't own the given purchase id.") % {
                            'row_num': row_num + 1}
                        break
                    producer_id = None
                    if row[_('producer')] in producer_2_id_dict:
                        producer_id = producer_2_id_dict[row[_('producer')]]
                    if producer_id != purchase.producer_id:
                        error = True
                        error_msg = _("Row %(row_num)d : No valid producer.") % {'row_num': row_num + 1}
                        break
                    customer_name = "%s" % row[_('customer')]
                    if customer_name in customer_2_id_dict:
                        customer_id = customer_2_id_dict[customer_name]
                        if customer_id != purchase.customer_id:
                            error = True
                            error_msg = _("Row %(row_num)d : No valid customer") % {'row_num': row_num + 1}
                            break
                    comment = cap(row[_('comment')], 100)
                    quantity_invoiced = DECIMAL_ZERO if row[_('quantity invoiced')] is None \
                        else Decimal(row[_('quantity invoiced')]).quantize(FOUR_DECIMALS)
                    producer_row_price = row[_('purchase price')]
                    if producer_row_price is not None:
                        producer_row_price = Decimal(producer_row_price).quantize(TWO_DECIMALS)
                        if purchase.purchase_price.amount != producer_row_price:
                            purchase_price_modified = True
                            if purchase.offer_item.order_unit in [
                                PRODUCT_ORDER_UNIT_KG, PRODUCT_ORDER_UNIT_PC_KG,
                                PRODUCT_ORDER_UNIT_LT
                            ]:
                                producer_unit_price = (purchase.offer_item.producer_unit_price.amount +
                                                       purchase.offer_item.unit_deposit.amount).quantize(TWO_DECIMALS)
                                if producer_unit_price != DECIMAL_ZERO:
                                    purchase.quantity_invoiced = (producer_row_price /
                                                                  producer_unit_price).quantize(FOUR_DECIMALS)
                                else:
                                    purchase.quantity_invoiced = DECIMAL_ZERO
                        else:
                            purchase_price_modified = False
                    elif purchase.quantity_invoiced != quantity_invoiced:
                        purchase.quantity_invoiced = quantity_invoiced
                        purchase_price_modified = False
                    else:
                        purchase_price_modified = False
                    if row_format == "A":
                        array_purchase = []
                        rule_of_3_source = DECIMAL_ZERO
                        if not purchase_price_modified:
                            producer_unit_price = Decimal(row[_('producer unit price')]).quantize(TWO_DECIMALS)
                            previous_producer_unit_price = purchase.get_producer_unit_price()
                            if producer_unit_price != previous_producer_unit_price:
                                offer_item = OfferItem.objects.filter(id=purchase.offer_item_id).order_by('?').first()
                                offer_item.producer_unit_price = producer_unit_price
                                recalculate_prices(offer_item, offer_item.producer_price_are_wo_vat,
                                                   offer_item.is_resale_price_fixed,
                                                   offer_item.price_list_multiplier)
                                offer_item.save()
                                recalculate_order_amount(
                                    permanence_id=offer_item.permanence_id,
                                    offer_item_queryset=OfferItem.objects.filter(id=offer_item.id).order_by('?'),
                                    send_to_producer=False
                                )
                    purchase.comment = comment
                    purchase.save()
                    rule_of_3_source += purchase.purchase_price.amount
                    array_purchase.append(purchase)
                elif row_format in ["C", "D"]:
                    rule_of_3_target = row[_('rule of 3')]
                    if rule_of_3_target is not None:
                        rule_of_3_target = Decimal(rule_of_3_target).quantize(TWO_DECIMALS)
                        if rule_of_3_target != rule_of_3_source:
                            max_purchase_counter = len(array_purchase)
                            if max_purchase_counter <= 1:
                                error = True
                                error_msg = _("Row %(row_num)d : Rule of 3 target in wrong context.") % {
                                    'row_num': row_num + 1}
                                return error, error_msg
                            else:
                                if rule_of_3_source != DECIMAL_ZERO:
                                    ratio = rule_of_3_target / rule_of_3_source
                                else:
                                    if rule_of_3_target == DECIMAL_ZERO:
                                        ratio = DECIMAL_ZERO
                                    else:
                                        ratio = DECIMAL_ONE
                                # Rule of 3
                                if ratio != DECIMAL_ONE:
                                    adjusted_invoice = DECIMAL_ZERO
                                    for i, purchase in enumerate(array_purchase, start=1):
                                        producer_unit_price = (purchase.offer_item.producer_unit_price.amount +
                                                               purchase.offer_item.unit_deposit.amount
                                                               ).quantize(TWO_DECIMALS)
                                        if i == max_purchase_counter:
                                            delta = rule_of_3_target - adjusted_invoice
                                            if producer_unit_price != DECIMAL_ZERO:
                                                purchase.quantity_invoiced = (delta / producer_unit_price).quantize(
                                                    FOUR_DECIMALS)
                                            else:
                                                purchase.quantity_invoiced = DECIMAL_ZERO
                                        else:
                                            purchase.quantity_invoiced = (purchase.quantity_invoiced * ratio).quantize(
                                                FOUR_DECIMALS)
                                            adjusted_invoice += purchase.quantity_invoiced * producer_unit_price
                                        purchase.save()

                row_num += 1
                row = get_row(worksheet, header, row_num)

            except KeyError, e:
                # Missing field
                error = True
                error_msg = _("Row %(row_num)d : A required column is missing %(error_msg)s.") % {
                    'row_num': row_num + 1, 'error_msg': str(e)}
            except Exception, e:
                error = True
                error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
    if import_counter == 0:
        error = True
        error_msg = "%s" % _("Nothing to import.")
    return error, error_msg


def handle_uploaded_file(request, permanences, file_to_import):
    error = False
    error_msg = None
    wb = load_workbook(file_to_import)
    # dict for performance optimisation purpose : read the DB only once
    customer_buyinggroup_id, customer_2_id_dict = get_customer_2_id_dict()
    producer_buyinggroup_id, producer_2_id_dict = get_producer_2_id_dict()
    if customer_buyinggroup_id is None:
        error = True
        error_msg = _("At least one customer must represent the buying group.")
    else:
        if producer_buyinggroup_id is None:
            error = True
            error_msg = _("At least one producer must represent the buying group.")

    if not error:
        permanence = permanences.first()
        if permanence is not None:
            if permanence.status == PERMANENCE_SEND:
                ws = wb.get_sheet_by_name(cap("%s" % (permanence), 31))
                error, error_msg = import_purchase_sheet(
                    ws, permanence=permanence,
                    customer_2_id_dict=customer_2_id_dict,
                    producer_2_id_dict=producer_2_id_dict
                )
                if error:
                    error_msg = cap("%s" % (permanence), 31) + " > " + error_msg
            else:
                error = True
                error_msg = _("The permanence has already been invoiced.")
        else:
            error = True
            error_msg = _("The permanence doesn't exists.")
    return error, error_msg


def admin_import(admin_ui, admin, request, queryset, action):
    return import_xslx_view(admin_ui, admin, request, queryset[:1], _("Import orders prepared"), handle_uploaded_file,
                            action=action)
