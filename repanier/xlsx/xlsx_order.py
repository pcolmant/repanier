# -*- coding: utf-8

from django.utils import translation
from django.utils.translation import ugettext_lazy as _
from openpyxl.style import Fill
from openpyxl.styles import Color

import repanier.apps
from repanier.xlsx.export_tools import *
from repanier.const import *
from repanier.models.configuration import Configuration
from repanier.models.customer import Customer
from repanier.models.deliveryboard import DeliveryBoard
from repanier.models.invoice import CustomerInvoice, ProducerInvoice
from repanier.models.offeritem import OfferItemWoReceiver
from repanier.models.permanence import Permanence
from repanier.models.permanenceboard import PermanenceBoard
from repanier.models.producer import Producer
from repanier.models.purchase import Purchase
from repanier.models.staff import Staff
from repanier.tools import get_base_unit, next_row
from .xlsx_stock import export_permanence_stock


def next_purchase(purchases):
    purchase = next_row(purchases)
    return purchase


def export_abstract(permanence, deliveries_id=None, group=False, wb=None):
    if permanence is not None:
        from repanier.apps import REPANIER_SETTINGS_CUSTOMERS_MUST_CONFIRM_ORDERS
        row_num = 1
        # Customer info
        Customer.objects.all().update(preparation_order=0)
        preparation_order = 1
        if deliveries_id is not None:
            header = [
                (_('Delivery point'),20),
                (_('Family'), 35),
                (_('Phone1'), 15),
                (_('Phone2'), 15),
                (_('Total with vat'), 15),
                (_('Email'), 35),
            ]
            wb, ws = new_portrait_a4_sheet(wb, permanence, EMPTY_STRING, header=header, add_print_title=False)
            for delivery_ref, delivery in enumerate(DeliveryBoard.objects.filter(id__in=deliveries_id).order_by("id")):
                customer_set = Customer.objects.filter(
                    customerinvoice__permanence_id=permanence.id,
                    represent_this_buyinggroup=False,
                    customerinvoice__delivery_id=delivery.id
                )
                for customer in customer_set:
                    invoice = CustomerInvoice.objects.filter(
                        permanence=permanence, customer=customer
                    ).order_by('?').first()
                    if invoice is not None and invoice.has_purchase:
                        customer.preparation_order = preparation_order
                        # customer.save(update_fields=['preparation_order'])
                        # use vvvv because ^^^^^ will call "pre_save" function which reset valid_email to None
                        Customer.objects.filter(id=customer.id).order_by('?').update(
                            preparation_order=preparation_order
                        )
                        preparation_order += 1
                        if REPANIER_SETTINGS_CUSTOMERS_MUST_CONFIRM_ORDERS and not invoice.is_order_confirm_send:
                            confirmed = "\n{}".format(_("/!\ This order isn't confirmed"))
                        else:
                            confirmed = EMPTY_STRING
                        row = [
                            "{} - {}".format(delivery_ref, delivery.get_delivery_display()),
                            "  {} - {}{}".format(customer.preparation_order, customer.long_basket_name, confirmed),
                            customer.phone1,
                            customer.phone2,
                            invoice.total_price_with_tax.amount,
                            # Used to send mail to customer with an order (via copy-paste to mail)
                            ";".join(
                                [customer.user.email, customer.email2, EMPTY_STRING]
                            ) if customer.email2 else ";".join(
                                [customer.user.email, EMPTY_STRING]
                            )
                        ]
                        for col_num in range(len(row)):
                            c = ws.cell(row=row_num, column=col_num)
                            c.value = row[col_num]
                            if col_num == 4:
                                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            else:
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                if col_num == 0:
                                    c.style.alignment.wrap_text = True
                                else:
                                    c.style.alignment.wrap_text = False
                            if row_num % 2 == 0:
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                        row_num += 1
        else:
            header = [
                (_("Basket"), 20),
                (_('Family'), 35),
                (_('Phone1'), 15),
                (_('Phone2'), 15),
                (_('Total with vat'), 15),
                (_('Email'), 35),
            ]
            wb, ws = new_portrait_a4_sheet(wb, permanence, EMPTY_STRING, header=header, add_print_title=False)
            customer_set = Customer.objects.filter(
                customerinvoice__permanence_id=permanence.id,
                represent_this_buyinggroup=False
            )
            for customer in customer_set:
                invoice = CustomerInvoice.objects.filter(
                    permanence=permanence, customer=customer
                ).order_by('?').first()
                if invoice is not None and invoice.has_purchase:
                    customer.preparation_order = preparation_order
                    # customer.save(update_fields=['preparation_order'])
                    # use vvvv because ^^^^^ will call "pre_save" function which reset valid_email to None
                    Customer.objects.filter(id=customer.id).order_by('?').update(
                        preparation_order=preparation_order
                    )
                    preparation_order += 1
                    if REPANIER_SETTINGS_CUSTOMERS_MUST_CONFIRM_ORDERS and not invoice.is_order_confirm_send:
                        confirmed = "\n{}".format(_("/!\ This order isn't confirmed"))
                    else:
                        confirmed = EMPTY_STRING
                    row = [
                        "{} - {}{}".format(customer.preparation_order, customer.long_basket_name, confirmed),
                        customer.long_basket_name,
                        customer.phone1,
                        customer.phone2,
                        invoice.total_price_with_tax.amount,
                        # Used to send mail to customer with an order (via copy-paste to mail)
                        ";".join(
                            [customer.user.email, customer.email2, EMPTY_STRING]
                        ) if customer.email2 else ";".join(
                            [customer.user.email, EMPTY_STRING]
                        )
                    ]
                    for col_num in range(len(row)):
                        c = ws.cell(row=row_num, column=col_num)
                        c.value = row[col_num]
                        if col_num == 4:
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                        else:
                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                            if col_num == 0:
                                c.style.alignment.wrap_text = True
                            else:
                                c.style.alignment.wrap_text = False
                        if row_num % 2 == 0:
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                    row_num += 1

        # Permanence board info
        permanence_date_save = None
        at_least_one = False
        next_permanence_set = Permanence.objects.filter(permanence_date__gte=permanence.permanence_date).order_by(
            "permanence_date")[:5]
        for next_permanence in next_permanence_set:

            for permanenceboard in PermanenceBoard.objects.filter(
                    permanence_id=next_permanence.id).order_by(
                "permanence_role__tree_id",
                "permanence_role__lft"
            ):
                customer = permanenceboard.customer
                if customer is not None:
                    if at_least_one == False:
                        for col_num in range(6):
                            c = ws.cell(row=row_num - 1, column=col_num)
                            c.style.borders.top.border_style = Border.BORDER_THIN
                            c.style.borders.bottom.border_style = Border.BORDER_THIN

                        c = ws.cell(row=row_num, column=0)
                        c.value = "-------"
                        c = ws.cell(row=row_num, column=1)
                        c.value = "{}".format(_('Permanence Board Member List'))
                        c.style.alignment.wrap_text = False
                        c.style.font.bold = True
                        c = ws.cell(row=row_num, column=2)
                        c.value = "-------"
                        row_num += 1

                    at_least_one = True
                    row = [
                        (next_permanence.permanence_date, NumberFormat.FORMAT_DATE_DMYSLASH),
                        (customer.long_basket_name, NumberFormat.FORMAT_TEXT),
                        (customer.phone1, NumberFormat.FORMAT_TEXT),
                        (customer.phone2, NumberFormat.FORMAT_TEXT),
                        (permanenceboard.permanence_role.short_name, NumberFormat.FORMAT_TEXT),
                    ]
                    for col_num in range(len(row)):
                        c = ws.cell(row=row_num, column=col_num)
                        c.value = "{}".format(row[col_num][0])
                        c.style.number_format.format_code = row[col_num][1]
                        c.style.alignment.wrap_text = False
                        if permanence_date_save != next_permanence.permanence_date:
                            c.style.font.bold = True
                            permanence_date_save = next_permanence.permanence_date
                    row_num += 1

        at_least_one = False

        for staff in Staff.objects.filter(is_active=True):
            customer_responsible = staff.customer_responsible
            if customer_responsible is not None:
                if not at_least_one:
                    for col_num in range(6):
                        c = ws.cell(row=row_num, column=col_num)
                        c.style.borders.top.border_style = Border.BORDER_THIN
                        c.style.borders.bottom.border_style = Border.BORDER_THIN

                    c = ws.cell(row=row_num, column=0)
                    c.value = "-------"
                    c = ws.cell(row=row_num, column=1)
                    c.value = "{}".format(_('Staff Member List'))
                    c.style.alignment.wrap_text = False
                    c.style.font.bold = True
                    c = ws.cell(row=row_num, column=2)
                    c.value = "-------"
                    row_num += 1

                    at_least_one = True

                staff_function = staff.safe_translation_getter(
                    'long_name', any_language=True, default=EMPTY_STRING
                )
                row = [
                    staff_function,
                    customer_responsible.long_basket_name,
                    customer_responsible.phone1,
                    customer_responsible.phone2
                ]
                for col_num in range(len(row)):
                    c = ws.cell(row=row_num, column=col_num)
                    c.value = row[col_num]
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.alignment.wrap_text = True
                row_num += 1

        if not group:
            # This is a private information which doesn't need to be give to customer's groups
            for col_num in range(6):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.top.border_style = Border.BORDER_THIN
                c.style.borders.bottom.border_style = Border.BORDER_THIN

            c = ws.cell(row=row_num, column=0)
            c.value = "-------"
            c = ws.cell(row=row_num, column=1)
            c.value = "{}".format(_('Producers'))
            c.style.alignment.wrap_text = False
            c.style.font.bold = True
            c = ws.cell(row=row_num, column=2)
            c.value = "-------"
            row_num += 1

            # Producer info
            for producer in Producer.objects.filter(permanence=permanence).order_by("short_profile_name"):
                invoice = ProducerInvoice.objects.filter(
                    permanence=permanence, producer=producer
                ).order_by('?').first()
                minimum_order_amount_reached = EMPTY_STRING
                if invoice is None:
                    total_price_with_tax = REPANIER_MONEY_ZERO
                else:
                    total_price_with_tax = invoice.total_price_with_tax
                    if invoice.total_price_with_tax < producer.minimum_order_value:
                        minimum_order_amount_reached = _('Minimum order amount not reached')
                row = [
                    producer.short_profile_name,
                    producer.long_profile_name,
                    producer.phone1,
                    producer.phone2,
                    total_price_with_tax.amount,
                    minimum_order_amount_reached
                ]
                for col_num in range(len(row)):
                    c = ws.cell(row=row_num, column=col_num)
                    c.value = row[col_num]
                    if col_num == 4:
                        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    else:
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.alignment.wrap_text = False
                row_num += 1

        return wb

    return


def export_customer_label(permanence, deliveries_id=None, wb=None):
    # Customer label
    wb, ws = new_portrait_a4_sheet(wb, _('Label'), permanence, add_print_title=False)
    row_num = 0
    customer_set = Customer.objects.filter(
        customerinvoice__permanence_id=permanence.id,
        represent_this_buyinggroup=False,
        preparation_order__gt=0
    ).order_by(
        'preparation_order'
    )

    if deliveries_id is not None:
        customer_set = customer_set.filter(customerinvoice__delivery_id__in=deliveries_id)

    dict_placement = dict(LUT_PRODUCT_PLACEMENT)
    freezer = []
    fridge = []
    out_of_basket = []

    for customer in customer_set:
        customer_identifier = "{} - {}".format(customer.preparation_order, customer.short_basket_name)
        s = [placement_id for placement_id in list(Purchase.objects.filter(
            permanence_id=permanence.id,
            customer_id=customer.id
        ).order_by(
            "offer_item__placement"
        ).values_list(
            "offer_item__placement", flat=True
        ).distinct(
            "offer_item__placement"
        ))]
        placements = ", ".join(["{}".format(dict_placement[placement_id]) for placement_id in s])
        if PRODUCT_PLACEMENT_FREEZER in s:
            freezer += [customer_identifier]
        if PRODUCT_PLACEMENT_FRIDGE in s:
            fridge += [customer_identifier]
        if PRODUCT_PLACEMENT_OUT_OF_BASKET in s:
            out_of_basket += [customer_identifier]

        row_num = customer_label(customer_identifier, placements, row_num, ws)

    placement_label = "[❄ {} ❄]".format(dict_placement[PRODUCT_PLACEMENT_FRIDGE])
    for customer_identifier in fridge:
        row_num = customer_label(customer_identifier, placement_label, row_num, ws)

    placement_label = "[❄❄❄ {} ❄❄❄]".format(dict_placement[PRODUCT_PLACEMENT_FREEZER])
    for customer_identifier in freezer:
        row_num = customer_label(customer_identifier, placement_label, row_num, ws)

    placement_label = "[⏭ {} ⏭]".format(dict_placement[PRODUCT_PLACEMENT_OUT_OF_BASKET])
    for customer_identifier in out_of_basket:
        row_num = customer_label(customer_identifier, placement_label, row_num, ws)

    if row_num > 0:
        ws.column_dimensions[get_column_letter(1)].width = 120

    return wb


def customer_label(customer_identifier, placements, row_num, ws):
    c = ws.cell(row=row_num, column=0)
    c.value = customer_identifier
    c.style.font.size = 36
    c.style.alignment.wrap_text = False
    c.style.borders.top.border_style = Border.BORDER_THIN
    c.style.borders.left.border_style = Border.BORDER_THIN
    c.style.borders.right.border_style = Border.BORDER_THIN
    c.style.alignment.vertical = 'center'
    c.style.alignment.horizontal = 'center'
    row_num += 1
    ws.row_dimensions[row_num].height = 60
    c = ws.cell(row=row_num, column=0)
    c.value = placements
    c.style.borders.left.border_style = Border.BORDER_THIN
    c.style.borders.right.border_style = Border.BORDER_THIN
    c.style.borders.bottom.border_style = Border.BORDER_THIN
    c.style.alignment.vertical = 'center'
    c.style.alignment.horizontal = 'center'
    row_num += 1
    # vvvv c = ... is nedeed for row_dimensions
    c = ws.cell(row=row_num, column=0)
    row_num += 1
    ws.row_dimensions[row_num].height = 5
    # ^^^^ c = ... is nedeed for row_dimensions
    return row_num


def export_preparation(permanence, deliveries_id=None, wb=None):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Purchase"), 10),
        (_("OfferItem"), 5),
        (_("Placement"), 7),
        (_("Producer"), 15),
        (EMPTY_STRING, 7), # This column isn't any more used
        (_("Product"), 60),
        (_("Customer"), 15),
        (_("Quantity ordered"), 10),
        (_("Unit"), 10),
        (_("Prepared"), 22),
        (_("To distribute"), 10),
    ]
    if deliveries_id is None:
        return export_preparation_for_a_delivery(0, None, header, permanence, wb, yellowFill)
    else:
        for delivery_ref, delivery_id in enumerate(deliveries_id):
            wb = export_preparation_for_a_delivery(delivery_ref, delivery_id, header, permanence, wb, yellowFill)
        return wb


def export_preparation_for_a_delivery(delivery_cpt, delivery_id, header, permanence, wb, yellowFill):
    producer_set = Producer.objects.filter(
        producerinvoice__permanence_id=permanence.id
    ).only('invoice_by_basket', 'short_profile_name')
    if delivery_id is not None:
        producer_set = producer_set.filter(purchase__customer_invoice__delivery_id=delivery_id)
    producers = producer_set.distinct().iterator()
    producer = next_row(producers)
    if producer is not None:
        wb, ws = new_landscape_a4_sheet(
            wb,
            _("Preparation") if delivery_id is None else "{}-{}".format(delivery_cpt, _("Preparation")),
            permanence,
            header
        )
        row_num = 1
        if delivery_id is not None:
            c = ws.cell(row=row_num, column=5)
            c.style.font.bold = True
            c.value = DeliveryBoard.objects.filter(id=delivery_id).order_by('?').first().get_delivery_display()
            row_num += 1
        producer_counter = 0
        hide_column_placement = True
        placement_save = None
        while producer is not None:
            producer_counter += 1
            producer_save = producer
            at_least_one_product = False
            ######################################################################################################################
            if producer.invoice_by_basket and not producer.represent_this_buyinggroup:
                # If the producer manage the production, he need to go into his field to pick up products.
                # In this cas, the preparation list must not be done by basket.
                # But the invoice must be done by basket.
                purchase_set = Purchase.objects.filter(
                    permanence_id=permanence.id,
                    producer_id=producer.id,
                    offer_item__translations__language_code=translation.get_language()
                ).order_by(
                    "customer__short_basket_name",
                    "offer_item__translations__preparation_sort_order"
                ).select_related('customer', 'offer_item', 'offer_item__department_for_customer')
                if delivery_id is not None:
                    purchase_set = purchase_set.filter(customer_invoice__delivery_id=delivery_id)
                purchases = purchase_set.iterator()
                purchase = next_purchase(purchases)
                while purchase is not None:
                    at_least_one_product = True
                    customer_save = purchase.customer
                    c = ws.cell(row=row_num, column=6)
                    c.style.font.bold = True
                    row_num += 1
                    count_purchase = 0
                    purchases_price = DECIMAL_ZERO
                    purchases_price_formula = []
                    while purchase is not None and customer_save.id == purchase.customer_id:
                        department_for_customer_save = purchase.offer_item.department_for_customer
                        department_for_customer_save__short_name = department_for_customer_save.short_name \
                            if department_for_customer_save is not None else EMPTY_STRING
                        while purchase is not None and customer_save.id == purchase.customer_id \
                                and department_for_customer_save == purchase.offer_item.department_for_customer:
                            qty = purchase.get_producer_quantity()
                            if qty != DECIMAL_ZERO:
                                base_unit = get_base_unit(
                                    qty,
                                    purchase.offer_item.order_unit,
                                    purchase.status,
                                    producer=True
                                )
                                c = ws.cell(row=row_num, column=0)
                                c.value = purchase.id
                                c = ws.cell(row=row_num, column=1)
                                c.value = purchase.offer_item_id
                                c.style.number_format.format_code = NumberFormat.FORMAT_DATE_DDMMYYYY
                                c = ws.cell(row=row_num, column=2)
                                c.value = purchase.offer_item.get_placement_display()
                                if placement_save is None:
                                    placement_save = c.value
                                elif hide_column_placement:
                                    if placement_save != c.value:
                                        hide_column_placement = False
                                c = ws.cell(row=row_num, column=3)
                                c.value = producer_save.short_profile_name
                                c = ws.cell(row=row_num, column=5)
                                if department_for_customer_save__short_name is not None:
                                    c.value = "{} - {}".format(
                                        purchase.get_long_name(), department_for_customer_save__short_name)
                                else:
                                    c.value = "{}".format(purchase.get_long_name())
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c.style.alignment.wrap_text = True
                                c = ws.cell(row=row_num, column=6)
                                c.value = "{} - {}".format(customer_save.preparation_order, customer_save.short_basket_name)
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=7)
                                c.value = qty
                                c.style.number_format.format_code = '#,##0.????'
                                c.style.font.color = Color(Color.BLUE)
                                ws.conditional_formatting.addCellIs(
                                    get_column_letter(8) + str(row_num + 1), 'notEqual',
                                    [str(qty)], True, wb,
                                    None, None, yellowFill
                                )
                                c = ws.cell(row=row_num, column=8)
                                c.value = "{}".format(base_unit)
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                customer_unit_price = purchase.get_customer_unit_price()
                                if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                                    price_qty = qty * purchase.offer_item.order_average_weight
                                    purchases_price_formula.append("ROUND(H{}*{}*{},2)".format(
                                                                   row_num + 1,
                                                                    purchase.offer_item.order_average_weight,
                                                                    customer_unit_price + purchase.offer_item.unit_deposit.amount))
                                else:
                                    price_qty = qty
                                    purchases_price_formula.append("ROUND(H{}*{},2)".format(
                                                                    row_num + 1,
                                                                    customer_unit_price + purchase.offer_item.unit_deposit.amount))
                                if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                                    c = ws.cell(row=row_num, column=9)
                                    if purchase.offer_item.wrapped:
                                        c.value = "{} :".format(repanier.apps.REPANIER_SETTINGS_CURRENCY_DISPLAY)
                                    else:
                                        c.value = "{}".format(_('kg :'))
                                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                else:
                                    if purchase.offer_item.wrapped:
                                        c = ws.cell(row=row_num, column=9)
                                        c.value = "{} :".format(repanier.apps.REPANIER_SETTINGS_CURRENCY_DISPLAY)
                                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                purchases_price += (price_qty *
                                                    (customer_unit_price + purchase.offer_item.unit_deposit.amount)
                                                    ).quantize(TWO_DECIMALS)
                                count_purchase += 1
                                delta = 6
                                for col_num in range(4):
                                    c = ws.cell(row=row_num, column=delta + col_num)
                                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                                row_num += 1

                            purchase = next_purchase(purchases)
                    c = ws.cell(row=row_num - 1, column=10)
                    if len(purchases_price_formula) > 0:
                        c.value = "={}".format("+".join(purchases_price_formula))
                    else:
                        c.value = DECIMAL_ZERO
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.color = Color(Color.BLUE)
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(11) + str(row_num), 'notEqual',
                        [str(purchases_price)], True, wb,
                        None, None, yellowFill
                    )
                if at_least_one_product:
                    row_num -= 1
                #############################################################################################################################
            else:
                # Using quantity_for_preparation_sort_order the order is by customer__short_basket_name if the product
                # is to be distributed by piece, otherwise by lower qty first.
                for offer_item in OfferItemWoReceiver.objects.filter(
                        permanence_id=permanence.id,
                        producer_id=producer.id,
                        translations__language_code=translation.get_language()
                ).order_by(
                    "translations__preparation_sort_order"
                ):
                    purchase_set = Purchase.objects.filter(
                        offer_item_id=offer_item.id
                    ).order_by(
                        "quantity_for_preparation_sort_order",
                        "customer__short_basket_name"
                    ).select_related('customer', 'offer_item', 'offer_item__department_for_customer')
                    if delivery_id is not None:
                        purchase_set = purchase_set.filter(customer_invoice__delivery_id=delivery_id)
                    purchases = purchase_set.iterator()
                    purchase = next_purchase(purchases)
                    while purchase is not None:
                        at_least_one_product = True
                        while purchase is not None:
                            department_for_customer_save = purchase.offer_item.department_for_customer
                            department_for_customer_save__short_name = department_for_customer_save.short_name \
                                if department_for_customer_save is not None else None
                            offer_item_save = purchase.offer_item
                            row_start_offer_item = row_num + 2
                            count_offer_item = 0
                            for col_num in range(11):
                                c = ws.cell(row=row_num, column=col_num)
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                            purchases_quantity = DECIMAL_ZERO
                            row_num += 1
                            while purchase is not None and offer_item_save == purchase.offer_item:
                                qty = purchase.get_producer_quantity()
                                if qty != DECIMAL_ZERO:
                                    base_unit = get_base_unit(
                                        qty,
                                        offer_item_save.order_unit,
                                        purchase.status,
                                        producer=True
                                    )
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = purchase.id
                                    c = ws.cell(row=row_num, column=1)
                                    c.value = purchase.offer_item_id
                                    if count_offer_item == 0:
                                        c = ws.cell(row=row_num, column=2)
                                        c.value = purchase.offer_item.get_placement_display()
                                        if placement_save is None:
                                            placement_save = c.value
                                        elif hide_column_placement:
                                            if placement_save != c.value:
                                                hide_column_placement = False
                                        c = ws.cell(row=row_num, column=3)
                                        c.value = producer_save.short_profile_name
                                    c = ws.cell(row=row_num, column=5)
                                    if department_for_customer_save__short_name is not None:
                                        c.value = "{} - {}".format(
                                            purchase.get_long_name(), department_for_customer_save__short_name)
                                    else:
                                        c.value = "{}".format(purchase.get_long_name())
                                    c.style.alignment.wrap_text = True
                                    if count_offer_item != 0:
                                        c.style.font.color.index = 'FF939393'
                                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                    c = ws.cell(row=row_num, column=6)
                                    c.value = "{} - {}".format(purchase.customer.preparation_order, purchase.customer.short_basket_name)
                                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                    c = ws.cell(row=row_num, column=7)
                                    c.value = qty
                                    c.style.number_format.format_code = '#,##0.????'
                                    c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(8) + str(row_num + 1), 'notEqual',
                                        [str(qty)], True, wb,
                                        None, None, yellowFill
                                    )
                                    c = ws.cell(row=row_num, column=8)
                                    c.value = "{}".format(base_unit)
                                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                    if offer_item_save.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                                        c = ws.cell(row=row_num, column=9)
                                        if offer_item_save.wrapped:
                                            c.value = "{} :".format(repanier.apps.REPANIER_SETTINGS_CURRENCY_DISPLAY)
                                        else:
                                            c.value = "{}".format(_('kg :'))
                                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                    else:
                                        if offer_item_save.wrapped:
                                            c = ws.cell(row=row_num, column=9)
                                            c.value = "{} :".format(repanier.apps.REPANIER_SETTINGS_CURRENCY_DISPLAY)
                                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                    purchases_quantity += qty
                                    count_offer_item += 1
                                    delta = 6
                                    for col_num in range(4):
                                        c = ws.cell(row=row_num, column=delta + col_num)
                                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                                    row_num += 1

                                purchase = next_purchase(purchases)
                            if count_offer_item > 1:
                                c = ws.cell(row=row_num - 1, column=10)
                                c.value = "=SUM(H{}:H{})".format(row_start_offer_item, row_num)
                                c.style.number_format.format_code = '#,##0.????'
                                if not offer_item_save.wrapped and offer_item_save.order_unit in [PRODUCT_ORDER_UNIT_KG,
                                                                                                  PRODUCT_ORDER_UNIT_PC_KG]:
                                    c.style.font.color = Color(Color.BLUE)
                                ws.conditional_formatting.addCellIs(
                                    get_column_letter(11) + str(row_num), 'notEqual',
                                    [str(purchases_quantity)], True, wb,
                                    None, None, yellowFill
                                )
                            row_num -= 1

            if at_least_one_product:
                for col_num in range(11):
                    c = ws.cell(row=row_num, column=col_num)
                    c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
                row_num += 1
            producer = next_row(producers)
        ws.column_dimensions[get_column_letter(1)].visible = False
        ws.column_dimensions[get_column_letter(2)].visible = False
        if hide_column_placement:
            ws.column_dimensions[get_column_letter(3)].visible = False
        if producer_counter <= 1:
            # hide producer name
            ws.column_dimensions[get_column_letter(4)].visible = False
        ws.column_dimensions[get_column_letter(5)].visible = False
    return wb


def export_producer_by_product(permanence, producer, wb=None):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    language_code = translation.get_language()
    translation.activate(producer.language)

    header = [
        (_("Basket"), 20),
        (_("Quantity"), 10),
        (_("Unit"), 10),
        (_("Reference"), 20),
        (_("Product"), 60),
        (_("Unit price"), 10),
        (_("Deposit"), 10),
        (_("Total Price"), 12),
    ]
    show_column_reference = False
    hide_column_short_basket_name = True
    hide_column_unit_deposit = True
    formula_main_total = []
    offer_items = OfferItemWoReceiver.objects.filter(
        permanence_id=permanence.id,
        producer_id=producer.id,
        translations__language_code=translation.get_language(),
        quantity_invoiced__gt=DECIMAL_ZERO
    ).order_by(
        "department_for_customer",
        "translations__producer_sort_order"
    ).iterator()
    offer_item = next_row(offer_items)
    if offer_item:
        wb, ws = new_landscape_a4_sheet(
            wb,
            "{} {}".format(producer.short_profile_name, _("by product")),
            permanence,
            header
        )
        row_num = 1
        producer_purchases_price = DECIMAL_ZERO
        status = ProducerInvoice.objects.filter(
            permanence_id=permanence.id, producer_id=producer.id
        ).order_by('?').only("status").first().status
        while offer_item is not None:
            department_for_customer_save = offer_item.department_for_customer
            c = ws.cell(row=row_num, column=1)
            c.value = department_for_customer_save.short_name \
                if department_for_customer_save is not None else "---"
            c.style.font.bold = True
            row_num += 1
            row_start_department = row_num
            department_purchases_price = DECIMAL_ZERO
            while offer_item is not None and department_for_customer_save == offer_item.department_for_customer:
                if offer_item.unit_deposit != DECIMAL_ZERO:
                    hide_column_unit_deposit = False
                show_column_reference = show_column_reference | len(offer_item.reference) < 36
                if offer_item.wrapped:
                    hide_column_short_basket_name = False
                    first_purchase = True
                    purchase_set = Purchase.objects.filter(
                        offer_item_id=offer_item.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).order_by(
                        "offer_item__translations__producer_sort_order",
                        "customer__short_basket_name"
                    )
                    for purchase in purchase_set:
                        if offer_item.limit_order_quantity_to_stock:
                            # Don't purchase anything to the producer in this case
                            qty = DECIMAL_ZERO
                        else:
                            qty = purchase.get_producer_quantity()
                        if qty != DECIMAL_ZERO:
                            base_unit = get_base_unit(
                                qty,
                                offer_item.order_unit,
                                status,
                                producer=True
                            )
                            c = ws.cell(row=row_num, column=0)
                            c.value = "{} - {}".format(purchase.customer.preparation_order, purchase.customer.short_basket_name)
                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=1)
                            c.value = qty
                            c.style.number_format.format_code = '#,##0.???'
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c.style.font.bold = True
                            c.style.font.color = Color(Color.BLUE)
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(2) + str(row_num + 1), 'notEqual',
                                [str(qty)], True, wb,
                                None, None, yellowFill
                            )
                            c = ws.cell(row=row_num, column=2)
                            c.value = "{}".format(base_unit)
                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=3)
                            reference = offer_item.reference if len(
                                offer_item.reference) < 36 else EMPTY_STRING
                            if not reference.isdigit():
                                # Avoid display of exponent by Excel
                                c.value = "[{}]".format(reference)
                            else:
                                c.value = reference
                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=4)
                            c.value = offer_item.get_long_name(customer_price=False)
                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                            c.style.alignment.wrap_text = True
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            if first_purchase:
                                first_purchase = False
                            else:
                                c.style.font.color.index = 'FF939393'
                            c = ws.cell(row=row_num, column=5)
                            producer_unit_price = purchase.get_producer_unit_price()
                            customer_unit_price = purchase.get_customer_unit_price()
                            if producer_unit_price < customer_unit_price:
                                unit_price = producer_unit_price
                            else:
                                unit_price = customer_unit_price
                            c.value = unit_price
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=6)
                            unit_deposit = offer_item.unit_deposit.amount
                            c.value = unit_deposit
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=7)
                            if offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                                price_qty = qty * offer_item.order_average_weight
                                c.value = "=ROUND(B{}*{}*(F{}+G{}),2)".format(
                                row_num + 1, offer_item.order_average_weight, row_num + 1, row_num + 1)
                            else:
                                price_qty = qty
                                c.value = "=ROUND(B{}*(F{}+G{}),2)".format(row_num + 1, row_num + 1, row_num + 1)
                            purchase_price = (
                                price_qty * (unit_price + unit_deposit)
                            ).quantize(TWO_DECIMALS)
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            department_purchases_price += purchase_price
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(8) + str(row_num + 1), 'notEqual',
                                [str(purchase_price)], True, wb,
                                None, None, yellowFill
                            )

                            row_num += 1
                else:
                    qty, taken_from_stock, customer_qty = offer_item.get_producer_qty_stock_invoiced()
                    if qty != DECIMAL_ZERO:
                        # Important : in this case, the qty comes from offer item.
                        # Offer item contains weight, not pieces
                        if offer_item.use_order_unit_converted:
                            if offer_item.order_average_weight > DECIMAL_ZERO:
                                qty = (qty / offer_item.order_average_weight).quantize(TWO_DECIMALS)
                        base_unit = get_base_unit(
                            qty,
                            offer_item.order_unit,
                            status,
                            producer=True
                        )
                        c = ws.cell(row=row_num, column=0)
                        c.value = repanier.apps.REPANIER_SETTINGS_GROUP_NAME
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=1)
                        c.value = qty
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c.style.font.bold = True
                        c.style.font.color = Color(Color.BLUE)
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(2) + str(row_num + 1), 'notEqual',
                            [str(qty)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=2)
                        c.value = "{}".format(base_unit)
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=3)
                        reference = offer_item.reference if len(
                            offer_item.reference) < 36 else EMPTY_STRING
                        if not reference.isdigit():
                            # Avoid display of exponent by Excel
                            c.value = "[{}]".format(reference)
                        else:
                            c.value = reference
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=4)
                        c.value = offer_item.get_long_name(customer_price=False)
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.alignment.wrap_text = True
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=5)
                        producer_unit_price = offer_item.producer_unit_price.amount
                        customer_unit_price = offer_item.customer_unit_price.amount
                        if producer_unit_price < customer_unit_price:
                            unit_price = producer_unit_price
                        else:
                            unit_price = customer_unit_price
                        c.value = unit_price
                        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=6)
                        unit_deposit = offer_item.unit_deposit.amount
                        c.value = unit_deposit
                        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=7)
                        if offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                            price_qty = qty * offer_item.order_average_weight
                            c.value = "=ROUND(B{}*{}*(F{}+G{}),2)".format(
                            row_num + 1, offer_item.order_average_weight, row_num + 1, row_num + 1)
                        else:
                            price_qty = qty
                            c.value = "=ROUND(B{}*(F{}+G{}),2)".format(row_num + 1, row_num + 1, row_num + 1)
                        purchase_price = (
                            price_qty * (unit_price + unit_deposit)
                        ).quantize(TWO_DECIMALS)
                        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        department_purchases_price += purchase_price
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(8) + str(row_num + 1), 'notEqual',
                            [str(purchase_price)], True, wb,
                            None, None, yellowFill
                        )

                        row_num += 1
                offer_item = next_row(offer_items)
            for col_num in range(8):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if col_num == 4:
                    if producer.producer_price_are_wo_vat:
                        c.value = "{} {} {}".format(
                            _("Total Price"), _("Wo tax"), department_for_customer_save.short_name \
                                if department_for_customer_save is not None else "---")
                    else:
                        c.value = "{} {} {}".format(
                            _("Total Price"), _("w tax"), department_for_customer_save.short_name \
                                if department_for_customer_save is not None else "---")
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                if col_num == 7:
                    formula = "SUM(H{}:H{})".format(row_start_department, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                    producer_purchases_price += department_purchases_price
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(8) + str(row_num + 1), 'notEqual',
                        [str(department_purchases_price)], True, wb,
                        None, None, yellowFill
                    )
                    formula_main_total.append(formula)
            row_num += 1
        for col_num in range(8):
            c = ws.cell(row=row_num, column=col_num)
            c.style.borders.bottom.border_style = Border.BORDER_THIN
            if col_num == 1:
                if producer.producer_price_are_wo_vat:
                    c.value = "{} {} {}".format(_("Total Price"), _("wo tax"), repanier.apps.REPANIER_SETTINGS_GROUP_NAME)
                else:
                    c.value = "{} {} {}".format(_("Total Price"), _("w tax"), repanier.apps.REPANIER_SETTINGS_GROUP_NAME)
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            if col_num == 7:
                c.value = "=" + "+".join(formula_main_total)
                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                c.style.font.bold = True
                ws.conditional_formatting.addCellIs(
                    get_column_letter(8) + str(row_num + 1), 'notEqual',
                    [str(producer_purchases_price)], True, wb,
                    None, None, yellowFill
                )

        if hide_column_short_basket_name:
            ws.column_dimensions[get_column_letter(1)].visible = False
        if not show_column_reference:
            ws.column_dimensions[get_column_letter(4)].visible = False
        if hide_column_unit_deposit:
            ws.column_dimensions[get_column_letter(7)].visible = False
    translation.activate(language_code)
    return wb


def export_producer_by_customer(permanence, producer, wb=None):
    language_code = translation.get_language()
    translation.activate(producer.language)

    header = [
        (_("Quantity"), 10),
        (_("Unit"), 10),
        (_("Reference"), 20),
        (_("Product"), 60),
        (_("Unit price"), 10),
        (_("Deposit"), 10),
        (_("Total Price"), 12),
        (EMPTY_STRING, 15),
    ]
    show_column_reference = False
    hide_column_unit_deposit = True
    formula_main_total = []
    purchase_set = Purchase.objects.filter(
        permanence_id=permanence.id,
        producer_id=producer.id,
        offer_item__translations__language_code=translation.get_language(),
    ).order_by(
        "customer__short_basket_name",
        "offer_item__translations__preparation_sort_order"
    ).select_related("offer_item")
    purchases = purchase_set.iterator()
    purchase = next_row(purchases)
    if purchase:
        wb, ws = new_landscape_a4_sheet(
            wb,
            "{} {}".format(producer.short_profile_name, _("by basket")),
            permanence,
            header
        )
        row_num = 1
        while purchase is not None:
            customer_save = purchase.customer
            c = ws.cell(row=row_num, column=0)
            c.value = "{} - {}".format(customer_save.preparation_order, customer_save.short_basket_name)
            c.style.font.bold = True
            row_num += 1
            row_start_customer = row_num
            first_purchase = True
            while purchase is not None and customer_save.id == purchase.customer_id:
                qty = purchase.get_producer_quantity()
                if qty != DECIMAL_ZERO:
                    offer_item_save = purchase.offer_item
                    base_unit = get_base_unit(
                        qty,
                        purchase.offer_item.order_unit,
                        purchase.status,
                        producer=True
                    )
                    c = ws.cell(row=row_num, column=0)
                    c.value = qty
                    c.style.number_format.format_code = '#,##0.???'
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c.style.font.bold = True
                    c = ws.cell(row=row_num, column=1)
                    c.value = "{}".format(base_unit)
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=2)
                    reference = offer_item_save.reference if len(offer_item_save.reference) < 36 else EMPTY_STRING
                    if not reference.isdigit():
                        # Avoid display of exponent by Excel
                        c.value = "[{}]".format(reference)
                    else:
                        c.value = reference
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=3)
                    c.value = offer_item_save.get_long_name(customer_price=False)
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.alignment.wrap_text = True
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    if first_purchase:
                        first_purchase = False
                    c = ws.cell(row=row_num, column=4)
                    producer_unit_price = purchase.get_producer_unit_price()
                    customer_unit_price = purchase.get_customer_unit_price()
                    if producer_unit_price < customer_unit_price:
                        unit_price = producer_unit_price
                    else:
                        unit_price = customer_unit_price
                    c.value = unit_price
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=5)
                    c.value = offer_item_save.unit_deposit.amount
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=6)
                    if offer_item_save.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                        c.value = "=ROUND(A{}*{}*(E{}+F{}),2)".format(
                        row_num + 1, offer_item_save.order_average_weight, row_num + 1, row_num + 1)
                    else:
                        c.value = "=ROUND(A{}*(E{}+F{}),2)".format(row_num + 1, row_num + 1, row_num + 1)
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    row_num += 1
                purchase = next_row(purchases)
            for col_num in range(7):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if col_num == 3:
                    if producer.producer_price_are_wo_vat:
                        c.value = "{} {} {}".format(
                            _("Total Price"), _("wo tax"), customer_save.short_basket_name)
                    else:
                        c.value = "{} {} {}".format(
                            _("Total Price"), _("w tax"), customer_save.short_basket_name)
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                if col_num == 6:
                    formula = "SUM(G{}:G{})".format(row_start_customer, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                    c.style.font.bold = True
                    formula_main_total.append(formula)
            row_num += 1
        for col_num in range(7):
            c = ws.cell(row=row_num, column=col_num)
            c.style.borders.bottom.border_style = Border.BORDER_THIN
            if col_num == 0:
                if producer.producer_price_are_wo_vat:
                    c.value = "{} {} {}".format(_("Total Price"), _("wo tax"), repanier.apps.REPANIER_SETTINGS_GROUP_NAME)
                else:
                    c.value = "{} {} {}".format(_("Total Price"), _("w tax"), repanier.apps.REPANIER_SETTINGS_GROUP_NAME)
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            if col_num == 6:
                c.value = "=" + "+".join(formula_main_total)
                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                c.style.font.bold = True
        if hide_column_unit_deposit:
            ws.column_dimensions[get_column_letter(6)].visible = False
        if not show_column_reference:
            ws.column_dimensions[get_column_letter(3)].visible = False
    translation.activate(language_code)
    return wb


def export_customer(
        permanence=None,
        customer=None,
        deliveries_id=None,
        deposit=False,
        xlsx_formula=True,
        wb=None, ws_preparation_title=None):

    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Purchase"), 5),
        (_("OfferItem"), 5),
        (_("Placement"), 15),
        (_("Producer"), 15),
        (_("Product"), 60),
        (_("Quantity"), 10),
        (_("Unit"), 10),
        (_("Unit price"), 10),
        (_("Deposit"), 10),
        (_("Total Price"), 12),
        (_("Basket"), 20),
    ]
    if deliveries_id is None:
        return export_customer_for_a_delivery(
            customer, 0, None, deposit, header, permanence, wb,
            ws_preparation_title, yellowFill, xlsx_formula
        )
    else:
        for delivery_cpt, delivery_id in enumerate(deliveries_id):
            wb = export_customer_for_a_delivery(
                customer, delivery_cpt, delivery_id, deposit, header, permanence, wb,
                ws_preparation_title, yellowFill, xlsx_formula
            )
        return wb


def export_customer_for_a_delivery(
        customer, delivery_cpt, delivery_id, deposit, header, permanence, wb, ws_preparation_title,
        yellowFill, xlsx_formula):
    language_code = translation.get_language()
    if customer is not None:
        translation.activate(customer.language)
        if deposit:
            purchase_set = Purchase.objects.filter(
                permanence_id=permanence.id,
                customer_id=customer.id,
                producer__isnull=False,
                offer_item__translations__language_code=customer.language,
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "offer_item__translations__long_name",
                "offer_item__order_average_weight",
            ).select_related('customer', 'offer_item', 'offer_item__department_for_customer')
        else:
            purchase_set = Purchase.objects.filter(
                permanence_id=permanence.id,
                customer_id=customer.id,
                producer__isnull=False,
                offer_item__translations__language_code=customer.language,
            ).exclude(
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "offer_item__translations__long_name",
                "offer_item__order_average_weight",
            ).select_related('customer', 'offer_item', 'offer_item__department_for_customer')
    else:
        if deposit:
            purchase_set = Purchase.objects.filter(
                permanence_id=permanence.id,
                producer__isnull=False,
                offer_item__translations__language_code=language_code,
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "customer__short_basket_name",
                "offer_item__producer",
                "offer_item__translations__long_name",
                "offer_item__order_average_weight",
            ).select_related('customer', 'offer_item', 'offer_item__department_for_customer')
        else:
            purchase_set = Purchase.objects.filter(
                permanence_id=permanence.id,
                producer__isnull=False,
                offer_item__translations__language_code=language_code,
            ).exclude(
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "customer__short_basket_name",
                "offer_item__producer",
                "offer_item__translations__long_name",
                "offer_item__order_average_weight",
            ).select_related('customer', 'offer_item', 'offer_item__department_for_customer')
        if delivery_id is not None:
            purchase_set = purchase_set.filter(customer_invoice__delivery_id=delivery_id)
    purchases = purchase_set.iterator()
    purchase = next_purchase(purchases)
    if purchase is not None:
        config = Configuration.objects.get(id=DECIMAL_ONE)
        group_label = config.group_label
        if deposit:
            wb, ws = new_portrait_a4_sheet(
                wb,
                _("Deposits") if delivery_id is None else "{}-{}".format(delivery_cpt, _("Deposits")),
                permanence,
                header
            )
        else:
            if repanier.apps.REPANIER_SETTINGS_PAGE_BREAK_ON_CUSTOMER_CHECK:
                # Change the orientation to reduce the number of page breaks, i.e. the number of printed pages
                wb, ws = new_portrait_a4_sheet(
                    wb,
                    _("Customer check") if delivery_id is None else "{}-{}".format(delivery_cpt, _("Customer check")),
                    permanence,
                    header
                )
            else:
                wb, ws = new_landscape_a4_sheet(
                    wb,
                    _("Customer check") if delivery_id is None else "{}-{}".format(delivery_cpt, _("Customer check")),
                    permanence,
                    header
                )
        hide_column_placement = True
        hide_column_producer = True
        offer_item_save = purchase.offer_item
        placement_save = offer_item_save.placement
        producer_save = offer_item_save.producer
        row_num = 1
        if delivery_id is not None:
            c = ws.cell(row=row_num, column=4)
            c.style.font.bold = True
            c.value = DeliveryBoard.objects.filter(id=delivery_id).order_by('?').first().get_delivery_display()
            row_num += 1
        if ws_preparation_title is not None and xlsx_formula:
            ref_preparation_sheet = ws_preparation_title if delivery_id is None else "{}-{}".format(
            delivery_cpt, ws_preparation_title)
        else:
            ref_preparation_sheet = None
        while purchase is not None:
            customer_save = purchase.customer
            row_start_customer = row_num + 1
            first_purchase = True
            total_price = DECIMAL_ZERO
            while purchase is not None and customer_save.id == purchase.customer_id:
                department_for_customer_save__id = offer_item_save.department_for_customer_id
                department_for_customer_save__short_name = offer_item_save.department_for_customer.short_name \
                    if offer_item_save.department_for_customer is not None else None
                while purchase is not None and customer_save.id == purchase.customer_id and \
                                offer_item_save.department_for_customer_id == department_for_customer_save__id:
                    if placement_save != offer_item_save.placement:
                        hide_column_placement = False
                    if producer_save != offer_item_save.producer:
                        hide_column_producer = False
                    qty = purchase.get_producer_quantity()
                    if deposit or qty != DECIMAL_ZERO:
                        base_unit = get_base_unit(
                            qty,
                            offer_item_save.order_unit,
                            purchase.status,
                            producer=True
                        )
                        c = ws.cell(row=row_num, column=0)
                        c.value = purchase.id
                        c = ws.cell(row=row_num, column=1)
                        c.value = offer_item_save.id
                        c = ws.cell(row=row_num, column=2)
                        c.value = offer_item_save.get_placement_display()
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=3)
                        c.value = purchase.producer.short_profile_name
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=4)
                        if department_for_customer_save__short_name is not None:
                            c.value = "{} - {}".format(purchase.get_long_name(), department_for_customer_save__short_name)
                        else:
                            c.value = "{}".format(purchase.get_long_name())
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.alignment.wrap_text = True
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=5)
                        if ref_preparation_sheet is None:
                            c.value = qty
                        else:
                            c.value = "=SUMIF('{}'!A2:A5000,A{},'{}'!H2:H5000)".format(
                                      ref_preparation_sheet, row_num + 1, ref_preparation_sheet
                            )

                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        if not deposit:
                            c.style.font.color = Color(Color.BLUE)
                        if xlsx_formula:
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(6) + str(row_num + 1), 'notEqual',
                                [str(qty)], True, wb,
                                None, None, yellowFill
                            )
                        c = ws.cell(row=row_num, column=6)
                        c.value = "{}".format(base_unit)
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        if purchase.is_box_content:
                            # No price infos : The customer pay for a box not for a product
                            for x in range(6, 10):
                                c = ws.cell(row=row_num, column=x)
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                        else:
                            c = ws.cell(row=row_num, column=7)
                            customer_unit_price = purchase.get_customer_unit_price()
                            c.value = customer_unit_price
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=8)
                            unit_deposit = offer_item_save.unit_deposit.amount
                            c.value = unit_deposit
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=9)
                            if offer_item_save.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                                price_qty = qty * offer_item_save.order_average_weight
                                c.value = "=ROUND(F{}*{}*(H{}+I{}),2)".format(
                                    row_num + 1, offer_item_save.order_average_weight, row_num + 1, row_num + 1)
                            else:
                                price_qty = qty
                                c.value = "=ROUND(F{}*(H{}+I{}),2)".format(row_num + 1, row_num + 1, row_num + 1)
                            purchases_price = (
                                price_qty * (customer_unit_price + unit_deposit)
                            ).quantize(TWO_DECIMALS)
                            if not xlsx_formula:
                                c.value = purchases_price
                                total_price += purchases_price
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            if xlsx_formula:
                                ws.conditional_formatting.addCellIs(
                                    get_column_letter(10) + str(row_num + 1), 'notEqual',
                                    [str(purchases_price)], True, wb,
                                    None, None, yellowFill
                                )

                        c = ws.cell(row=row_num, column=10)
                        if xlsx_formula:
                            c.value = "{} - {}".format(purchase.customer.preparation_order, purchase.customer.short_basket_name)
                        else:
                            c.value = purchase.customer.long_basket_name
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        if first_purchase:
                            c.style.font.bold = True
                            first_purchase = False
                        c.style.font.italic = True
                        row_num += 1
                    purchase = next_purchase(purchases)
                    offer_item_save = purchase.offer_item if purchase is not None else None
                # if not deposit and repanier.apps.REPANIER_SETTINGS_PAGE_BREAK_ON_CUSTOMER_CHECK:
                #     # Need Openpyxl 2.x
                #     page_break = Break(id=row_number)  # create Break obj
                #     ws.page_breaks.append(page_break)
            row_num += 1
            c = ws.cell(row=row_num, column=4)
            c.value = group_label
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            c.style.font.bold = True
            c = ws.cell(row=row_num, column=8)
            c.value = "{} {}".format(_("Total Price"), customer_save.long_basket_name)
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            c.style.font.bold = True
            c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
            c = ws.cell(row=row_num, column=9)
            if xlsx_formula:
                c.value = "=SUM(J{}:J{})".format(row_start_customer, row_num)
            else:
                c.value = total_price
            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
            c.style.font.bold = True
            # Display a separator line between customers
            row_num += 1
            for col_num in range(11):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
            row_num += 2

        ws.column_dimensions[get_column_letter(1)].visible = False
        ws.column_dimensions[get_column_letter(2)].visible = False
        if hide_column_placement:
            ws.column_dimensions[get_column_letter(3)].visible = False
        if hide_column_producer:
            ws.column_dimensions[get_column_letter(4)].visible = False
    translation.activate(language_code)
    return wb


def generate_customer_xlsx(permanence, deliveries_id=None, customer=None, group=False):
    if customer is not None:
        wb = export_customer(
            permanence=permanence, customer=customer, xlsx_formula=False, wb=None
        )
        return wb, None
    else:
        wb = export_abstract(
            permanence=permanence, deliveries_id=deliveries_id, group=group, wb=None
        )
        if wb is not None:
            # At least one order
            abstract_ws = wb.get_active_sheet()
            ws_preparation_title = cap("{}".format(_("Preparation")), 31)
            ws_customer_title = cap("{}".format(_('Customer check')), 31)
            wb = export_customer_label(
                permanence=permanence, deliveries_id=deliveries_id, wb=wb
            )
            wb = export_preparation(
                permanence=permanence, deliveries_id=deliveries_id, wb=wb
            )
            wb = export_customer(
                permanence=permanence, deliveries_id=deliveries_id, deposit=True, wb=wb
            )
            wb = export_customer(
                permanence=permanence, deliveries_id=deliveries_id, deposit=False, wb=wb,
                ws_preparation_title=ws_preparation_title
            )
            wb = export_permanence_stock(
                permanence=permanence, deliveries_id=deliveries_id, customer_price=True, wb=wb,
                ws_customer_title=ws_customer_title
            )
            return wb, abstract_ws
        else:
            return


def generate_producer_xlsx(permanence, producer=None, wb=None):
    wb = export_producer_by_product(
        permanence=permanence, producer=producer, wb=wb
    )
    if not (wb is None or producer.manage_replenishment):
        # At least one order and we don't manage replenishment for this producer
        wb = export_producer_by_customer(
            permanence=permanence, producer=producer, wb=wb
        )
    return wb
