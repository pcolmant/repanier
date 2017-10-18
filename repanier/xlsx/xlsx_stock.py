# -*- coding: utf-8
from __future__ import unicode_literals

from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.utils import translation
from django.utils.translation import ugettext_lazy as _
from openpyxl import load_workbook
from openpyxl.style import Fill
from openpyxl.styles import Color

import repanier.apps
from repanier.xlsx.export_tools import *
from repanier.const import *
from repanier.models.offeritem import OfferItem
from repanier.models.product import Product
from repanier.tools import update_offer_item, next_row
from repanier.xlsx.import_tools import get_row, get_header


def export_permanence_stock(permanence, deliveries_id=None, customer_price=False, wb=None, ws_customer_title=None):
    if settings.DJANGO_SETTINGS_STOCK and wb is not None:
        yellowFill = Fill()
        yellowFill.start_color.index = 'FFEEEE11'
        yellowFill.end_color.index = 'FFEEEE11'
        yellowFill.fill_type = Fill.FILL_SOLID

        header = [
            (_("Id"), 5),
            (_("OfferItem"), 5),
            (_("Reference"), 20),
            (_("Product"), 60),
            (_("Customer unit price") if customer_price else _("Producer unit price"), 10),
            (_("Deposit"), 10),
            (_("Asked"), 10),
            (_("Quantity ordered"), 10),
            (_("Initial stock"), 10),
            (repanier.apps.REPANIER_SETTINGS_CURRENCY_DISPLAY, 15),
            (_("Stock used"), 10),
            (_("Additional"), 10),
            (_("Final stock"), 10),
            (repanier.apps.REPANIER_SETTINGS_CURRENCY_DISPLAY, 15),
        ]
        offer_items = OfferItem.objects.filter(
            Q(
                permanence_id=permanence.id,
                manage_replenishment=True,
                translations__language_code=translation.get_language()
            ) |
            Q(
                permanence_id=permanence.id,
                manage_production=True,
                translations__language_code=translation.get_language()
            )
        ).order_by(
            "producer",
            "translations__long_name",
            "order_average_weight",
        ).select_related(
            'producer', 'department_for_customer'
        ).iterator()
        offer_item = next_row(offer_items)
        if offer_item is not None:
            # Check if there are deliveries_ws
            deliveries_ws = []
            if deliveries_id is not None:
                for delivery_cpt, delivery_id in enumerate(deliveries_id):
                    ws_sc_name = cap('%d-%s' % (delivery_cpt, ws_customer_title), 31)
                    for sheet in wb.worksheets:
                        if ws_sc_name == sheet.title:
                            deliveries_ws.append(ws_sc_name)
                            break
            else:
                ws_sc_name = cap(ws_customer_title, 31)
                for sheet in wb.worksheets:
                    if ws_sc_name == sheet.title:
                        deliveries_ws.append(ws_sc_name)
                        break
            wb, ws = new_landscape_a4_sheet(
                wb,
                _('Stock check'),
                permanence,
                header
            )
            formula_main_total_a = []
            formula_main_total_b = []
            show_column_reference = False
            show_column_qty_ordered = False
            show_column_add2stock = False
            row_num = 1
            while offer_item is not None:
                producer_save = offer_item.producer
                row_start_producer = row_num + 1
                c = ws.cell(row=row_num, column=2)
                c.value = producer_save.short_profile_name
                c.style.font.bold = True
                c.style.font.italic = True
                while offer_item is not None and producer_save.id == offer_item.producer_id:
                    department_for_customer_save__id = offer_item.department_for_customer_id
                    department_for_customer_save__short_name = offer_item.department_for_customer.short_name \
                        if offer_item.department_for_customer is not None else None
                    while offer_item is not None and producer_save.id == offer_item.producer_id \
                            and department_for_customer_save__id == offer_item.department_for_customer_id:
                        if len(offer_item.reference) < 36:
                            if offer_item.reference.isdigit():
                                # Avoid display of exponent by Excel
                                offer_item_reference = '[%s]' % offer_item.reference
                            else:
                                offer_item_reference = offer_item.reference
                            show_column_reference = True
                        else:
                            offer_item_reference = EMPTY_STRING
                        if offer_item.order_unit < PRODUCT_ORDER_UNIT_DEPOSIT:

                            asked = offer_item.quantity_invoiced - offer_item.add_2_stock
                            stock = offer_item.stock
                            add_2_stock = offer_item.add_2_stock
                            c = ws.cell(row=row_num, column=0)
                            c.value = offer_item.producer_id
                            c = ws.cell(row=row_num, column=1)
                            c.value = offer_item.id
                            c = ws.cell(row=row_num, column=2)
                            c.value = offer_item_reference
                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=3)
                            if department_for_customer_save__short_name is not None:
                                c.value = "%s - %s" % (
                                    offer_item.get_long_name(),
                                    department_for_customer_save__short_name
                                )
                            else:
                                c.value = "%s" % offer_item.get_long_name()
                            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                            c.style.alignment.wrap_text = True
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=4)
                            unit_price = offer_item.customer_unit_price if customer_price else offer_item.producer_unit_price
                            c.value = unit_price.amount
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=5)
                            c.value = offer_item.unit_deposit.amount
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=6)
                            if ws_customer_title is None:
                                c.value = asked
                            else:
                                if len(deliveries_ws) > 0:
                                    #     # Without any deleveriy point
                                    #     sum_value = "SUMIF('%s'!B2:B5000,B%s,'%s'!F2:F5000)" % \
                                    #         (ws_customer_title, row_num + 1, ws_customer_title)
                                    # else:
                                    sum_value = "+".join("SUMIF('%s'!B2:B5000,B%s,'%s'!F2:F5000)" % \
                                                         (delivery_ws, row_num + 1, delivery_ws)
                                                         for delivery_ws in deliveries_ws
                                                         )
                                    c.value = "=%s" % sum_value
                                else:
                                    c.value = DECIMAL_ZERO
                            c.style.number_format.format_code = '#,##0.???'
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=7)
                            c.value = '=G%s-K%s+L%s' % (row_num + 1, row_num + 1, row_num + 1)
                            if not show_column_qty_ordered:
                                show_column_qty_ordered = (asked - min(asked, stock) + add_2_stock) > 0
                            c.style.number_format.format_code = '#,##0.???'
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=8)
                            c.value = stock
                            c.style.number_format.format_code = '#,##0.???'
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c.style.font.color = Color(Color.BLUE)
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(9) + str(row_num + 1), 'notEqual',
                                [str(stock)], True, wb,
                                None, None, yellowFill
                            )
                            c = ws.cell(row=row_num, column=9)
                            c.value = '=ROUND(I%s*(E%s+F%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=10)
                            c.value = '=MIN(G%s,I%s)' % (row_num + 1, row_num + 1)
                            c.style.number_format.format_code = '#,##0.???'
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c = ws.cell(row=row_num, column=11)
                            c.value = add_2_stock
                            c.style.number_format.format_code = '#,##0.???'
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c.style.font.color = Color(Color.BLUE)
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(12) + str(row_num + 1), 'notEqual',
                                [str(add_2_stock)], True, wb,
                                None, None, yellowFill
                            )
                            if not show_column_add2stock:
                                show_column_add2stock = add_2_stock > 0
                            c = ws.cell(row=row_num, column=12)
                            c.value = '=I%s-K%s+L%s' % (row_num + 1, row_num + 1, row_num + 1)
                            c.style.number_format.format_code = '#,##0.???'
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            c.style.font.bold = True
                            c = ws.cell(row=row_num, column=13)
                            c.value = '=ROUND(M%s*(E%s+F%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                            c.style.borders.bottom.border_style = Border.BORDER_THIN
                            row_num += 1
                        offer_item = next_row(offer_items)
                row_num += 1
                c = ws.cell(row=row_num, column=3)
                c.value = "%s %s" % (_("Total price"), producer_save.short_profile_name)
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                c.style.font.bold = True
                c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
                c = ws.cell(row=row_num, column=9)
                formula = 'SUM(J%s:J%s)' % (row_start_producer, row_num)
                c.value = '=' + formula
                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                c.style.font.bold = True
                formula_main_total_a.append(formula)
                c = ws.cell(row=row_num, column=13)
                formula = 'SUM(N%s:N%s)' % (row_start_producer, row_num)
                c.value = '=' + formula
                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                c.style.font.bold = True
                formula_main_total_b.append(formula)

                if offer_items is not None:
                    # Display a separator line between producers
                    row_num += 1
                    for col_num in range(16):
                        c = ws.cell(row=row_num, column=col_num)
                        c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
                    row_num += 2

            c = ws.cell(row=row_num, column=3)
            c.value = "%s" % _("Total price")
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            c.style.font.bold = True
            c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
            c = ws.cell(row=row_num, column=9)
            c.value = "=" + "+".join(formula_main_total_a)
            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
            c.style.font.bold = True
            c = ws.cell(row=row_num, column=13)
            c.value = "=" + "+".join(formula_main_total_b)
            c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
            c.style.font.bold = True

            row_num += 1
            for col_num in range(16):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED

            ws.column_dimensions[get_column_letter(1)].visible = False
            ws.column_dimensions[get_column_letter(2)].visible = False
            ws.column_dimensions[get_column_letter(11)].visible = False
            if not show_column_reference:
                ws.column_dimensions[get_column_letter(3)].visible = False
            if not show_column_qty_ordered:
                ws.column_dimensions[get_column_letter(8)].visible = False
            if not show_column_add2stock:
                ws.column_dimensions[get_column_letter(12)].visible = False
    return wb


# @transaction.atomic
# def import_stock_sheet(worksheet, permanence=None):
#     error = False
#     error_msg = None
#     if permanence.status < PERMANENCE_DONE:
#         header = get_header(worksheet)
#         if header:
#             row_num = 1
#             row = get_row(worksheet, header, row_num)
#             while row and not error:
#                 try:
#                     # with transaction.atomic():
#                     stock = None if row[_('Initial stock')] is None else Decimal(row[_('Initial stock')]).quantize(THREE_DECIMALS)
#                     add_2_stock = None if row[_('Add 2 stock')] is None else Decimal(row[_('Add 2 stock')]).quantize(THREE_DECIMALS)
#                     if stock is not None:
#                         producer_id = None if row[_('Id')] is None else Decimal(row[_('Id')])
#                         offer_item_id = None if row[_('OfferItem')] is None else Decimal(row[_('OfferItem')])
#                         offer_item = OfferItem.objects.filter(
#                             id=offer_item_id,
#                             permanence_id=permanence.id,
#                             producer_id=producer_id
#                         ).order_by('?').first()
#                         if offer_item is not None \
#                                 and (offer_item.stock != stock or offer_item.add_2_stock != add_2_stock):
#                             offer_item.stock = stock
#                             offer_item.add_2_stock = add_2_stock
#                             offer_item.save()
#                             Product.objects.filter(
#                                 id=offer_item.product_id,
#                                 producer_id=producer_id
#                             ).update(stock=stock)
#                     row_num += 1
#                     row = get_row(worksheet, header, row_num)
#                 except KeyError, e:
#                     # Missing field
#                     error = True
#                     error_msg = _("Row %(row_num)d : A required column is missing.") % {'row_num': row_num + 1}
#                 except Exception, e:
#                     error = True
#                     error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
#     else:
#         error = True
#         error_msg = _("The status of this permanence prohibit you to update the stock.")
#     return error, error_msg


def export_producer_stock(producers, customer_price=False, wb=None):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Id"), 5),
        (_("Producer"), 60),
        (_("Reference"), 20),
        (_("Product"), 60),
        (_("Customer unit price") if customer_price else _("Producer unit price"), 10),
        (_("Deposit"), 10),
        (_("Current stock"), 10),
        (repanier.apps.REPANIER_SETTINGS_CURRENCY_DISPLAY, 15),
    ]
    producers = producers.iterator()
    producer = next_row(producers)
    wb, ws = new_landscape_a4_sheet(
        wb,
        _('Current stock'),
        _('Current stock'),
        header
    )
    show_column_reference = False
    row_num = 1
    while producer is not None:
        products = Product.objects.filter(
            producer_id=producer.id,
            is_active=True,
            translations__language_code=translation.get_language()
        ).order_by(
            "translations__long_name",
            "order_average_weight",
        ).select_related(
            'producer', 'department_for_customer'
        ).iterator()
        product = next_row(products)
        while product is not None:
            if product.order_unit < PRODUCT_ORDER_UNIT_DEPOSIT:
                c = ws.cell(row=row_num, column=0)
                c.value = product.id
                c = ws.cell(row=row_num, column=1)
                c.value = "%s" % product.producer
                if len(product.reference) < 36:
                    if product.reference.isdigit():
                        # Avoid display of exponent by Excel
                        product_reference = '[%s]' % product.reference
                    else:
                        product_reference = product.reference
                    show_column_reference = True
                else:
                    product_reference = EMPTY_STRING
                c = ws.cell(row=row_num, column=2)
                c.value = product_reference
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=3)
                if product.department_for_customer is not None:
                    c.value = "%s - %s" % (
                        product.get_long_name(),
                        product.department_for_customer.short_name
                    )
                else:
                    c.value = product.get_long_name()
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                c.style.alignment.wrap_text = True
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=4)
                unit_price = product.customer_unit_price if customer_price else product.producer_unit_price
                c.value = unit_price.amount
                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=5)
                c.value = product.unit_deposit.amount
                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=6)
                c.value = product.stock
                c.style.number_format.format_code = '_ * #,##0.00_ ;_ * -#,##0.00_ ;_ * "-"??_ ;_ @_ '
                c.style.font.color = Color(Color.BLUE)
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=7)
                c.value = '=ROUND((E%s+F%s)*G%s,2)' % (row_num + 1, row_num + 1, row_num + 1)
                c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
                ws.conditional_formatting.addCellIs(
                    get_column_letter(8) + str(row_num + 1), 'notEqual',
                    [str(((unit_price.amount + product.unit_deposit.amount) * product.stock).quantize(TWO_DECIMALS))],
                    True, wb,
                    None, None, yellowFill
                )
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                row_num += 1
            product = next_row(products)
        row_num += 1
        c = ws.cell(row=row_num, column=4)
        c.value = "%s" % (_("Total"),)
        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
        c.style.font.bold = True
        c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
        c = ws.cell(row=row_num, column=7)
        formula = 'SUM(H%s:H%s)' % (2, row_num)
        c.value = '=' + formula
        c.style.number_format.format_code = repanier.apps.REPANIER_SETTINGS_CURRENCY_XLSX
        c.style.font.bold = True

        ws.column_dimensions[get_column_letter(1)].visible = False
        if not show_column_reference:
            ws.column_dimensions[get_column_letter(3)].visible = False
        producer = next_row(producers)
    return wb


@transaction.atomic
def import_producer_stock(worksheet, producers=None):
    error = False
    error_msg = None
    header = get_header(worksheet)
    if header:
        row_num = 1
        row = get_row(worksheet, header, row_num)
        while row and not error:
            try:
                # with transaction.atomic():
                product_id = None if row[_('Id')] is None else Decimal(row[_('Id')])
                if product_id is not None:
                    stock = DECIMAL_ZERO if row[_('Current stock')] is None else Decimal(
                        row[_('Current stock')]).quantize(
                        THREE_DECIMALS)
                    stock = stock if stock >= DECIMAL_ZERO else DECIMAL_ZERO
                    Product.objects.filter(
                        id=product_id, producer__in=producers
                    ).update(stock=stock)
                update_offer_item(product_id=product_id)
                row_num += 1
                row = get_row(worksheet, header, row_num)
            except KeyError as e:
                # Missing field
                error = True
                error_msg = _("Row %(row_num)d : A required column is missing.") % {'row_num': row_num + 1}
            except Exception as e:
                error = True
                error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
    return error, error_msg


def handle_uploaded_stock(request, producers, file_to_import, *args):
    error = False
    error_msg = None
    wb = load_workbook(file_to_import)
    if wb is not None:
        ws = wb.get_sheet_by_name(cap(slugify("%s" % _('Current stock')), 31))
        if ws is None:
            ws = wb.get_sheet_by_name(cap("%s" % _('Current stock'), 31))
        if ws is not None:
            error, error_msg = import_producer_stock(
                ws,
                producers=producers
            )
            if error:
                error_msg = cap("%s" % _('Current stock'), 31) + " > " + error_msg
    return error, error_msg
