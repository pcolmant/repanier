# -*- coding: utf-8
from __future__ import unicode_literals

from django.contrib.sites.models import Site
from django.utils.text import slugify
from django.utils.translation import ugettext_lazy as _
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat
from openpyxl.styles import Border
from openpyxl.workbook import Workbook

from repanier.xlsx.const import *
from repanier.const import EMPTY_STRING
from repanier.models.staff import Staff
from repanier.tools import cap


def worksheet_setup_a4(worksheet, title1, title2, add_print_title=True):
    title1 = "%s" % title1
    worksheet.title = cap(slugify(title1), 31)
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.fitToWidth = 1
    worksheet.print_gridlines = True
    if add_print_title:
        worksheet.add_print_title(1, rows_or_cols='rows')
        try:
            # Python 2
            worksheet.freeze_panes = 'A2'.encode("utf8")
        except:
            # Python 3
            worksheet.freeze_panes = 'A2'
    worksheet.header_footer.left_header.text = Site.objects.get_current().name
    worksheet.header_footer.left_footer.text = "%s" % (title2)
    worksheet.header_footer.center_footer.text = title1
    worksheet.header_footer.right_footer.text = 'Page &[Page]/&[Pages]'
    orders_responsible = Staff.objects.filter(is_reply_to_order_email=True, is_active=True).order_by('?').first()
    invoices_responsible = Staff.objects.filter(is_reply_to_invoice_email=True, is_active=True).order_by('?').first()
    s1 = EMPTY_STRING
    if orders_responsible:
        c = orders_responsible.customer_responsible
        if c is not None:
            s1 = "%s: %s, %s" % (_("Orders"), c.long_basket_name, c.phone1)
    s2 = EMPTY_STRING
    if invoices_responsible:
        c = invoices_responsible.customer_responsible
        if c is not None:
            s2 = "%s: %s, %s" % (_("Invoices"), c.long_basket_name, c.phone1)
    separator = chr(10) + " "
    worksheet.header_footer.right_header.text = separator.join((s1, s2))
    return worksheet


def worksheet_setup_portrait_a4(worksheet, title1, title2, add_print_title=True):
    worksheet = worksheet_setup_a4(worksheet, title1, title2, add_print_title)
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    return worksheet


def new_portrait_a4_sheet(workbook, title1, title2, header=None, add_print_title=True):
    if workbook is None:
        workbook = Workbook()
        worksheet = workbook.get_active_sheet()
    else:
        worksheet = workbook.create_sheet()
    worksheet = worksheet_setup_portrait_a4(worksheet, title1, title2, add_print_title=add_print_title)
    if header is not None:
        worksheet_set_header(worksheet, header)
    return workbook, worksheet


def worksheet_setup_landscape_a4(worksheet, title1, title2, add_print_title=True):
    worksheet = worksheet_setup_a4(worksheet, title1, title2, add_print_title)
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    return worksheet


def new_landscape_a4_sheet(workbook, title1, title2, header=None, add_print_title=True):
    if workbook is None:
        workbook = Workbook()
        worksheet = workbook.get_active_sheet()
    else:
        worksheet = workbook.create_sheet()
    worksheet = worksheet_setup_landscape_a4(worksheet, title1, title2, add_print_title=add_print_title)
    if header is not None:
        worksheet_set_header(worksheet, header)
    return workbook, worksheet


def worksheet_set_header(worksheet, header):
    for col_num in range(len(header)):
        c = worksheet.cell(row=0, column=col_num)
        c.value = (header[col_num][ROW_TITLE]).encode("utf8")
        c.style.font.bold = True
        c.style.alignment.wrap_text = False
        c.style.borders.bottom.border_style = Border.BORDER_THIN
        worksheet.column_dimensions[get_column_letter(col_num + 1)].width = header[col_num][ROW_WIDTH]
        if header[col_num][ROW_TITLE] == _("Id"):
            worksheet.column_dimensions[get_column_letter(col_num + 1)].visible = False


def get_validation_formula(wb=None, valid_values=None):
    if valid_values:

        ws_dv_name = cap(slugify("%s" % (_("data validation"))), 31)
        ws_dv = wb.get_sheet_by_name(ws_dv_name)
        if ws_dv is None:
            ws_dv = wb.create_sheet(index=0)
            worksheet_setup_landscape_a4(ws_dv, ws_dv_name, EMPTY_STRING)
        col_dv = 0
        c = ws_dv.cell(row=0, column=col_dv)
        while (c.value is not None) and (col_dv < 20):
            col_dv += 1
            c = ws_dv.cell(row=0, column=col_dv)

        col_letter_dv = get_column_letter(col_dv + 1)
        row_num = 0
        for v in valid_values:
            c = ws_dv.cell(row=row_num, column=col_dv)
            c.value = "%s" % (v)
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            row_num += 1
        return "'%s'!$%s$1:$%s$%s" % (ws_dv_name, col_letter_dv, col_letter_dv, row_num + 1)
    else:
        return EMPTY_STRING
