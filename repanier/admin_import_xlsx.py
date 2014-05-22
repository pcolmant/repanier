# -*- coding: utf-8 -*-
from const import *
from tools import *
from decimal import *
from django import forms
from django.http import HttpResponseRedirect
from django.db.models import F
from django.utils.translation import ugettext_lazy as _
from django.contrib import messages
from views import render_response
from openpyxl import load_workbook
from repanier.views import render_response

import datetime

from repanier.models import Producer
from repanier.models import Product
from repanier.models import LUT_DepartmentForCustomer
from repanier.models import LUT_ProductionMode
from repanier.models import OfferItem
from repanier.models import Purchase
from repanier.models import Customer
# from repanier.models import BankAccount
from repanier.models import Permanence

class ImportXlsxForm(forms.Form):
	_selected_action = forms.CharField(widget=forms.MultipleHiddenInput)
	file_to_import = forms.FileField(
		label=_('File to import'), 
		allow_empty_file=False
	)

def get_header(worksheet):
	header = []
	if worksheet:
		row_num = 0
		col_num = 0
		c = worksheet.cell(row=row_num, column=col_num)
		while (c.value!=None ) and (col_num < 50):
			header.append(c.value)
			col_num+=1
			c = worksheet.cell(row=row_num, column=col_num)
	return header

def get_row(worksheet, header, row_num):
	row = {}
	if worksheet:
		# last_row is a row with all cells empty
		last_row = True
		for col_num, col_header in enumerate(header):
			c = worksheet.cell(row=row_num, column=col_num)
			# Important c.value==0 : Python (or Python lib) mix 0 and None
			if c.value or c.value==0:
				last_row = False
			row[col_header] = None if c.data_type == c.TYPE_FORMULA else c.value
		if last_row:
			row = {}
	return row

def import_producer_products(worksheet, producer = None, db_write=False,
		department_for_customer_2_id_dict = None,
		production_mode_2_id_dict = None
	):
	vat_level_dict = dict(LUT_VAT_REVERSE)
	order_unit_dict = dict(LUT_PRODUCT_ORDER_UNIT_REVERSE)
	error = False
	error_msg = None
	header = get_header(worksheet)
	if header:
		row_num = 1
		row = get_row(worksheet, header, row_num)
		while row and not error:
			try:
				row_id = None
				if row[_('Id')] != None:
					row_id = Decimal(row[_('Id')])
				# product_reorder += 1

				department_for_customer_id = None
				if row[_('department_for_customer')] in department_for_customer_2_id_dict:
					department_for_customer_id = department_for_customer_2_id_dict[row[_('department_for_customer')]]
				else:
					error = True
					error_msg = _("Row %(row_num)d : No valid departement for customer") % {'row_num': row_num + 1}
					break

				production_mode_id = None
				if row[_('production mode')] in production_mode_2_id_dict:
					production_mode_id = production_mode_2_id_dict[row[_('production mode')]]
				else:
					error = True
					error_msg = _("Row %(row_num)d : No valid production mode") % {'row_num': row_num + 1}
					break

				original_unit_price = None if row[_('original_unit_price')] == None else Decimal(row[_('original_unit_price')])
				unit_deposit = None if row[_('deposit')] == None else Decimal(row[_('deposit')])
				order_average_weight = None if row[_('order_average_weight')] == None else Decimal(row[_('order_average_weight')])
				customer_minimum_order_quantity = None if row[_('customer_minimum_order_quantity')] == None else Decimal(row[_('customer_minimum_order_quantity')])
				customer_increment_order_quantity = None if row[_('customer_increment_order_quantity')] == None else Decimal(row[_('customer_increment_order_quantity')])
				customer_alert_order_quantity = None if row[_('customer_alert_order_quantity')] == None else Decimal(row[_('customer_alert_order_quantity')])

				order_unit = None
				if row[_("order unit")] in order_unit_dict:
					order_unit = order_unit_dict[row[_("order unit")]]
				if order_unit == None:
					order_unit = PRODUCT_ORDER_UNIT_LOOSE_PC

				long_name = cap(row[_('long_name')], 100)
				product_set = Product.objects.filter(
					producer_id = producer.id,
					long_name = long_name
				).order_by()[:1]
				# print(long_name.encode('utf8'))
				if product_set:
					product = product_set[0]
					if row_id == product.id:
						# Detect VAT LEVEL. Fall back on product.
						vat_level = None
						if row[_("vat or compensation")] in vat_level_dict:
							vat_level = vat_level_dict[row[_("vat or compensation")]]
						elif product != None:
							vat_level = product.vat_level
						else:
							vat_level = producer.vat_level
						# Let only update if the given id is the same as the product found id
						product.producer_id = producer.id
						product.long_name = long_name
						product.production_mode_id = production_mode_id
						product.department_for_customer_id = department_for_customer_id
						product.order_unit = order_unit
						if order_average_weight != None:
							product.order_average_weight = order_average_weight
						if original_unit_price != None:
							product.original_unit_price = original_unit_price
						if unit_deposit != None:
							product.unit_deposit = unit_deposit
						if customer_minimum_order_quantity != None:
							product.customer_minimum_order_quantity = customer_minimum_order_quantity
						if customer_increment_order_quantity != None:
							product.customer_increment_order_quantity = customer_increment_order_quantity
						if customer_alert_order_quantity != None:
							product.customer_alert_order_quantity = customer_alert_order_quantity
						product.is_into_offer = ( row[_('is_into_offer')] != None )
						product.vat_level = vat_level
						product.is_active = True
						# product.product_reorder = product_reorder
						if db_write:
							product.save()
					else:
						error = True
						if row[_('Id')]==None:
							error_msg = _("Row %(row_num)d : No id given, or the product %(producer)s - %(product)s already exist.") % {'row_num': row_num + 1, 'producer': producer.short_profile_name, 'product': row[_('long_name')]}
						else:
							error_msg = _("Row %(row_num)d : The given id %(record_id)s is not the id of %(producer)s - %(product)s.") % {'row_num': row_num + 1, 'record_id':row[_('Id')], 'producer': producer.short_profile_name, 'product': row[_('long_name')]}
						break
				else:
					if row_id == None:
						# Let only create product if non id in the row
						if db_write:
							# Detect VAT LEVEL. Can't fall back on product.
							vat_level = None
							if row[_("vat or compensation")] in vat_level_dict:
								vat_level = vat_level_dict[row[_("vat or compensation")]]
							else:
								vat_level = producer.vat_level

							if order_average_weight == None:
								order_average_weight = 0
							if original_unit_price == None:
								original_unit_price = 0
							if unit_deposit == None:
								unit_deposit = 0
							if customer_minimum_order_quantity == None:
								customer_minimum_order_quantity = 0
							if customer_increment_order_quantity == None:
								customer_increment_order_quantity = 0
							if customer_alert_order_quantity == None:
								customer_alert_order_quantity = 0
							Product.objects.create(
								producer = producer,
								long_name = long_name,
								production_mode_id = production_mode_id,
								department_for_customer_id = department_for_customer_id,
								order_unit = order_unit,
								order_average_weight = order_average_weight,
								original_unit_price = original_unit_price,
								unit_deposit = unit_deposit,
								customer_minimum_order_quantity = customer_minimum_order_quantity,
								customer_increment_order_quantity = customer_increment_order_quantity,
								customer_alert_order_quantity = customer_alert_order_quantity,
								is_into_offer = ( row[_('is_into_offer')] != None ),
								is_active = True,
								# product_reorder = product_reorder
							)
					else:
						error = True
						error_msg = _("Row %(row_num)d : The given id %(record_id)s is not the id of %(producer)s - %(product)s.") % {'row_num': row_num + 1, 'record_id':row[_('Id')], 'producer': producer.short_profile_name, 'product': row[_('long_name')]}
						break

				row_num += 1
				row = get_row(worksheet, header, row_num)
			except KeyError, e:
				# Missing field
				error = True
				error_msg = _("Row %(row_num)d : A required column is missing.") % {'row_num': row_num + 1}
			except Exception, e:
				error = True
				error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg':str(e)}
	return error, error_msg

def handle_product_uploaded_file(request, queryset, file_to_import):
	error = False
	error_msg = None
	wb = load_workbook(file_to_import)
	# dict for performance optimisation purpose : read the DB only once
	department_for_customer_2_id_dict=get_department_for_customer_2_id_dict()
	production_mode_2_id_dict=get_production_mode_2_id_dict()
	# product_reorder = 0
	for producer in queryset:
		error, error_msg = import_producer_products(wb.get_sheet_by_name(unicode(cap(producer.short_profile_name,31), "utf8")), producer=producer, db_write=False,
				department_for_customer_2_id_dict=department_for_customer_2_id_dict,
				production_mode_2_id_dict=production_mode_2_id_dict
			)
		if error:
			error_msg = producer.short_profile_name + " > " + error_msg
			break
	if not error:
		# Product.objects.all().update(product_reorder=F('product_order') + 100000)
		# product_reorder = 0
		for producer in queryset:
			error_flag, error_msg  = import_producer_products(wb.get_sheet_by_name(unicode(cap(producer.short_profile_name,31), "utf8")), producer=producer, db_write=True,
					department_for_customer_2_id_dict=department_for_customer_2_id_dict,
					production_mode_2_id_dict=production_mode_2_id_dict
				)
			if error:
				error_msg = producer.short_profile_name + " > " + error_msg
				break
		# order = 0
		# for obj in Product.objects.all().order_by('producer__short_profile_name', 'product_reorder'):
		# 	order += 1
		# 	obj.product_order = order
		# 	obj.save(update_fields=['product_order'])
		for permanence in Permanence.objects.filter(status=PERMANENCE_OPENED):
			recalculate_order_amount(permanence.id)
	return error, error_msg

def import_product_xlsx(producer, admin, request, queryset):
	form = None
	error_msg = None
	if 'apply' in request.POST:
		form = ImportXlsxForm(request.POST, request.FILES)
		if form.is_valid():
			file_to_import = request.FILES['file_to_import']
			if('.xlsx' in file_to_import.name) and (file_to_import.size <= 1000000):
				error, error_msg = handle_product_uploaded_file(request, queryset, file_to_import)
				if error:
					if error_msg == None:
						producer.message_user(request, 
							_("Error when importing %s : Content not valid") % (file_to_import.name),
							level=messages.WARNING
						)
					else:
						producer.message_user(request, 
							_("Error when importing %(file_name)s : %(error_msg)s") % {'file_name': file_to_import.name, 'error_msg':error_msg},
							level=messages.ERROR
						)
				else:
					producer.message_user(request, _("Successfully imported %s.") % (file_to_import.name))
			else:
				producer.message_user(request, 
					_("Error when importing %s : File size must be <= 1 Mb and extension must be .xlsx") % (file_to_import.name),
					level=messages.ERROR
				)
		else:
			producer.message_user(request, _("No file to import."),level=messages.WARNING)
		return HttpResponseRedirect(request.get_full_path())
	elif 'cancel' in request.POST:
		producer.message_user(request, _("Action canceled by the user."),level=messages.WARNING)
		return HttpResponseRedirect(request.get_full_path())
	form = ImportXlsxForm(
		initial={'_selected_action': request.POST.getlist(admin.ACTION_CHECKBOX_NAME)}
	)
	return render_response(request, 'repanier/import_xlsx.html', {
		'title': _("Import products"),
		'objects': queryset,
		'import_xlsx_form': form,
	})


def import_permanence_purchases(worksheet, permanence = None, db_write = False,
		customer_2_id_dict=None,
		id_2_customer_vat_id_dict=None,
		producer_2_id_dict=None,
		id_2_producer_vat_level_dict=None,
		id_2_producer_price_list_multiplier_dict=None
	):
	vat_level_dict = dict(LUT_VAT_REVERSE)
	order_unit_dict = dict(LUT_PRODUCT_ORDER_UNIT_REVERSE)
	error = False
	error_msg = None
	header = get_header(worksheet)
	if header:
		row_num = 1
		array_purchase = []
		new_invoiced = None
		row = get_row(worksheet, header, row_num)
		while row and not error:
			# print(str(row[_('Id')]))
			try:
				if row[_('producer')] == None and row[_('product')] == None and row[_('customer')] == None:
					if db_write:
						max_purchase_counter = len(array_purchase)
						if max_purchase_counter > 1:
							old_invoiced = None if row[_('Id')] == None else Decimal(row[_('Id')])
							if old_invoiced == None:
								error = True
								error_msg = _("Row %(row_num)d : No purchase invoice given.") % {'row_num': row_num + 1}
								break
							producer_id = None
							actual_invoice = DECIMAL_ZERO
							invoice_by_basket = None
							for i, purchase in enumerate(array_purchase):
								if i == 0:
									producer_id = purchase.producer_id
									invoice_by_basket = purchase.producer.invoice_by_basket
								else:
									if producer_id != purchase.producer_id:
										error = True
										error_msg = _("Row %(row_num)d : The system cannot combine purchases of different producers.") % {'row_num': row_num + 1}
										break
								if invoice_by_basket:
									actual_invoice += purchase.original_price
								else:
									actual_invoice += purchase.quantity
							if error:
								break
							if invoice_by_basket == None:
								error = True
								error_msg = _("Row %(row_num)d : The system cannot determine if purchases are invoiced by basket or not.") % {'row_num': row_num + 1}
								break

							if new_invoiced != None:
								ratio = DECIMAL_ONE
								if actual_invoice != DECIMAL_ZERO:
									ratio = new_invoiced / actual_invoice
								else:
									if new_invoiced == DECIMAL_ZERO:
										ratio = DECIMAL_ZERO
									else:
										error = True
										error_msg = _("Row %(row_num)d : The actual invoiced amount is zero and you want to distribute more than zero.") % {'row_num': row_num + 1}
										break
								# Rule of 3
								adjusted_invoice = 0
								for i, purchase in enumerate(array_purchase, start=1):
									if i == max_purchase_counter:
										if invoice_by_basket:
											purchase.original_price = new_invoiced - adjusted_invoice
											if ( purchase.original_unit_price + purchase.unit_deposit ) != DECIMAL_ZERO:
												purchase.quantity = purchase.original_price / ( purchase.original_unit_price + purchase.unit_deposit )
										else:
											purchase.quantity = new_invoiced - adjusted_invoice
											purchase.original_price = purchase.quantity * ( purchase.original_unit_price + purchase.unit_deposit )
									else:
										if invoice_by_basket:
											purchase.original_price *= ratio
											purchase.original_price = purchase.original_price.quantize(DECIMAL_0_01, rounding=ROUND_UP)
											adjusted_invoice += purchase.original_price
											if ( purchase.original_unit_price + purchase.unit_deposit ) != DECIMAL_ZERO:
												purchase.quantity = purchase.original_price / ( purchase.original_unit_price + purchase.unit_deposit )
										else:
											purchase.quantity *= ratio
											purchase.quantity = purchase.quantity.quantize(DECIMAL_0_0001, rounding=ROUND_HALF_UP)
											adjusted_invoice += purchase.quantity
											purchase.original_price = purchase.quantity * ( purchase.original_unit_price + purchase.unit_deposit )
						# Adjust tax and save updated purchase
						price_list_multiplier = 1
						if producer_id in id_2_producer_price_list_multiplier_dict:
							price_list_multiplier = id_2_producer_price_list_multiplier_dict[producer_id]
						for purchase in array_purchase:

							unit_price_with_vat = (purchase.original_unit_price * price_list_multiplier).quantize(DECIMAL_0_01, rounding=ROUND_UP)
							purchase.price_with_vat = purchase.quantity * ( unit_price_with_vat + purchase.unit_deposit ).quantize(DECIMAL_0_01, rounding=ROUND_UP)
							unit_price_with_compensation = unit_price_with_vat
							if purchase.vat_level == VAT_200:
								unit_price_with_compensation = (unit_price_with_vat * Decimal(1.02)).quantize(DECIMAL_0_01, rounding=ROUND_UP)
							elif purchase.vat_level == VAT_300:
								unit_price_with_compensation = (unit_price_with_vat * Decimal(1.06)).quantize(DECIMAL_0_01, rounding=ROUND_UP)
							purchase.price_with_compensation = purchase.quantity * ( unit_price_with_compensation + purchase.unit_deposit ).quantize(DECIMAL_0_01, rounding=ROUND_UP)

							purchase.invoiced_price_with_compensation = False
							if (purchase.vat_level in [VAT_200, VAT_300]) and (id_2_customer_vat_id_dict[purchase.customer_id] != None):
								purchase.invoiced_price_with_compensation = True
							purchase.save()

					sum_original_price = 0
					sum_quantity = 0
					array_purchase = []
					new_invoiced = None
				else:
					row_id = None
					if row[_('Id')] == None:
						error = True
						error_msg = _("Row %(row_num)d : No purchase id given.") % {'row_num': row_num + 1}
						break
					row_id = Decimal(row[_('Id')])

					purchase = None
					purchase_set = Purchase.objects.filter(id = row_id).order_by()[:1]
					if purchase_set:
						purchase = purchase_set[0]
					else:
						error = True
						error_msg = _("Row %(row_num)d : No purchase corresponding to the given purchase id.") % {'row_num': row_num + 1}
						break
					if purchase.permanence_id != permanence.id:
						error = True
						error_msg = _("Row %(row_num)d : The given permanence doesn't own the given purchase id.") % {'row_num': row_num + 1}
						break
					producer_id = None
					if row[_('producer')] in producer_2_id_dict:
						producer_id = producer_2_id_dict[row[_('producer')]]
					if producer_id != purchase.producer_id:
						error = True
						error_msg = _("Row %(row_num)d : No valid producer.") % {'row_num': row_num + 1}
						break
					# long_name = row[_('product')]
					# if purchase.product!= None and long_name != purchase.product.long_name:
					# 	error = True
					# 	error_msg = _("Row %(row_num)d : No valid product name.") % {'row_num': row_num + 1}
					# 	break
					customer_id = None
					if row[_('customer')] in customer_2_id_dict:
						customer_id = customer_2_id_dict[row[_('customer')]]
					if customer_id != purchase.customer_id:
						error = True
						error_msg = _("Row %(row_num)d : No valid customer") % {'row_num': row_num + 1}
						break
					vat_level = None
					if row[_("vat or compensation")] in vat_level_dict:
						vat_level = vat_level_dict[row[_("vat or compensation")]]
					else:
						error = True
						error_msg = _("Row %(row_num)d : No valid vat or compensation level") % {'row_num': row_num + 1}
						break

					# Q 
					quantity = DECIMAL_ZERO if row[_('quantity')] == None else Decimal(row[_('quantity')]).quantize(Decimal('.0001'), rounding=ROUND_HALF_DOWN)
					# PU 
					original_unit_price = DECIMAL_ZERO if row[_('original unit price')] == None else Decimal(row[_('original unit price')]).quantize(Decimal('.01'), rounding=ROUND_HALF_DOWN)
					# C 
					unit_deposit = DECIMAL_ZERO if row[_('deposit')] == None else Decimal(row[_('deposit')]).quantize(Decimal('.01'), rounding=ROUND_HALF_DOWN)
					# PL 
					original_price = DECIMAL_ZERO if row[_('original price')] == None  else Decimal(row[_('original price')]).quantize(Decimal('.01'), rounding=ROUND_HALF_UP)
					new_invoiced = None if row[_('invoiced')] in [None, " "] else Decimal(row[_('invoiced')])

					comment = cap(row[_('comment')], 100)

					if db_write:

						quantity_modified = quantity != purchase.quantity
						original_unit_price_modified = original_unit_price != purchase.original_unit_price
						unit_deposit_modified = unit_deposit != purchase.unit_deposit
						original_price_modified = original_price != purchase.original_price

						# A1	if (PU + C) != 0 then: Q = PL / (PU + C) else: Q = 1, PU = PL - C
						# A2	if Q != 0 then: PU = ( PL / Q ) - C else: PL = 0
						# A3	if Q != 0 then: C = ( PL / Q ) - PU else: PL = 0
						# A4	if (PU + C) != 0 then: Q = PL / (PU + C) else: Q = 1, C = PL - PU
						# A5	PL = Q * ( PU + C )
						# A6	Nothing

						if original_price_modified:
							if unit_deposit_modified:
								if quantity_modified and not original_unit_price_modified:
									# A2
									if quantity != DECIMAL_ZERO:
										original_unit_price = ( original_price / quantity ) - unit_deposit
										original_unit_price = original_unit_price.quantize(DECIMAL_0_01, rounding=ROUND_HALF_DOWN)
									else:
										original_price = DECIMAL_ZERO
								else:
									# A1
									if ( original_unit_price + unit_deposit ) != DECIMAL_ZERO:
										quantity = original_price / ( original_unit_price + unit_deposit )
										quantity = quantity.quantize(DECIMAL_0_0001, rounding=ROUND_HALF_DOWN)
									else:
										quantity = DECIMAL_ONE
										original_unit_price = original_price - unit_deposit
							else:
								if quantity_modified:
									if original_unit_price_modified:
										# A3
										if quantity != DECIMAL_ZERO:
											unit_deposit = ( original_price / quantity ) - original_unit_price
											unit_deposit = unit_deposit.quantize(DECIMAL_0_01, rounding=ROUND_HALF_DOWN)
										else:
											original_price = DECIMAL_ZERO
									else:
										# A2
										if quantity != DECIMAL_ZERO:
											original_unit_price = ( original_price / quantity ) - unit_deposit
											original_unit_price = original_unit_price.quantize(DECIMAL_0_01, rounding=ROUND_HALF_DOWN)
										else:
											original_price = DECIMAL_ZERO
								else:
									if original_unit_price_modified:
										# A4
										if ( original_unit_price + unit_deposit ) != DECIMAL_ZERO:
											quantity = original_price / ( original_unit_price + unit_deposit )
											quantity = quantity.quantize(DECIMAL_0_0001, rounding=ROUND_HALF_DOWN)
										else:
											quantity = DECIMAL_ONE
											original_unit_price = original_price - original_unit_price
									else:
										# A1
										if ( original_unit_price + unit_deposit ) != DECIMAL_ZERO:
											quantity = original_price / ( original_unit_price + unit_deposit )
											quantity = quantity.quantize(DECIMAL_0_0001, rounding=ROUND_HALF_DOWN)
										else:
											quantity = DECIMAL_ONE
											original_unit_price = original_price - unit_deposit
						else:
							if quantity_modified or original_unit_price_modified or unit_deposit_modified:
								# A5
								original_price = quantity * ( original_unit_price + unit_deposit )
								original_price = original_price.quantize(DECIMAL_0_01, rounding=ROUND_HALF_UP)
							else:
								# A6
								pass

	# print (Decimal('1.0')/Decimal('2.0')).quantize(Decimal('.0001'), rounding=ROUND_HALF_UP )
	# print (Decimal('2.0')/Decimal('3.0')).quantize(Decimal('.0001'), rounding=ROUND_HALF_UP  )
						purchase.quantity = quantity
						purchase.original_unit_price = original_unit_price
						purchase.unit_deposit = unit_deposit
						purchase.original_price = original_price
						purchase.vat_level = vat_level
						purchase.comment = comment
						array_purchase.append(purchase)

				row_num += 1
				row = get_row(worksheet, header, row_num)

			except KeyError, e:
				# Missing field
				error = True
				error_msg = _("Row %(row_num)d : A required column is missing %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
			except Exception, e:
				error = True
				error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
	return error, error_msg


def handle_permanence_done_uploaded_file(request, queryset, file_to_import):
	error = False
	error_msg = None
	wb = load_workbook(file_to_import)
	# dict for performance optimisation purpose : read the DB only once
	customer_buyinggroup_id, customer_2_id_dict = get_customer_2_id_dict()
	id_2_customer_vat_id_dict=get_customer_2_vat_id_dict()
	producer_buyinggroup_id, producer_2_id_dict = get_producer_2_id_dict()
	id_2_producer_vat_level_dict = get_id_2_producer_vat_level_dict()
	id_2_producer_price_list_multiplier_dict = get_id_2_producer_price_list_multiplier_dict()
	if customer_buyinggroup_id == None:
		error=True
		error_msg = _("At least one customer must represent the buying group.")	
	else:
		if producer_buyinggroup_id == None:
			error=True
			error_msg = _("At least one producer must represent the buying group.")	

	if not error:
		for permanence in queryset:
			if permanence.status == PERMANENCE_SEND:
				ws = wb.get_sheet_by_name(unicode(cap(permanence.__unicode__(),31), "utf8"))
				if ws:
					error, error_msg = import_permanence_purchases(ws, permanence=permanence, db_write=False,
						customer_2_id_dict=customer_2_id_dict,
						id_2_customer_vat_id_dict=id_2_customer_vat_id_dict,
						producer_2_id_dict=producer_2_id_dict,
						id_2_producer_vat_level_dict=id_2_producer_vat_level_dict,
						id_2_producer_price_list_multiplier_dict=id_2_producer_price_list_multiplier_dict
					)
					if error:
						error_msg = cap(permanence.__unicode__(),31) + " > " + error_msg
						break
			else:
				error=True
				error_msg = _("At least one of the permanences has already been invoiced.")
				break

	if not error:
		for permanence in queryset:
			ws = wb.get_sheet_by_name(cap(permanence.__unicode__(),31))
			if ws:
				error, error_msg  = import_permanence_purchases(ws, permanence=permanence, db_write=True,
					customer_2_id_dict=customer_2_id_dict,
					id_2_customer_vat_id_dict=id_2_customer_vat_id_dict,
					producer_2_id_dict=producer_2_id_dict,
					id_2_producer_vat_level_dict=id_2_producer_vat_level_dict,
					id_2_producer_price_list_multiplier_dict=id_2_producer_price_list_multiplier_dict
				)
				if error:
					error_msg = cap(permanence.__unicode__(),31) + " > " + error_msg
					break

	return error, error_msg

def import_permanence_done_xlsx(permanence, admin, request, queryset):
	form = None
	if 'apply' in request.POST:
		form = ImportXlsxForm(request.POST, request.FILES)
		if form.is_valid():
			file_to_import = request.FILES['file_to_import']
			if('.xlsx' in file_to_import.name) and (file_to_import.size <= 1000000):
				error, error_msg = handle_permanence_done_uploaded_file(request, queryset, file_to_import)
				if error:
					if error_msg == None:
						permanence.message_user(request, 
							_("Error when importing %s : Content not valid") % (file_to_import.name),
							level=messages.WARNING
						)
					else:
						permanence.message_user(request, 
							_("Error when importing %(file_name)s : %(error_msg)s") % {'file_name': file_to_import.name, 'error_msg':error_msg},
							level=messages.ERROR
						)
				else:
					permanence.message_user(request, _("Successfully imported %s.") % (file_to_import.name))
			else:
				permanence.message_user(request, 
					_("Error when importing %s : File size must be <= 1 Mb and extension must be .xlsx") % (file_to_import.name),
					level=messages.ERROR
				)
		else:
			permanence.message_user(request, _("No file to import."), level=messages.WARNING)
		return HttpResponseRedirect(request.get_full_path())
	elif 'cancel' in request.POST:
		permanence.message_user(request, _("Action canceled by the user."),level=messages.WARNING)
		return HttpResponseRedirect(request.get_full_path())
	form = ImportXlsxForm(
		initial={'_selected_action': request.POST.getlist(admin.ACTION_CHECKBOX_NAME)}
	)
	return render_response(request, 'repanier/import_xlsx.html', {
		'title': _("Import purchases"),
		'objects': queryset,
		'import_xlsx_form': form,
	})
