# -*- coding: utf-8 -*-
from repanier.const import *
from tools import *

from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.db.models import Sum
from django.contrib.sites.models import Site
# Alternative to openpyxl : XlsxWriter
from openpyxl.workbook import Workbook
from openpyxl.cell import get_column_letter
from openpyxl.style import Border
from openpyxl.style import NumberFormat, Color, Fill
from openpyxl.datavalidation import DataValidation, ValidationType, ValidationOperator

import datetime
from django.utils import timezone
from django.utils.timezone import utc
from django.contrib.sites.models import get_current_site

from repanier.models import Permanence
from repanier.models import Customer
from repanier.models import Producer
from repanier.models import CustomerInvoice
from repanier.models import ProducerInvoice
from repanier.models import Purchase
from repanier.models import Product
from repanier.models import BankAccount
from repanier.models import LUT_DepartmentForCustomer
from repanier.models import LUT_ProductionMode
from repanier.models import PermanenceBoard
from repanier.models import Staff


ROW_TITLE = 0
ROW_WIDTH = 1
ROW_VALUE = 2
ROW_FORMAT = 3
ROW_BOX = 4

def worksheet_setup_portait_a4(worksheet, title1, title2):
	worksheet.title = unicode(cap(title1,31), "utf8")
	worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT 
	worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
	worksheet.page_setup.fitToPage = True
	worksheet.page_setup.fitToHeight = 0
	worksheet.page_setup.fitToWidth = 1
	worksheet.print_gridlines = True
	worksheet.add_print_title(1, rows_or_cols='rows')
	worksheet.freeze_panes = 'A2'
	worksheet.header_footer.left_header.text = Site.objects.get_current().name
	s1 = ""
	s2 = ""
	for staff in Staff.objects.all().active().order_by():
		if staff.is_reply_to_order_email:
			c = staff.customer_responsible
			if c != None:
				s1 = unicode(_("Orders")) + ": " + c.long_basket_name + ", " + c.phone1
		if staff.is_reply_to_invoice_email:
			c = staff.customer_responsible
			if c != None:
				s2 += unicode(_("Invoices")) + ": " + c.long_basket_name + ", " + c.phone1
	s = s1 + chr(10) + s2 if s1 != "" and s2 != "" else s1 + s2
	worksheet.header_footer.right_header.text = s
	worksheet.header_footer.left_footer.text = title2
	worksheet.header_footer.center_footer.text = title1
	worksheet.header_footer.right_footer.text = 'Page &[Page]/&[Pages]'
	return worksheet.title

def worksheet_setup_landscape_a4(worksheet, title1, title2):
	worksheet.title = unicode(cap(title1,31), "utf8")
	worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE 
	worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
	worksheet.page_setup.fitToPage = True
	worksheet.page_setup.fitToHeight = 0
	worksheet.page_setup.fitToWidth = 1
	worksheet.print_gridlines = True
	worksheet.add_print_title(1, rows_or_cols='rows')
	worksheet.freeze_panes = 'A2'
	worksheet.header_footer.left_header.text = Site.objects.get_current().name
	worksheet.header_footer.left_footer.text = title2
	worksheet.header_footer.center_footer.text = title1
	worksheet.header_footer.right_footer.text = 'Page &[Page]/&[Pages]'
	return worksheet.title

def worksheet_set_header(worksheet, row_num, header):
	for col_num in xrange(len(header)):
		c = worksheet.cell(row=row_num, column=col_num)
		c.value = header[col_num][ ROW_TITLE ]
		c.style.font.bold = True
		c.style.alignment.wrap_text = False
		worksheet.column_dimensions[get_column_letter(col_num+1)].width = header[col_num][ ROW_WIDTH ]
		if header[col_num][ ROW_TITLE ] == unicode(_("Id")):
			worksheet.column_dimensions[get_column_letter(col_num+1)].visible = False

def export_orders_xlsx(permanence, wb = None):

	ws=None
	if wb==None:
		wb = Workbook()
		ws = wb.get_active_sheet()
	else:
		ws = wb.create_sheet()

# Customer info
	worksheet_setup_portait_a4(ws, unicode(permanence), '')

	header = [
		(unicode(_("Basket")), 20),
		(unicode(_('Family')), 35),
		(unicode(_('Phone1')), 15),
		(unicode(_('Phone2')), 15),
	]
	row_num = 0
	worksheet_set_header(ws, row_num, header)
	row_num += 1
	customer_set = Customer.objects.filter(
		purchase__permanence_id=permanence.id).not_the_buyinggroup().distinct()
	for customer in customer_set:
		row = [
			customer.short_basket_name,
			customer.long_basket_name,
			customer.phone1,
			customer.phone2
		]
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num, column=col_num)
			c.value = row[col_num]
			c.style.alignment.wrap_text = True
		row_num += 1
	c = ws.cell(row=row_num, column=1)
	c.value = unicode(_('Permanence Board Member List'))
	c.style.alignment.wrap_text = False
	c.style.font.bold = True
	row_num += 1
	distribution_date_save = None
	next_permanence_set = Permanence.objects.filter(distribution_date__gte = permanence.distribution_date).order_by("distribution_date")[:3]
	for next_permanence in next_permanence_set:
		for permanenceboard in PermanenceBoard.objects.filter(
			permanence=next_permanence.id):
			c = permanenceboard.customer
			if c != None:
				row = [
					next_permanence.distribution_date,
					c.long_basket_name,
					c.phone1,
					c.phone2,
					permanenceboard.permanence_role.short_name
				]
				for col_num in xrange(len(row)):
					c = ws.cell(row=row_num, column=col_num)
					c.value = row[col_num]
					c.style.alignment.wrap_text = False
					if distribution_date_save != next_permanence.distribution_date:
						c.style.font.bold = True
						distribution_date_save = next_permanence.distribution_date
				row_num += 1
	c = ws.cell(row=row_num, column=1)
	c.value = unicode(_('Staff Member List'))
	c.style.alignment.wrap_text = False
	c.style.font.bold = True
	row_num += 1
	for staff in Staff.objects.all().active().order_by():
		c = staff.customer_responsible
		if c != None:
			row = [
				staff.long_name,
				c.long_basket_name,
				c.phone1,
				c.phone2
			]
			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.value = row[col_num]
				c.style.alignment.wrap_text = True
			row_num += 1

	c = ws.cell(row=row_num, column=1)
	c.value = unicode(_('producers'))
	c.style.alignment.wrap_text = False
	c.style.font.bold = True
	row_num += 1
	for producer in Producer.objects.filter(permanence=permanence).order_by("short_profile_name"):
		row = [
			producer.short_profile_name,
			producer.long_profile_name,
			producer.phone1,
			producer.phone2
		]
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num, column=col_num)
			c.value = row[col_num]
			c.style.alignment.wrap_text = True
		row_num += 1

	if PERMANENCE_WAIT_FOR_SEND <= permanence.status <= PERMANENCE_SEND:
# Customer label
		ws = wb.create_sheet()
		worksheet_setup_portait_a4(ws, unicode(_('Label')), unicode(permanence))
		row_num = 0
		customer_set = Customer.objects.filter(
			purchase__permanence_id=permanence.id).not_the_buyinggroup().distinct()
		for customer in customer_set:
			c = ws.cell(row=row_num, column=0)
			c.value = customer.short_basket_name
			c.style.font.size = 36
			c.style.alignment.wrap_text = False
			c.style.borders.top.border_style = Border.BORDER_THIN
			c.style.borders.bottom.border_style = Border.BORDER_THIN
			c.style.borders.left.border_style = Border.BORDER_THIN
			c.style.borders.right.border_style = Border.BORDER_THIN
			c.style.alignment.vertical = 'center'
			c.style.alignment.horizontal = 'center'
			row_num += 1
			ws.row_dimensions[row_num].height = 60
		if row_num > 0:
			ws.column_dimensions[get_column_letter(1)].width = 120

	if PERMANENCE_WAIT_FOR_SEND <= permanence.status <= PERMANENCE_SEND:
# Basket check list, by customer
		ws = wb.create_sheet()
		worksheet_setup_portait_a4(ws, unicode(_('Customer check')), unicode(permanence))

		row_num = 0
		department_for_customer_save = None
		customer_save = None

		purchase_set = Purchase.objects.filter(
			permanence_id=permanence.id, producer__isnull=False).order_by(
			"customer__short_basket_name", 
			"product__placement", 
			"producer__short_profile_name",
			"department_for_customer",
			"product__long_name"
		)

		for purchase in purchase_set:
			qty = purchase.quantity
			if (qty != 0 or purchase.order_unit == PRODUCT_ORDER_UNIT_DEPOSIT):

				unit = get_unit(order_unit=purchase.order_unit, qty=qty)

				row = [
					(unicode(_("Placement")), 15, purchase.product.get_placement_display() if purchase.product != None else "", NumberFormat.FORMAT_TEXT),
					(unicode(_("Producer")), 15, purchase.producer.short_profile_name, NumberFormat.FORMAT_TEXT),
					(unicode(_("Product")), 60, purchase.long_name, NumberFormat.FORMAT_TEXT),
					(unicode(_("Quantity")), 10, qty, '#,##0.???'),
					(unicode(_("Unit")), 12, unit, NumberFormat.FORMAT_TEXT),
					(unicode(_("Basket")), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT),
				]

				if row_num == 0:
					worksheet_set_header(ws, row_num, row)
					row_num += 1

				if customer_save!= purchase.customer.id or department_for_customer_save != purchase.department_for_customer:
					if customer_save!= purchase.customer.id:
						# Chloé asked to add an empty line for the scissors.
						row_num += 1
					c = ws.cell(row=row_num, column=2)
					c.style.borders.bottom.border_style = Border.BORDER_THIN
					c.style.alignment.horizontal = 'right'
					c.style.font.italic = True
					department_for_customer_save = purchase.department_for_customer
					if department_for_customer_save != None:
						c.value = department_for_customer_save.short_name
					else:
						c.value = ""
					if customer_save!= purchase.customer.id:
						for col_num in xrange(len(row)):
							c = ws.cell(row=row_num, column=col_num)
							c.style.borders.top.border_style = Border.BORDER_THIN
							if col_num == 5: 
								c.value = purchase.customer.short_basket_name
								c.style.font.bold = True
						# Force the display of the department for next customer
						customer_save = purchase.customer.id
					else:
						c = ws.cell(row=row_num, column=5)
						c.value = purchase.customer.short_basket_name
					row_num += 1

				for col_num in xrange(len(row)):
					c = ws.cell(row=row_num, column=col_num)
					c.value = row[col_num][ ROW_VALUE ]
					c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
					c.style.borders.bottom.border_style = Border.BORDER_HAIR    

				row_num += 1

	if PERMANENCE_OPENED <= permanence.status <= PERMANENCE_SEND:

# Preparation list

		ws=None
		row_num = 0
		department_for_customer_save = None
		long_name_save = None
		producer_save = None
		producer_bold = False
		product_bold = False
		product_counter = 0
		previous_product_qty_sum = 0
		customer_save = None
		customer_bold = False

		producer_set = Producer.objects.filter(
          purchase__permanence_id=permanence.id).distinct()

		for producer in producer_set:
			purchase_set = Purchase.objects.none()
			if producer.invoice_by_basket:
				purchase_set = Purchase.objects.filter(
					permanence_id=permanence.id, 
					producer_id=producer.id,
					order_unit__in=[
					PRODUCT_ORDER_UNIT_LOOSE_PC, PRODUCT_ORDER_UNIT_LOOSE_KG, PRODUCT_ORDER_UNIT_LOOSE_PC_KG, 
					PRODUCT_ORDER_UNIT_NAMED_PC, PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG, 
					PRODUCT_ORDER_UNIT_DEPOSIT]
				).order_by(
					"customer__short_basket_name",
					# "product__placement",
					"department_for_customer",
					"long_name"
				)
			else:
				# Using quantity_for_preparation_order the order is by customer__short_basket_name if the product
				# is to be distributed by piece, otherwise by lower qty first.
				purchase_set = Purchase.objects.filter(
					permanence_id=permanence.id,
					producer_id=producer.id, 
					order_unit__in=[
					PRODUCT_ORDER_UNIT_LOOSE_PC, PRODUCT_ORDER_UNIT_LOOSE_KG, PRODUCT_ORDER_UNIT_LOOSE_PC_KG, 
					PRODUCT_ORDER_UNIT_NAMED_PC, PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG, 
					PRODUCT_ORDER_UNIT_DEPOSIT]
				).order_by(
					# "product__placement",
					"department_for_customer",
					"long_name",
					"quantity_for_preparation_order",
					"customer__short_basket_name"
				)


			for purchase in purchase_set:

				if (purchase.quantity != 0 or purchase.order_unit == PRODUCT_ORDER_UNIT_DEPOSIT):

					if 	ws==None:
						if wb==None:
							wb = Workbook()
							ws = wb.get_active_sheet()
						else:
							ws = wb.create_sheet()
						worksheet_setup_portait_a4(ws, unicode(_("Preparation")), unicode(permanence))

					if producer_save != purchase.producer_id or long_name_save != purchase.long_name:
						if producer_save != purchase.producer_id:
							producer_bold = True
							producer_save = purchase.producer_id
						else:
							producer_bold = False
						if producer.invoice_by_basket:
							product_bold = False
						else:
							product_bold = True
						long_name_save = purchase.long_name

						if product_counter > 1:
							c = ws.cell(row=row_num - 1, column=6)
							c.value = previous_product_qty_sum
							c.style.number_format.format_code = '#,##0.???'
						
						product_counter = 0
						previous_product_qty_sum = 0
					else:
						producer_bold = False
						product_bold = False

					previous_product_qty_sum += purchase.quantity

					product_counter += 1
					unit = None
					if  purchase.order_unit in [PRODUCT_ORDER_UNIT_LOOSE_PC,PRODUCT_ORDER_UNIT_NAMED_PC, PRODUCT_ORDER_UNIT_DEPOSIT]:
						unit = unicode(_("Piece(s) :"))
					elif purchase.order_unit in [PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG]:
						unit = unicode(_(u"€ :"))
					else:
						unit = unicode(_("Kg :"))

					row = [
						(unicode(_("Placement")), 7, purchase.product.get_placement_display() if purchase.product != None else "", NumberFormat.FORMAT_TEXT, False),
						(unicode(_("Producer")), 15, purchase.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
						(unicode(_("Product")), 40, purchase.long_name, NumberFormat.FORMAT_TEXT, False),
						(unicode(_("Quantity")), 10, purchase.quantity, '#,##0.???', False),
						(unicode(_("Basket")), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT, False),
						(unicode(_("Unit")), 10, get_unit(order_unit=purchase.order_unit, qty=purchase.quantity), NumberFormat.FORMAT_TEXT, False),
						(unicode(_("To distribute")), 9, "", '#,##0.???', False),
						(unicode(_("Prepared Quantity")), 22,  unit, NumberFormat.FORMAT_TEXT, True if purchase.order_unit in [PRODUCT_ORDER_UNIT_LOOSE_PC_KG, PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG] else False),
					]

					if row_num == 0:
						worksheet_set_header(ws, row_num, row)
						row_num += 1

					if department_for_customer_save != purchase.department_for_customer or ( purchase.producer.invoice_by_basket and customer_save != purchase.customer):
						if purchase.producer.invoice_by_basket and customer_save != purchase.customer:
							customer_bold = True
						else:
							customer_bold = False
						customer_save = purchase.customer
						c = ws.cell(row=row_num, column=2)
						c.style.borders.bottom.border_style = Border.BORDER_THIN
						c.style.alignment.horizontal = 'right'
						c.style.font.italic = True
						department_for_customer_save = purchase.department_for_customer
						if department_for_customer_save != None:
							c.value = department_for_customer_save.short_name
						else:
							c.value = ""
						if producer_bold:
							c = ws.cell(row=row_num, column=1)
							c.value = purchase.producer.short_profile_name
							c.style.font.bold = True
							for col_num in xrange(len(row)):
								c = ws.cell(row=row_num, column=col_num)
								c.style.borders.top.border_style = Border.BORDER_THIN
						row_num += 1
					else:
						customer_bold = False

					for col_num in xrange(len(row)):
						c = ws.cell(row=row_num, column=col_num)
						c.value = row[col_num][ ROW_VALUE ]
						c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
						if row[col_num][ ROW_BOX ]:
							c.style.borders.top.border_style = Border.BORDER_THIN   
							c.style.borders.bottom.border_style = Border.BORDER_THIN   
							c.style.borders.left.border_style = Border.BORDER_THIN   
							c.style.borders.right.border_style = Border.BORDER_THIN   							
						else:
							c.style.borders.bottom.border_style = Border.BORDER_HAIR
						if col_num == 2:
							c.style.alignment.wrap_text = True
							if product_bold:
								c.style.font.bold = True
						if col_num == 4:
							if customer_bold:
								c.style.font.bold = True
						if product_bold:
							c.style.borders.top.border_style = Border.BORDER_THIN

					row_num += 1

			if product_counter > 1:
				c = ws.cell(row=row_num - 1, column=6)
				c.value = previous_product_qty_sum
				c.style.number_format.format_code = '#,##0.???'

	# if PERMANENCE_WAIT_FOR_SEND <= permanence.status <= PERMANENCE_SEND:
	if PERMANENCE_OPENED <= permanence.status <= PERMANENCE_SEND:
# Order adressed to our producers, 
		producer_set = Producer.objects.filter(permanence=permanence).order_by("short_profile_name")
		for producer in producer_set:

			export_order_producer_xlsx(permanence=permanence, producer=producer, wb=wb)

	return wb

def export_permanence_planified_xlsx(request, queryset):

	wb = Workbook()
	ws = wb.get_active_sheet()

	for permanence in queryset[:1]:

# Product selected in a planified permanence
		worksheet_setup_landscape_a4(ws, unicode(_("Planified")), unicode(permanence))
		row_num = 0

		if permanence.status == PERMANENCE_PLANIFIED:

			producers_in_this_permanence = Producer.objects.filter(
				permanence=permanence).active()

			for product in Product.objects.filter(
				producer__in = producers_in_this_permanence
				).active().is_selected_for_offer().order_by(
				"producer__short_profile_name",
				"department_for_customer",
				"long_name"):
				row = [
					(unicode(_("Producer")), 15, product.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("Department")), 15, product.department_for_customer.short_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("Product")), 60, product.long_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("Unit Price")), 10, product.original_unit_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
				]

				if row_num == 0:
					worksheet_set_header(ws, row_num, row)
					row_num += 1

				for col_num in xrange(len(row)):
					c = ws.cell(row=row_num, column=col_num)
					c.value = row[col_num][ ROW_VALUE ]
					c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
					if row[col_num][ ROW_BOX ]:
						c.style.borders.top.border_style = Border.BORDER_THIN   
						c.style.borders.bottom.border_style = Border.BORDER_THIN   
						c.style.borders.left.border_style = Border.BORDER_THIN   
						c.style.borders.right.border_style = Border.BORDER_THIN   							
					else:
						c.style.borders.bottom.border_style = Border.BORDER_HAIR


				col_num = len(row)
				q_min = product.customer_minimum_order_quantity
				q_alert = product.customer_alert_order_quantity
				q_step = product.customer_increment_order_quantity
				# The q_min cannot be 0. In this case try to replace q_min by q_step.
				# In last ressort by q_alert.
				if q_step <= 0:
					q_step = q_min
				if q_min <= 0:
					q_min = q_step
				if q_min <= 0:
					q_min = q_alert
					q_step = q_alert
				c = ws.cell(row=row_num, column=col_num)
				c.value = unicode('---')
				ws.column_dimensions[get_column_letter(col_num+1)].width = 2.3
				c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
				col_num += 1
				q_valid = q_min
				q_counter = 0 # Limit to avoid too long selection list
				while q_valid <= q_alert and q_counter <= 20:
					q_counter += 1
					c = ws.cell(row=row_num, column=col_num)
					c.value = get_qty_display(
						q_valid,
					 	product.order_average_weight,
						product.order_unit
					)
					ws.column_dimensions[get_column_letter(col_num+1)].width = 15
					c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
					col_num += 1
					if q_valid < q_step:
						# 1; 2; 4; 6; 8 ... q_min = 1; q_step = 2
						# 0,5; 1; 2; 3 ... q_min = 0,5; q_step = 1
						q_valid = q_step
					else:
						# 1; 2; 3; 4 ... q_min = 1; q_step = 1
						# 0,125; 0,175; 0,225 ... q_min = 0,125; q_step = 0,50
						q_valid = q_valid + q_step

				row_num += 1

		if permanence.status == PERMANENCE_OPENED:

			for offer_item in OfferItem.objects.all().permanence(permanence).active().order_by(
				'product__producer__short_profile_name', 
				'product__department_for_customer', 
				'product__long_name'):
				row = [
					(unicode(_("Producer")), 15, offer_item.product.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("Department")), 15, offer_item.product.department_for_customer.short_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("Product")), 60, offer_item.product.long_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("Unit Price")), 10, offer_item.product.original_unit_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
				]

				if row_num == 0:
					worksheet_set_header(ws, row_num, row)
					row_num += 1

				for col_num in xrange(len(row)):
					c = ws.cell(row=row_num, column=col_num)
					c.value = row[col_num][ ROW_VALUE ]
					c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
					if row[col_num][ ROW_BOX ]:
						c.style.borders.top.border_style = Border.BORDER_THIN   
						c.style.borders.bottom.border_style = Border.BORDER_THIN   
						c.style.borders.left.border_style = Border.BORDER_THIN   
						c.style.borders.right.border_style = Border.BORDER_THIN   							
					else:
						c.style.borders.bottom.border_style = Border.BORDER_HAIR


				col_num = len(row)
				q_min = offer_item.product.customer_minimum_order_quantity
				q_alert = offer_item.product.customer_alert_order_quantity
				q_step = offer_item.product.customer_increment_order_quantity
				# The q_min cannot be 0. In this case try to replace q_min by q_step.
				# In last ressort by q_alert.
				if q_step <= 0:
					q_step = q_min
				if q_min <= 0:
					q_min = q_step
				if q_min <= 0:
					q_min = q_alert
					q_step = q_alert
				c = ws.cell(row=row_num, column=col_num)
				c.value = unicode('---')
				ws.column_dimensions[get_column_letter(col_num+1)].width = 2.3
				c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
				col_num += 1
				q_valid = q_min
				q_counter = 0 # Limit to avoid too long selection list
				while q_valid <= q_alert and q_counter <= 20:
					q_counter += 1
					c = ws.cell(row=row_num, column=col_num)
					c.value = get_qty_display(
						q_valid,
					 	offer_item.product.order_average_weight,
						offer_item.product.order_unit
					)
					ws.column_dimensions[get_column_letter(col_num+1)].width = 15
					c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
					col_num += 1
					if q_valid < q_step:
						# 1; 2; 4; 6; 8 ... q_min = 1; q_step = 2
						# 0,5; 1; 2; 3 ... q_min = 0,5; q_step = 1
						q_valid = q_step
					else:
						# 1; 2; 3; 4 ... q_min = 1; q_step = 1
						# 0,125; 0,175; 0,225 ... q_min = 0,125; q_step = 0,50
						q_valid = q_valid + q_step

				row_num += 1


	response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	filename = (unicode(_("Preview")) + u" - " + permanence.__unicode__() + u'.xlsx').encode('latin-1', errors='ignore')
	response['Content-Disposition'] = 'attachment; filename=' + filename
	wb.save(response)
	return response

def export_order_producer_xlsx(permanence, producer, wb = None):

	ws=None
	row_num = 0
	row_inc = 0
	current_site_name = Site.objects.get_current().name

	date_save = None
	department_for_customer_save = None
	department_for_customer_short_name_save = None
	product_bold = False
	qty_sum = 0
	unit_sum = None
	long_name_save = None
	unit_price_save = 0
	unit_deposit_save = 0
	total_price_sum = 0
	total_price_sum_sum = 0
	row_start_sum_sum = 0
	total_price_sum_sum_sum = 0
	formula_sum_sum_sum = []
	hidde_column_short_basket_name = True
	hidde_column_unit_deposit = True
	unit_save = None
	hidde_column_unit = True

	purchase_set = Purchase.objects.filter(
		permanence_id=permanence.id, producer_id=producer.id,
		order_unit__in=[
			PRODUCT_ORDER_UNIT_LOOSE_PC, PRODUCT_ORDER_UNIT_LOOSE_KG, PRODUCT_ORDER_UNIT_LOOSE_PC_KG,
			PRODUCT_ORDER_UNIT_NAMED_PC, PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG,
			PRODUCT_ORDER_UNIT_TRANSPORTATION
		]
		).exclude(quantity=0
		).order_by(
		"department_for_customer",
		"long_name",
		"customer__short_basket_name"
	)
	for purchase in purchase_set:

		if 	ws==None:
			if wb==None:
				wb = Workbook()
				ws = wb.get_active_sheet()
			else:
				ws = wb.create_sheet()
			worksheet_setup_landscape_a4(ws, unicode(producer.short_profile_name) + unicode(_(" by product")), unicode(permanence))

		short_basket_name = ""

		if long_name_save != purchase.long_name:
			product_bold = True
			row_start_sum = row_num
			if department_for_customer_save != purchase.department_for_customer_id:
				if department_for_customer_short_name_save != None:
					row_num += 1
					for col_num in xrange(7):
						c = ws.cell(row=row_num, column=col_num)
						c.style.borders.bottom.border_style = Border.BORDER_THIN
						if col_num == 2:
							c.value = unicode(_("Total Price")) + " " + department_for_customer_short_name_save
							c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
						if col_num == 6:
							formula = 'SUM(G%s:G%s)' % (row_start_sum_sum + 3, row_num)
							c.value = '=' + formula
							formula_sum_sum_sum.append(formula)
							c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
							c.style.font.bold = True
					total_price_sum_sum = 0
				row_start_sum_sum = row_num
				department_for_customer_save = purchase.department_for_customer_id
				if purchase.department_for_customer != None:
					department_for_customer_short_name_save = purchase.department_for_customer.short_name
				else:
					department_for_customer_short_name_save = ""
				c = None
				if long_name_save != None:
					row_num += 1
					c = ws.cell(row=row_num, column=1)
				else:
					c = ws.cell(row=1, column=1)
				c.value = department_for_customer_short_name_save
				c.style.font.bold = True

			long_name_save = purchase.long_name
			unit_price_save = purchase.original_unit_price
			unit_deposit_save = purchase.unit_deposit

			if unit_deposit_save != 0:
				hidde_column_unit_deposit = False

			if purchase.order_unit in [PRODUCT_ORDER_UNIT_NAMED_PC, PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG]:
				short_basket_name = purchase.customer.short_basket_name
				hidde_column_short_basket_name = False
			else:
				short_basket_name = current_site_name

			qty_sum = 0
			total_price_sum = 0
			row_inc = 1
		else:

			product_bold = False

			if unit_deposit_save != 0:
				hidde_column_unit_deposit = False

			if purchase.order_unit in [PRODUCT_ORDER_UNIT_NAMED_PC, PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG]:
				short_basket_name = purchase.customer.short_basket_name
				hidde_column_short_basket_name = False
				qty_sum = 0
				total_price_sum = 0
				row_inc = 1
			else:
				short_basket_name = current_site_name
				row_inc = 0

		qty_sum += purchase.quantity
		unit_sum = get_producer_unit(order_unit=purchase.order_unit, qty=qty_sum)
		if unit_save == None:
			unit_save = purchase.order_unit
		else:
			if purchase.order_unit in [PRODUCT_ORDER_UNIT_LOOSE_KG, PRODUCT_ORDER_UNIT_NAMED_KG]:
				hidde_column_unit = False
		# if purchase.order_unit not in [PRODUCT_ORDER_UNIT_LOOSE_PC_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG]:
		total_price_sum += purchase.original_price
		total_price_sum_sum += purchase.original_price
		total_price_sum_sum_sum += purchase.original_price

		row = [
			(unicode(_("Basket")), 20, short_basket_name, NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Quantity")), 10, qty_sum, '#,##0.???', True),
			(unicode(_("Unit")), 12, unit_sum, NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Product")), 60, long_name_save, NumberFormat.FORMAT_TEXT, product_bold),
			(unicode(_("Unit Price")), 10, unit_price_save, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("Desposit")), 10, unit_deposit_save, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("Total Price")), 12, total_price_sum, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False)
		]

		if row_num == 0:
			worksheet_set_header(ws, row_num, row)
			row_num += 1

		row_num += row_inc

		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num, column=col_num)
			c.value = row[col_num][ ROW_VALUE ]
			c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
			if row[col_num][ ROW_BOX ]:
				c.style.font.bold = True   
			c.style.borders.bottom.border_style = Border.BORDER_THIN

	if ws != None:
		if hidde_column_unit_deposit:
			ws.column_dimensions[get_column_letter(6)].visible = False
		if hidde_column_unit:
			ws.column_dimensions[get_column_letter(3)].visible = False
		if hidde_column_short_basket_name:
			ws.column_dimensions[get_column_letter(1)].visible = False
		row_num += 1
		for col_num in xrange(7):
			c = ws.cell(row=row_num, column=col_num)
			c.style.borders.bottom.border_style = Border.BORDER_THIN
			if col_num == 2:
				if department_for_customer_short_name_save != None:
					c.value = unicode(_("Total Price")) + " " + department_for_customer_short_name_save
				else:
					c.value = unicode(_("Total Price"))
				c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
			if col_num == 6:
				formula = 'SUM(G%s:G%s)' % (row_start_sum_sum + 3, row_num)
				c.value = '=' + formula
				formula_sum_sum_sum.append(formula)
				c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
				c.style.font.bold = True
		row_num += 1
		for col_num in xrange(7):
			c = ws.cell(row=row_num, column=col_num)
			c.style.borders.bottom.border_style = Border.BORDER_THIN
			if col_num == 1:
				c.value = unicode(_("Total Price")) + " " + current_site_name
				c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
			if col_num == 6:
				c.value = "=" + "+".join(formula_sum_sum_sum)
				c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
				c.style.font.bold = True

	ws=None
	row_num = 0
	row_inc = 0

	department_for_customer_save = None
	basket_save = None
	basket_bold = False
	total_price = 0
	total_price_sum = 0
	row_start_sum = 0
	total_price_sum_sum = 0
	formula_sum_sum = []
	hidde_column_unit_deposit = True
	unit_save = None
	hidde_column_unit = True

	purchase_set = Purchase.objects.filter(
		permanence_id=permanence.id, producer_id=producer.id,
		order_unit__in=[
			PRODUCT_ORDER_UNIT_LOOSE_PC, PRODUCT_ORDER_UNIT_LOOSE_KG, PRODUCT_ORDER_UNIT_LOOSE_PC_KG,
			PRODUCT_ORDER_UNIT_NAMED_PC, PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG,
			PRODUCT_ORDER_UNIT_TRANSPORTATION
		]
		).exclude(quantity=0
		).order_by(
		"customer__short_basket_name",
		"department_for_customer",
		"long_name"
	)
	for purchase in purchase_set:

		if 	ws==None:
			if wb==None:
				wb = Workbook()
				ws = wb.get_active_sheet()
			else:
				ws = wb.create_sheet()
			worksheet_setup_landscape_a4(ws, unicode(producer.short_profile_name) + unicode(_(" by basket")), unicode(permanence))

		if basket_save != purchase.customer.short_basket_name:
			basket_bold = True
			if basket_save != None:
				c = ws.cell(row=row_num, column=2)
				c.value = unicode(_("Total Price")) + " " + basket_save
				c = ws.cell(row=row_num, column=6)
				formula = 'SUM(G%s:G%s)' % (row_start_sum + 2, row_num)
				c.value = '=' + formula
				formula_sum_sum.append(formula)
				c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
				c.style.font.bold = True
				row_start_sum = row_num
				total_price_sum = 0
				row_num += 1
			basket_save = purchase.customer.short_basket_name
		else:
			basket_bold = False

		if unit_deposit_save != 0:
			hidde_column_unit_deposit = False

		# if purchase.order_unit not in [PRODUCT_ORDER_UNIT_LOOSE_PC_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG]:
		total_price = purchase.original_price
		total_price_sum += purchase.original_price
		total_price_sum_sum += purchase.original_price
		# else:
		# 	total_price = 0

		qty = purchase.quantity
		unit = get_producer_unit(order_unit=purchase.order_unit, qty=qty)

		if unit_save == None:
			unit_save = purchase.order_unit
		else:
			if (purchase.order_unit != unit_save and purchase.order_unit < PRODUCT_ORDER_UNIT_DEPOSIT) or (purchase.order_unit in [PRODUCT_ORDER_UNIT_LOOSE_KG, PRODUCT_ORDER_UNIT_NAMED_KG]):
				hidde_column_unit = False

		row = [
			(unicode(_("Basket")), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT, basket_bold),
			(unicode(_("Quantity")), 10, qty, '#,##0.???', True),
			(unicode(_("Unit")), 12, unit, NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Product")), 60, purchase.long_name, NumberFormat.FORMAT_TEXT, True),
			(unicode(_("Unit Price")), 10, purchase.original_unit_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("Desposit")), 10, unit_deposit_save, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("Total Price")), 12, total_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False)
		]

		if row_num == 0:
			worksheet_set_header(ws, row_num, row)
			row_num += 1

		if basket_bold:
			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.style.borders.top.border_style = Border.BORDER_THIN

		if basket_bold or (department_for_customer_save != purchase.department_for_customer):
			department_for_customer_save = purchase.department_for_customer
			if department_for_customer_save != None:
				c = ws.cell(row=row_num, column=1)
				c.value = department_for_customer_save.short_name
				row_num += 1

		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num, column=col_num)
			c.value = row[col_num][ ROW_VALUE ]
			c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
			if row[col_num][ ROW_BOX ]:
				c.style.font.bold = True
			c.style.borders.bottom.border_style = Border.BORDER_THIN

		row_num += 1

	if ws != None:
		if hidde_column_unit_deposit:
			ws.column_dimensions[get_column_letter(6)].visible = False
		if hidde_column_unit:
			ws.column_dimensions[get_column_letter(3)].visible = False
		c = ws.cell(row=row_num, column=2)
		c.value = unicode(_("Total Price")) + " " + basket_save
		c = ws.cell(row=row_num, column=6)
		formula = 'SUM(G%s:G%s)' % (row_start_sum + 2, row_num)
		c.value = '=' + formula
		formula_sum_sum.append(formula)
		c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
		c.style.font.bold = True
		row_num += 1
		for col_num in xrange(7):
			c = ws.cell(row=row_num, column=col_num)
			c.style.borders.top.border_style = Border.BORDER_THIN
			c.style.borders.bottom.border_style = Border.BORDER_THIN
			if col_num == 1:
				c.value = unicode(_("Total Price")) + " " + current_site_name
				# c.style.font.bold = True
			if col_num == 6:
				# c.value = total_price_sum_sum
				c.value = "=" + "+".join(formula_sum_sum)
				c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
				c.style.font.bold = True

	return wb

def export_order_customer_xlsx(permanence, customer, wb = None):

	ws=None
	if wb==None:
		wb = Workbook()
		ws = wb.get_active_sheet()
	else:
		ws = wb.create_sheet()
	worksheet_setup_landscape_a4(ws, unicode(_('Customer check')), unicode(permanence))

	row_num = 0

	purchase_set = Purchase.objects.filter(
		permanence_id=permanence.id, customer_id=customer.id, producer__isnull=False).order_by(
		"product__placement", 
		"producer__short_profile_name", 
		"department_for_customer",
		"long_name"
	)
	customer_save = None
	for purchase in purchase_set:
		qty = purchase.quantity
		if (qty != 0 or purchase.order_unit == PRODUCT_ORDER_UNIT_DEPOSIT):
			unit = get_unit(order_unit=purchase.order_unit, qty=qty)

			row = [
				(unicode(_("Placement")), 15, purchase.product.get_placement_display() if purchase.product != None else "", NumberFormat.FORMAT_TEXT, False),
				(unicode(_("Producer")), 15, purchase.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
				(unicode(_("Department")), 15, purchase.department_for_customer.short_name if purchase.department_for_customer != None else "N/A", NumberFormat.FORMAT_TEXT, False),
				(unicode(_("Product")), 60, purchase.long_name, NumberFormat.FORMAT_TEXT, False),
				(unicode(_("Basket")), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT, False),
				(unicode(_("Quantity")), 10, qty, '#,##0.???', False),
				(unicode(_("Unit")), 10, unit, NumberFormat.FORMAT_TEXT, True if purchase.order_unit in [PRODUCT_ORDER_UNIT_NAMED_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG] else False),
			]

			if row_num == 0:
				worksheet_set_header(ws, row_num, row)
				row_num += 1

			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.value = row[col_num][ ROW_VALUE ]
				c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
				if row[col_num][ ROW_BOX ]:
					c.style.borders.top.border_style = Border.BORDER_THIN   
					c.style.borders.bottom.border_style = Border.BORDER_THIN   
					c.style.borders.left.border_style = Border.BORDER_THIN   
					c.style.borders.right.border_style = Border.BORDER_THIN   							
				else:
					c.style.borders.bottom.border_style = Border.BORDER_HAIR    
				if customer_save!= purchase.customer.id:
					# Display the customer in bold when changing
					if col_num == 4: 
						c.style.font.bold = True
					c.style.borders.top.border_style = Border.BORDER_THIN
			if customer_save!= purchase.customer.id:
				customer_save = purchase.customer.id
			row_num += 1
	return wb


def export_invoices_xlsx(permanence, customer = None, producer = None, wb = None, sheet_name = ""):

	ws=None

# Detail of what has been prepared
	purchase_set = Purchase.objects.none()
	if customer == None and producer == None:

		if ws == None:
			if wb == None:
				wb = Workbook()
				ws = wb.get_active_sheet()
			else:
				ws = wb.create_sheet()
			worksheet_setup_landscape_a4(ws, unicode(_('Account summary')) + " " + unicode(sheet_name), unicode(permanence))

		row_num = 0

		max_customer_invoice_id = 0
		customer_invoice_set = CustomerInvoice.objects.filter(permanence=permanence).order_by("-id")[:1]
		if customer_invoice_set:
			max_customer_invoice_id = customer_invoice_set[0].id

		customer_set = Customer.objects.all()
		for customer in customer_set:
			balance_before = 0
			payment = 0
			prepared = 0
			balance_after = 0
			customer_invoice_set = CustomerInvoice.objects.filter(customer=customer,
				permanence=permanence
				).order_by()[:1]
			if customer_invoice_set:
				customer_invoice = customer_invoice_set[0]
				balance_before = customer_invoice.previous_balance
				payment = customer_invoice.bank_amount_in - customer_invoice.bank_amount_out
				prepared = customer_invoice.total_price_with_tax
				balance_after = customer_invoice.balance
			else:
				customer_invoice_set = CustomerInvoice.objects.filter(customer=customer,
					# Do not filter on date_balance : You may close the permanences in any date order
					# date_balance__lte=permanence.distribution_date,
					id__lt=max_customer_invoice_id
					).order_by("-id")[:1]
				if customer_invoice_set:
					customer_invoice = customer_invoice_set[0]
					balance_before = customer_invoice.balance
					balance_after = customer_invoice.balance
				else:
					# No invoice yet.
					balance_before = customer.initial_balance
					balance_after = customer.initial_balance
			row = [
				(unicode(_('Name')), 40, customer.long_basket_name, NumberFormat.FORMAT_TEXT),
				(unicode(_('Balance before')), 15, balance_before, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Payment')), 10, payment, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Prepared')), 10, prepared, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Balance after')), 15, balance_after, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Name')), 20, customer.short_basket_name, NumberFormat.FORMAT_TEXT),
			]

			if row_num == 0:
				worksheet_set_header(ws, row_num, row)
				row_num += 1

			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.value = row[col_num][ ROW_VALUE ]
				c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]

			row_num += 1

		row_break = row_num
		row_num += 1

		max_producer_invoice_id = 0
		producer_invoice_set = ProducerInvoice.objects.filter(permanence=permanence).order_by("-id")[:1]
		if producer_invoice_set:
			max_producer_invoice_id = producer_invoice_set[0].id
		producer_set = Producer.objects.all().active()
		for producer in producer_set:
			balance_before = 0
			payment = 0
			prepared = 0
			balance_after = 0
			producer_invoice_set = ProducerInvoice.objects.filter(producer=producer,
				permanence=permanence
				).order_by()[:1]
			if producer_invoice_set:
				producer_invoice = producer_invoice_set[0]
				balance_before = -producer_invoice.previous_balance
				payment = producer_invoice.bank_amount_out - producer_invoice.bank_amount_in
				prepared = producer_invoice.total_price_with_tax
				balance_after = -producer_invoice.balance
			else:
				producer_invoice_set = ProducerInvoice.objects.filter(producer=producer,
					# Do not filter on date_balance : You may close the permanences in any date order
					# date_balance__lte=permanence.distribution_date,
					id__lt=max_producer_invoice_id
					).order_by("-id")[:1]
				if producer_invoice_set:
					producer_invoice = producer_invoice_set[0]
					balance_before = -producer_invoice.balance
					balance_after = -producer_invoice.balance

			row = [
				(unicode(_('Name')), 40, producer.long_profile_name, NumberFormat.FORMAT_TEXT),
				(unicode(_('Balance before')), 15, balance_before, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Payment')), 10, payment, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Prepared')), 10, prepared, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Balance after')), 15, balance_after, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
				(unicode(_('Name')), 20, producer.short_profile_name, NumberFormat.FORMAT_TEXT),
			]

			if row_num == 0:
				worksheet_set_header(ws, row_num, row)
				row_num += 1

			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.value = row[col_num][ ROW_VALUE ]
				c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]

			row_num += 1

		initial_bank_amount = 0
		final_bank_amount = 0
		bank_account_set = BankAccount.objects.filter(permanence=permanence,
			producer=None,
			customer=None).order_by()[:1]
		if bank_account_set:
			final_bank_amount = bank_account_set[0].bank_amount_in - bank_account_set[0].bank_amount_out
			bank_account_set = BankAccount.objects.filter(id__lt=bank_account_set[0].id,
				producer=None,
				customer=None).order_by("-id")[:1]
			if bank_account_set:
				initial_bank_amount = bank_account_set[0].bank_amount_in - bank_account_set[0].bank_amount_out

		row_num +=1
		c = ws.cell(row=row_num, column=1)
		c.value = initial_bank_amount
		c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
		c = ws.cell(row=row_num, column=4)
		formula = 'B%s+SUM(C%s:C%s)-SUM(C%s:C%s)' % (row_num + 1, 2, row_break, row_break + 2, row_num - 1)
		c.value = '=' + formula
		c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '

		row_num +=1
		c = ws.cell(row=row_num, column=4)
		formula = 'SUM(E%s:E%s)-SUM(E%s:E%s)' % (2, row_break, row_break + 2, row_num - 2)
		c.value = '=' + formula
		c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '

		row_num +=1
		c = ws.cell(row=row_num, column=4)
		c.value = final_bank_amount
		c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '


		ws=None

		purchase_set = Purchase.objects.filter(
			permanence_id=permanence.id, producer__isnull=False, customer__isnull=False).order_by(
			"producer__short_profile_name",
			"department_for_customer",
			"long_name",
			"customer__short_basket_name"
		)
	elif customer != None:
		purchase_set = Purchase.objects.filter(
			permanence_id=permanence.id, producer__isnull=False, customer=customer).order_by(
			"producer__short_profile_name",
			"department_for_customer",
			"long_name",
			"customer__short_basket_name"
		)
	else:
		purchase_set = Purchase.objects.filter(
			permanence_id=permanence.id, producer=producer, customer__isnull=False).order_by(
			"producer__short_profile_name",
			"department_for_customer",
			"long_name",
			"customer__short_basket_name"
		)

	row_num = 0

	hidde_column_vat = True
	hidde_column_compensation = True

	for purchase in purchase_set:

		if ws == None:
			if wb == None:
				wb = Workbook()
				ws = wb.get_active_sheet()
			else:
				ws = wb.create_sheet()
			worksheet_setup_landscape_a4(ws, unicode(_('Invoice')) + " - " + unicode(sheet_name), unicode(permanence))

		qty = purchase.quantity
		# if (qty != 0):
		a_total_price = 0
		a_total_vat = 0
		a_total_compensation = 0
		if purchase.invoiced_price_with_compensation:
			a_total_price = purchase.price_with_compensation
			a_total_vat = 0
			a_total_compensation = purchase.price_with_compensation - purchase.price_with_vat
		else:
			a_total_price = purchase.price_with_vat
			a_total_vat = 0
			a_total_without_deposit = a_total_price - (purchase.unit_deposit * purchase.quantity)
			if purchase.vat_level == VAT_400:
				a_total_vat = (a_total_without_deposit * Decimal(0.06)).quantize(DECIMAL_0_001, rounding=ROUND_HALF_UP)
			elif purchase.vat_level == VAT_500:
				a_total_vat = (a_total_without_deposit * Decimal(0.12)).quantize(DECIMAL_0_001, rounding=ROUND_HALF_UP)
			elif purchase.vat_level == VAT_600:
				a_total_vat = (a_total_without_deposit * Decimal(0.21)).quantize(DECIMAL_0_001, rounding=ROUND_HALF_UP)
			a_total_compensation = 0
		if a_total_vat != 0:
			hidde_column_vat = False
		if a_total_compensation != 0:
			hidde_column_compensation = False
		a_unit_price = ( a_total_price / qty ) if qty != 0 else 0
		unit = get_unit(order_unit=purchase.order_unit, qty=qty)
		row = [
			(unicode(_("Producer")), 15, purchase.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Basket")), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Department")), 15, purchase.product.department_for_customer.short_name if purchase.product != None else "", NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Product")), 60, purchase.long_name, NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Quantity")), 10, qty, '#,##0.???', True if purchase.order_unit in [PRODUCT_ORDER_UNIT_LOOSE_PC_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG] else False),
			(unicode(_("Unit")), 10, unit, NumberFormat.FORMAT_TEXT, False),
			(unicode(_("Unit Invoided Price")), 10, a_unit_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("Total Invoiced Price")), 10, a_total_price , u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("Vat")), 10, a_total_vat , u'_ € * #,##0.000_ ;_ € * -#,##0.000_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("Compensation")), 10, a_total_compensation, u'_ € * #,##0.000_ ;_ € * -#,##0.000_ ;_ € * "-"??_ ;_ @_ ', False),
			(unicode(_("comment")), 30, purchase.comment, NumberFormat.FORMAT_TEXT, False),
		]

		if row_num == 0:
			worksheet_set_header(ws, row_num, row)
			row_num += 1

		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num, column=col_num)
			c.value = row[col_num][ ROW_VALUE ]
			c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
			if row[col_num][ ROW_BOX ]:
				c.style.borders.top.border_style = Border.BORDER_THIN
				c.style.borders.bottom.border_style = Border.BORDER_THIN
				c.style.borders.left.border_style = Border.BORDER_THIN
				c.style.borders.right.border_style = Border.BORDER_THIN
			else:
				c.style.borders.bottom.border_style = Border.BORDER_HAIR
			if col_num == 7:
				c.style.font.bold = True

		row_num += 1

	if wb != None:
		if hidde_column_vat:
			ws.column_dimensions[get_column_letter(9)].visible = False
		if hidde_column_compensation:
			ws.column_dimensions[get_column_letter(10)].visible = False

		current_site_name = Site.objects.get_current().name
		for col_num in xrange(11):
			c = ws.cell(row=row_num, column=col_num)
			c.style.borders.top.border_style = Border.BORDER_THIN
			c.style.borders.bottom.border_style = Border.BORDER_THIN
			if col_num == 1:
				c.value = unicode(_("Total Price")) + " " + current_site_name
				# c.style.font.bold = True
			if col_num == 7:
				formula = 'SUM(H%s:H%s)' % (2, row_num)
				c.value = '=' + formula
				c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
				c.style.font.bold = True
			if col_num == 8:
				formula = 'SUM(I%s:I%s)' % (2, row_num)
				c.value = '=' + formula
				c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
				c.style.font.bold = True
			if col_num == 9:
				formula = 'SUM(J%s:J%s)' % (2, row_num)
				c.value = '=' + formula
				c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
				c.style.font.bold = True

	return wb


def export_product_xlsx(request, queryset):

	wb = Workbook()
	ws = wb.get_active_sheet()
	# list of Departement for customer
	valid_values=[]
	department_for_customer_set = LUT_DepartmentForCustomer.objects.all().active().order_by()
	for department_for_customer in department_for_customer_set:
		valid_values.append(department_for_customer.short_name)
	valid_values.sort()
	department_for_customer_list = get_list(wb=wb, valid_values=valid_values)
	# List of Production mode
	valid_values=[]
	production_mode_set = LUT_ProductionMode.objects.all().active().order_by()
	for production_mode in production_mode_set:
		valid_values.append(production_mode.short_name)
	valid_values.sort()
	production_mode_list = get_list(wb=wb, valid_values=valid_values)
	# List of Yes/ 
	valid_values=[unicode(_('Yes')),]
	yes_list = get_list(wb=wb, valid_values=valid_values)
	# List of Vat or Compensation
	valid_values=[]
	for record in LUT_VAT:
		valid_values.append(unicode(record[1]))
	vat_list = get_list(wb=wb, valid_values=valid_values)
	# List of Unit
	valid_values=[]
	for record in LUT_PRODUCT_ORDER_UNIT:
		valid_values.append(unicode(record[1]))
	unit_list = get_list(wb=wb, valid_values=valid_values)

	queryset = queryset.order_by("-short_profile_name")
	for producer in queryset:

		row_num = 0
		product_set = Product.objects.filter(
			producer_id=producer.id, is_active=True
		)
		product_save = None
		if ws == None:
			ws = wb.create_sheet()
		now = timezone.localtime(timezone.now())
		worksheet_setup_landscape_a4(ws, unicode(producer.short_profile_name), unicode(now.strftime('%d-%m-%Y %H:%M')))
		for product in product_set:
			row = [
				(unicode(_("Id")), 10, product.id, '#,##0', False),
				(unicode(_("department_for_customer")), 15, product.department_for_customer.short_name, NumberFormat.FORMAT_TEXT, False),
				(unicode(_("production mode")), 15, product.production_mode.short_name if product.production_mode != None else u"", NumberFormat.FORMAT_TEXT, False),
				(unicode(_("is_into_offer")), 7,  unicode(_("Yes")) if product.is_into_offer else None, NumberFormat.FORMAT_TEXT, False),
				(unicode(_("long_name")), 60, product.long_name, NumberFormat.FORMAT_TEXT, False),
				(unicode(_("order unit")), 15, product.get_order_unit_display(), NumberFormat.FORMAT_TEXT, False),
				(unicode(_("order_average_weight")), 10, product.order_average_weight, '#,##0.???', False),
				(unicode(_("original_unit_price")), 10, product.original_unit_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
				(unicode(_("deposit")), 10, product.unit_deposit, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
				(unicode(_("vat or compensation")), 10, product.get_vat_level_display(), NumberFormat.FORMAT_TEXT, False),
				(unicode(_("customer_minimum_order_quantity")), 10, product.customer_minimum_order_quantity, '#,##0.???', False),
				(unicode(_("customer_increment_order_quantity")), 10, product.customer_increment_order_quantity, '#,##0.???', False),
				(unicode(_("customer_alert_order_quantity")), 10, product.customer_alert_order_quantity, '#,##0.???', False),
			]

			if row_num==0:
				worksheet_set_header(ws, row_num, row)
				row_num += 1

			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.value = row[col_num][ ROW_VALUE ]
				c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
				if row[col_num][ ROW_BOX ]:
					c.style.borders.top.border_style = Border.BORDER_THIN   
					c.style.borders.bottom.border_style = Border.BORDER_THIN   
					c.style.borders.left.border_style = Border.BORDER_THIN   
					c.style.borders.right.border_style = Border.BORDER_THIN   							
				else:
					c.style.borders.bottom.border_style = Border.BORDER_HAIR    
				if product_save!= product.department_for_customer.id:
					c.style.borders.top.border_style = Border.BORDER_THIN
			if product_save!= product.department_for_customer.id:
				product_save = product.department_for_customer.id
			row_num += 1
		#  Now, for helping the user encoding new purchases
		row = [
			(unicode(_("Id")), 10, u"", NumberFormat.FORMAT_TEXT),
			(unicode(_("department_for_customer")), 15, u"", NumberFormat.FORMAT_TEXT),
			(unicode(_("production mode")), 15, u"", NumberFormat.FORMAT_TEXT),
			(unicode(_("is_into_offer")), 7, u"", NumberFormat.FORMAT_TEXT),
			(unicode(_("long_name")), 60, u"", NumberFormat.FORMAT_TEXT),
			(unicode(_("order unit")), 15, u"", NumberFormat.FORMAT_TEXT),
			(unicode(_("order_average_weight")), 10, u"", '#,##0.???'),
			(unicode(_("original_unit_price")), 10, u"", u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
			(unicode(_("deposit")), 10, u"", u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
			(unicode(_("vat or compensation")), 10, u"", NumberFormat.FORMAT_TEXT),
			(unicode(_("customer_minimum_order_quantity")), 10, u"", '#,##0.???'),
			(unicode(_("customer_increment_order_quantity")), 10, u"", '#,##0.???'),
			(unicode(_("customer_alert_order_quantity")), 10, u"", '#,##0.???'),
		]

		if row_num==0:
			#  add a header if there is no previous movement.
			worksheet_set_header(ws, row_num, row)
			row_num +=1

		# Data validation Id
		dv = DataValidation(ValidationType.WHOLE,
			ValidationOperator.EQUAL,
			0)
		ws.add_data_validation(dv)
		dv.ranges.append('A%s:A5000' % (row_num+1))
		# Data validation Departement for customer
		dv = DataValidation(ValidationType.LIST, formula1=department_for_customer_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('B2:B5000')
		# Data validation Production mode
		dv = DataValidation(ValidationType.LIST, formula1=production_mode_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('C2:C5000')
		# Data validation Unit
		dv = DataValidation(ValidationType.LIST, formula1=unit_list, allow_blank=False)
		ws.add_data_validation(dv)
		dv.ranges.append('F2:F5000')
		# Data validation Yes/ 
		dv = DataValidation(ValidationType.LIST, formula1=yes_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('D2:D5000')
		# Data validation qty / weight
		dv = DataValidation(ValidationType.DECIMAL,
			ValidationOperator.GREATER_THAN_OR_EQUAL,
			0)
		ws.add_data_validation(dv)
		dv.ranges.append('G2:I5000')
		dv = DataValidation(ValidationType.DECIMAL,
			ValidationOperator.GREATER_THAN_OR_EQUAL,
			0)
		ws.add_data_validation(dv)
		dv.ranges.append('K2:M5000')
		# Data validation Vat or Compensation
		dv = DataValidation(ValidationType.LIST, formula1=vat_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('J2:J5000')
		# End of data validation

		#  Add formating for empty cells.
		for row_num in xrange(row_num,row_num+30):
			for col_num in xrange(len(row)):
				c = ws.cell(row=row_num, column=col_num)
				c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]

		# if row_num > 0:
		# Let's create a new sheet for the next producer
		ws = None
	wb.worksheets.reverse()
	response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	filename = str(unicode(_("products")) + '.xlsx')
	response['Content-Disposition'] = 'attachment; filename=' + filename
	wb.save(response)
	return response

def export_permanence_done_xlsx(request, queryset):

	wb = Workbook()
	ws = wb.get_active_sheet()
	# list of Customer
	valid_values=[]
	customer_set = Customer.objects.all().active().order_by()
	for customer in customer_set:
		valid_values.append(customer.short_basket_name)
	valid_values.sort()
	customer_list = get_list(wb=wb, valid_values=valid_values)
	# List of Producer
	valid_values=[]
	producer_set = Producer.objects.all().active().order_by()
	for producer in producer_set:
		valid_values.append(producer.short_profile_name)
	valid_values.sort()
	producer_list = get_list(wb=wb, valid_values=valid_values)
	# list of Departement for customer
	valid_values=[]
	department_for_customer_set = LUT_DepartmentForCustomer.objects.all().active().order_by()
	for department_for_customer in department_for_customer_set:
		valid_values.append(department_for_customer.short_name)
	valid_values.sort()
	department_for_customer_list = get_list(wb=wb, valid_values=valid_values)
	# List of Vat or Compensation
	valid_values=[]
	for record in LUT_VAT:
		valid_values.append(unicode(record[1]))
	vat_list = get_list(wb=wb, valid_values=valid_values)

	last_permanence_name = None
	producer_valid_values = []
	customer_valid_values = []

	yellowFill = Fill()
	yellowFill.start_color.index = 'FFEEEE11'
	yellowFill.end_color.index = 'FFEEEE11'
	yellowFill.fill_type = Fill.FILL_SOLID

	for permanence in queryset[:1]:

		if ws == None:
			ws = wb.create_sheet()

		last_permanence_name = worksheet_setup_landscape_a4(ws, unicode(permanence), unicode(_('invoices')))
		producer_valid_values = []
		customer_valid_values = []

		row_num = 0

		producer_set = Producer.objects.all().active()
		for producer in producer_set:

			purchase_set = Purchase.objects.none()
			if producer.invoice_by_basket:
				purchase_set = Purchase.objects.filter(
					permanence_id=permanence.id, 
					producer_id=producer.id
				).order_by(
					"customer__short_basket_name",
					# "product__placement",
					"department_for_customer",
					"long_name"
				)
			else:
				# Using quantity_for_preparation_order the order is by customer__short_basket_name if the product
				# is to be distributed by piece, otherwise by lower qty first.
				purchase_set = Purchase.objects.filter(
					permanence_id=permanence.id,
					producer_id=producer.id
				).order_by(
					# "product__placement",
					"department_for_customer",
					"long_name",
					"quantity_for_preparation_order",
					"customer__short_basket_name"
				)

			sum_on = None
			sum_counter = 0
			sum_original_price = 0
			sum_quantity = 0
			customer_short_basket_name = None
			product_long_name = None

			for purchase in purchase_set:
				customer_short_basket_name = purchase.customer.short_basket_name
				producer_short_profile_name = purchase.producer.short_profile_name
				product_long_name = purchase.long_name
				if producer_short_profile_name not in producer_valid_values:
					producer_valid_values.append(producer_short_profile_name)
				if customer_short_basket_name not in customer_valid_values:
					customer_valid_values.append(customer_short_basket_name)

				if producer.invoice_by_basket:
					if sum_on != customer_short_basket_name:
						if sum_on != None:
							if sum_counter > 1:
								c = ws.cell(row=row_num-1, column=9)
								c.value = sum_original_price
								c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
								ws.conditional_formatting.addCellIs(get_column_letter(10)+str(row_num),'notEqual', [get_column_letter(1)+str(row_num+1)], True, wb, None, None, yellowFill)
								c.style.font.bold = True
							c = ws.cell(row=row_num, column=0)
							c.value = sum_original_price
							c = ws.cell(row=row_num-1, column=1)
							c.style.font.bold = True
							c = ws.cell(row=row_num-1, column=4)
							c.style.font.bold = True
							for col_num in xrange(12):
								c = ws.cell(row=row_num-1, column=col_num)
								c.style.borders.bottom.border_style = Border.BORDER_THIN
							row_num += 1
							ws.row_dimensions[row_num].visible = False
						sum_on = customer_short_basket_name
						sum_original_price = 0
						sum_quantity = 0
						sum_counter = 0
				else:
					if sum_on != product_long_name:
						if sum_on != None:
							if sum_counter > 1:
								c = ws.cell(row=row_num-1, column=9)
								c.value = sum_quantity
								c.style.number_format.format_code = '#,##0.????'
								ws.conditional_formatting.addCellIs(get_column_letter(10)+str(row_num),'notEqual', [get_column_letter(1)+str(row_num+1)], True, wb, None, None, yellowFill)
								c.style.font.bold = True
							c = ws.cell(row=row_num, column=0)
							c.value = sum_quantity
							c = ws.cell(row=row_num-1, column=1)
							c.style.font.bold = True
							c = ws.cell(row=row_num-1, column=3)
							c.style.font.bold = True
							for col_num in xrange(12):
								c = ws.cell(row=row_num-1, column=col_num)
								c.style.borders.bottom.border_style = Border.BORDER_THIN
							row_num += 1
							ws.row_dimensions[row_num].visible = False
						sum_on = product_long_name
						sum_original_price = 0
						sum_quantity = 0
						sum_counter = 0
					else:
						if PRODUCT_ORDER_UNIT_NAMED_PC <= purchase.product.order_unit <= PRODUCT_ORDER_UNIT_NAMED_PC_KG:
							# Don't display the sum_quantity of NAMED products
							sum_counter = 0

				sum_original_price += purchase.original_price
				sum_quantity += purchase.quantity
				sum_counter += 1

				row = [
					(unicode(_("Id")), 10, purchase.id, '#,##0', False),
					(unicode(_("producer")), 15, producer_short_profile_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("Department")), 15, "" if purchase.department_for_customer == None else purchase.department_for_customer.short_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("product")), 60, product_long_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("customer")), 15, customer_short_basket_name, NumberFormat.FORMAT_TEXT, False),
					(unicode(_("quantity")), 10, purchase.quantity, '#,##0.????', True if purchase.order_unit in [PRODUCT_ORDER_UNIT_LOOSE_PC_KG, PRODUCT_ORDER_UNIT_NAMED_PC_KG] else False),
					(unicode(_("original unit price")), 10, purchase.original_unit_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
					(unicode(_("deposit")), 10, purchase.unit_deposit, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
					(unicode(_("original price")), 10, purchase.original_price, u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
					(unicode(_("invoiced")), 10, "", NumberFormat.FORMAT_TEXT, False),
					(unicode(_("comment")), 30, cap(purchase.comment,100), NumberFormat.FORMAT_TEXT, False),
					(unicode(_("vat or compensation")), 10, purchase.get_vat_level_display(), NumberFormat.FORMAT_TEXT, False),
				]

				if row_num == 0:
					worksheet_set_header(ws, row_num, row)
					row_num += 1

				for col_num in xrange(len(row)):
					c = ws.cell(row=row_num, column=col_num)
					c.value = row[col_num][ ROW_VALUE ]
					c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]
					if row[col_num][ ROW_BOX ]:
						c.style.borders.top.border_style = Border.BORDER_THIN   
						c.style.borders.bottom.border_style = Border.BORDER_THIN   
						c.style.borders.left.border_style = Border.BORDER_THIN   
						c.style.borders.right.border_style = Border.BORDER_THIN   							
					else:
						if c.style.borders.bottom.border_style != Border.BORDER_THIN:
							c.style.borders.bottom.border_style = Border.BORDER_HAIR
					if col_num in [0, 5, 6, 7, 8]:
						ws.conditional_formatting.addCellIs(get_column_letter(col_num+1)+str(row_num+1),'notEqual', [str(row[col_num][ ROW_VALUE ])], True, wb, None, None, yellowFill)
				row_num += 1
			if sum_on != None:
				c = ws.cell(row=row_num-1, column=9)
				if producer.invoice_by_basket:
					if sum_counter > 1:
						c.value = sum_original_price
						c.style.number_format.format_code = u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
						c.style.font.bold = True
						ws.conditional_formatting.addCellIs(get_column_letter(10)+str(row_num),'notEqual', [get_column_letter(1)+str(row_num+1)], True, wb, None, None, yellowFill)
					c = ws.cell(row=row_num, column=0)
					c.value = sum_original_price
					c = ws.cell(row=row_num-1, column=1)
					c.style.font.bold = True
					c = ws.cell(row=row_num-1, column=4)
					c.style.font.bold = True
				else:
					if sum_counter > 1:
						c.value = sum_quantity
						c.style.number_format.format_code = '#,##0.????'
						c.style.font.bold = True
						ws.conditional_formatting.addCellIs(get_column_letter(10)+str(row_num),'notEqual', [get_column_letter(1)+str(row_num+1)], True, wb, None, None, yellowFill)
					c = ws.cell(row=row_num, column=0)
					c.value = sum_quantity
					c = ws.cell(row=row_num-1, column=1)
					c.style.font.bold = True
					c = ws.cell(row=row_num-1, column=3)
					c.style.font.bold = True
				for col_num in xrange(12):
					c = ws.cell(row=row_num-1, column=col_num)
					c.style.borders.bottom.border_style = Border.BORDER_THIN
				row_num +=1

		# Data validation Producer
		dv = DataValidation(ValidationType.LIST, formula1=producer_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('B2:B5000')
		# Data validation Departement for customer
		dv = DataValidation(ValidationType.LIST, formula1=department_for_customer_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('C2:C5000')
		# Data validation Customer
		dv = DataValidation(ValidationType.LIST, formula1=customer_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('E2:E5000')
		# Data validation Vat or Compensation
		dv = DataValidation(ValidationType.LIST, formula1=vat_list, allow_blank=True)
		ws.add_data_validation(dv)
		dv.ranges.append('L2:L5000')
		# End of data validation

		# use a new ws if needed for another permanence
		ws = None

	ws = wb.create_sheet()
	worksheet_setup_landscape_a4(ws, unicode(_('Account summary')), unicode(permanence))
	row_num = 0

	row = [
		(unicode(_("Who")), 15, u"", NumberFormat.FORMAT_TEXT),
		(unicode(_("Bank_amount_in")), 12, u"", u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
		(unicode(_("Bank_amount_out")), 12, u"", u'_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
		(unicode(_("Operation_comment")), 60, u"", NumberFormat.FORMAT_TEXT),
	]

	worksheet_set_header(ws, row_num, row)
	row_num +=1

	today = timezone.localtime(timezone.now()).date()
	current_site_name = Site.objects.get_current().name
	producer_valid_values.sort()
	for v in producer_valid_values:
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num, column=col_num)
			if col_num == 0:
				c.value = v
			if col_num == 2:
				c.value = "=SUMIF('" + last_permanence_name + "'!$B$2:$B$5000,A" + str(row_num + 1) + ",'" + last_permanence_name + "'!$I$2:$I$5000)"
			if col_num == 3:
				c.value = unicode(_('Delivery')) + " " + current_site_name + " - " + last_permanence_name + ". " + unicode(_('Thanks!'))
			c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]

		row_num +=1

	customer_valid_values.sort()
	for v in customer_valid_values:
		for col_num in xrange(len(row)):
			c = ws.cell(row=row_num, column=col_num)
			if col_num == 0:
				c.value = v
			if col_num == 1:
				c.value = "=SUMIF('" + last_permanence_name + "'!$E$2:$E$5000,A" + str(row_num + 1) + ",'" + last_permanence_name + "'!$I$2:$I$5000)"
			if col_num == 3:
				c.value = unicode(_('Delivery')) + " - " + last_permanence_name + "."
			c.style.number_format.format_code = row[col_num][ ROW_FORMAT ]

		row_num +=1


	wb.worksheets.reverse()
	response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	filename = (unicode(_("invoices")) + u" - " + permanence.__unicode__() + u'.xlsx').encode('latin-1', errors='ignore')
	response['Content-Disposition'] = 'attachment; filename=' + filename
	wb.save(response)
	return response

def get_worksheet_and_col_for_data_validation(wb=None):
	ws_dv_name = cap(unicode(_("data validation")),31)
	ws = wb.get_sheet_by_name(ws_dv_name)
	if ws==None:
		ws = wb.create_sheet(index=0)
		worksheet_setup_landscape_a4(ws, ws_dv_name, "")
	col_num = 0
	c = ws.cell(row=0, column=col_num)
	while (c.value!=None ) and (col_num < 20):
		col_num+=1
		c = ws.cell(row=0, column=col_num)
	return ws, ws_dv_name, col_num, get_column_letter(col_num + 1)

def get_list(wb=None, valid_values=None):
	if valid_values:
		ws_dv, ws_dv_name, col_dv, col_letter_dv = get_worksheet_and_col_for_data_validation(wb=wb)
		row_num = 0
		for v in valid_values:
			c = ws_dv.cell(row=row_num, column=col_dv)
			c.value = v
			c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
			row_num +=1
		formula1 = "'%s'!$%s$1:$%s$%s" % (ws_dv_name, col_letter_dv, col_letter_dv, row_num + 1)
		return formula1