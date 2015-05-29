# -*- coding: utf-8
from __future__ import unicode_literals
from django.conf import settings
from django.db import transaction
from django.utils.translation import ugettext_lazy as _
from django.utils import translation
from django.http import HttpResponse
from openpyxl import load_workbook
from openpyxl.style import Border
from openpyxl.style import Fill
from openpyxl.style import NumberFormat
from openpyxl.styles import Color
from openpyxl.workbook import Workbook

from export_tools import *
from import_tools import *
from repanier.const import *
from repanier.models import OfferItem
from repanier.models import Product
from views import import_xslx_view


def export_stock(permanence, customer_price=False, wb=None, ws_customer_title=None):

    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Id"), 5),
        (_("OfferItem"), 5),
        (_("Reference"), 20),
        (_("Product"), 60),
        (_("producer unit price") if customer_price else _("customer unit price"), 10),
        (_("Deposit"), 10),
        (_("Asked"), 10),
        (_("quantity ordered"), 10),
        (_("Initial stock"), 10),
        (_('€'), 15),
        (_("Stock used"), 10),
        (_("Add 2 stock"), 10),
        (_("Final stock"), 10),
        (_('€'), 15),
        (_("Current stock"), 10),
        (_('€'), 15),
    ]
    offer_items = OfferItem.objects.filter(
        permanence_id=permanence.id,
        manage_stock=True,
        translations__language_code=translation.get_language()
    ).order_by(
        "producer",
        "translations__order_sort_order"
    ).iterator()
    offer_item = next_row(offer_items)
    if offer_item is not None:
        wb, ws = new_landscape_a4_sheet(
            wb,
            _('Stock check'),
            permanence,
            header
        )
        formula_main_total_a = []
        formula_main_total_b = []
        formula_main_total_c = []
        show_column_reference = False
        show_column_current_stock = False
        permanence_status = permanence.status
        row_num = 1
        while offer_item is not None:
            producer_save = offer_item.producer
            row_start_producer = row_num + 1
            c = ws.cell(row=row_num, column=2)
            c.value = producer_save.short_profile_name
            c.style.font.bold = True
            c.style.font.italic = True
            while offer_item is not None and producer_save.id == offer_item.producer_id:
                department_for_customer_save = offer_item.department_for_customer
                c = ws.cell(row=row_num, column=3)
                c.value = producer_save.short_profile_name + " - " + department_for_customer_save.short_name
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c.style.alignment.horizontal = 'right'
                c.style.font.italic = True
                row_num += 1
                while offer_item is not None and producer_save.id == offer_item.producer_id and department_for_customer_save.id == offer_item.department_for_customer_id:
                    if len(offer_item.reference) < 36:
                        offer_item_reference = offer_item.reference
                        show_column_reference = True
                    else:
                        offer_item_reference = ""
                    if offer_item.order_unit < PRODUCT_ORDER_UNIT_DEPOSIT:

                        asked = offer_item.quantity_invoiced - offer_item.add_2_stock
                        stock = offer_item.stock
                        add_2_stock = offer_item.add_2_stock
                        product_stock = offer_item.product.stock
                        to_order = asked - min(asked, stock) + add_2_stock

                        c = ws.cell(row=row_num, column=0)
                        c.value = offer_item.producer_id
                        c = ws.cell(row=row_num, column=1)
                        c.value = offer_item.id
                        c = ws.cell(row=row_num, column=2)
                        c.value = offer_item_reference
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=3)
                        c.value = offer_item.get_long_name(is_quantity_invoiced=True)
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=4)
                        unit_price = offer_item.customer_unit_price if customer_price else offer_item.producer_unit_price
                        c.value = unit_price
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=5)
                        c.value = offer_item.unit_deposit
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=6)
                        if ws_customer_title is None:
                            c.value = asked
                        else:
                            c.value = "=SUMIF('" + ws_customer_title + "'!B2:B5000,B" + str(
                                row_num + 1) + ",'" + ws_customer_title + "'!F2:F5000)"
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=7)
                        c.value = '=G%s-K%s+L%s' % (row_num + 1, row_num + 1, row_num + 1)
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(8) + str(row_num+1), 'notEqual',
                            [str(to_order)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=8)
                        c.value = stock
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c.style.font.color = Color(Color.BLUE)
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(9) + str(row_num+1), 'notEqual',
                            [str(stock)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=9)
                        purchases_price = (stock *
                            (unit_price +
                            offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                        c.value = '=ROUND(I%s*(E%s+F%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(10) + str(row_num+1), 'notEqual',
                            [str(purchases_price)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=10)
                        c.value = '=MIN(G%s,I%s)' % (row_num + 1, row_num + 1)
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=11)
                        c.value = add_2_stock
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c.style.font.color = Color(Color.BLUE)
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(12) + str(row_num+1), 'notEqual',
                            [str(add_2_stock)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=12)
                        c.value = '=I%s-K%s+L%s' % (row_num + 1, row_num + 1, row_num + 1)
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c.style.font.bold = True
                        final_stock = stock - min(asked, stock) + add_2_stock
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(13) + str(row_num+1), 'notEqual',
                            [str(final_stock)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=13)
                        purchases_price = (final_stock *
                            (unit_price +
                            offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                        c.value = '=ROUND(M%s*(E%s+F%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(14) + str(row_num+1), 'notEqual',
                            [str(purchases_price)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=14)
                        c.value = product_stock
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(15) + str(row_num+1), 'notEqual',
                            [str(final_stock)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=15)
                        c.value = '=ROUND(O%s*(E%s+F%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        if permanence_status <= PERMANENCE_SEND:
                            show_column_current_stock |= stock != product_stock
                        else:
                            show_column_current_stock |= final_stock != product_stock

                        row_num += 1
                    offer_item = next_row(offer_items)
            row_num += 1
            c = ws.cell(row=row_num, column=3)
            c.value = "%s %s" % (_("Total price"), producer_save.short_profile_name)
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            c.style.font.bold = True
            c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
            c = ws.cell(row=row_num, column=9)
            formula = 'SUM(J%s:J%s)' % (row_start_producer, row_num)
            c.value = '=' + formula
            c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
            c.style.font.bold = True
            formula_main_total_a.append(formula)
            c = ws.cell(row=row_num, column=13)
            formula = 'SUM(N%s:N%s)' % (row_start_producer, row_num)
            c.value = '=' + formula
            c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
            c.style.font.bold = True
            formula_main_total_b.append(formula)
            c = ws.cell(row=row_num, column=15)
            formula = 'SUM(P%s:P%s)' % (row_start_producer, row_num)
            c.value = '=' + formula
            c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
            c.style.font.bold = True
            formula_main_total_c.append(formula)

            if offer_items is not None:
                # Display a separator line between producers
                row_num += 1
                for col_num in xrange(16):
                    c = ws.cell(row=row_num, column=col_num)
                    c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
                row_num += 2

        c = ws.cell(row=row_num, column=3)
        c.value = "%s" % _("Total price")
        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
        c.style.font.bold = True
        c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
        c = ws.cell(row=row_num, column=9)
        c.value = "=" + "+".join(formula_main_total_a)
        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
        c.style.font.bold = True
        c = ws.cell(row=row_num, column=13)
        c.value = "=" + "+".join(formula_main_total_b)
        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
        c.style.font.bold = True
        c = ws.cell(row=row_num, column=15)
        c.value = "=" + "+".join(formula_main_total_c)
        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
        c.style.font.bold = True

        row_num += 1
        for col_num in xrange(16):
            c = ws.cell(row=row_num, column=col_num)
            c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED

        ws.column_dimensions[get_column_letter(1)].visible = False
        ws.column_dimensions[get_column_letter(2)].visible = False
        # ws.column_dimensions[get_column_letter(7)].visible = False
        ws.column_dimensions[get_column_letter(11)].visible = False
        ws.column_dimensions[get_column_letter(15)].visible = show_column_current_stock
        ws.column_dimensions[get_column_letter(16)].visible = show_column_current_stock
        if not show_column_reference:
            ws.column_dimensions[get_column_letter(3)].visible = False
    return wb


def admin_export(request, queryset):
    permanence = queryset.first()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = ("%s - %s.xlsx" % (_("Stock"), permanence)).encode('ascii', errors='replace').replace('?', '_')
    # filename = (_("Stock") + " - " + permanence.__str__() + '.xlsx').encode('latin-1', errors='ignore')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = export_stock(permanence=permanence, wb=None)

    if wb is not None:
        wb.save(response)
        return response
    else:
        return None


@transaction.atomic
def import_stock_sheet(worksheet, permanence=None):
    error = False
    error_msg = None
    if permanence.status < PERMANENCE_DONE:
        header = get_header(worksheet)
        if header:
            row_num = 1
            row = get_row(worksheet, header, row_num)
            while row and not error:
                try:
                    # with transaction.atomic():
                    stock = None if row[_('Initial stock')] is None else Decimal(row[_('Initial stock')]).quantize(THREE_DECIMALS)
                    add_2_stock = None if row[_('Add 2 stock')] is None else Decimal(row[_('Add 2 stock')]).quantize(THREE_DECIMALS)
                    if stock is not None:
                        producer_id = None if row[_('Id')] is None else Decimal(row[_('Id')])
                        offer_item_id = None if row[_('OfferItem')] is None else Decimal(row[_('OfferItem')])
                        offer_item = OfferItem.objects.filter(
                            id=offer_item_id,
                            permanence_id=permanence.id,
                            producer_id=producer_id
                        ).order_by().first()
                        if offer_item is not None \
                                and (offer_item.stock != stock or offer_item.add_2_stock != add_2_stock)  :
                            offer_item.stock = stock
                            offer_item.add_2_stock = add_2_stock
                            offer_item.save()
                            Product.objects.filter(
                                id=offer_item.product_id,
                                producer_id=producer_id
                            ).update(stock=stock)
                    row_num += 1
                    row = get_row(worksheet, header, row_num)
                except KeyError, e:
                    # Missing field
                    error = True
                    error_msg = _("Row %(row_num)d : A required column is missing.") % {'row_num': row_num + 1}
                except Exception, e:
                    error = True
                    error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
    else:
        error = True
        error_msg = _("The status of this permanence prohibit you to update the stock.")
    return error, error_msg


def handle_uploaded_file(request, queryset, file_to_import):
    error = False
    error_msg = None
    wb = load_workbook(file_to_import)
    if wb is not None:
        ws = None
        permanence=queryset.first()
        ws_sc_name = cap(_("Stock check"), 31).decode('utf-8')
        for sheet in wb.worksheets:
            if ws_sc_name == sheet.title:
                ws = sheet
        if ws is not None:
            error, error_msg = import_stock_sheet(
                ws,
                permanence=permanence
            )
            # ws = wb.get_sheet_by_name(ws_sc_name)
            if error:
                error_msg = cap(permanence.__str__(), 31) + " > " + error_msg
    return error, error_msg


def admin_import(admin_ui, admin, request, queryset, action):
    return import_xslx_view(admin_ui, admin, request, queryset, handle_uploaded_file, action=action)
