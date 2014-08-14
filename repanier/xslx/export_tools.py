# -*- coding: utf-8 -*-
from const import *
from django.contrib.sites.models import Site
from django.utils.translation import ugettext_lazy as _
from openpyxl.cell import get_column_letter
from openpyxl.style import NumberFormat
from repanier.models import Customer
from repanier.models import LUT_DepartmentForCustomer
from repanier.models import LUT_ProductionMode
from repanier.models import Producer
from repanier.models import Staff
from repanier.tools import cap


def worksheet_setup_a4(worksheet, title1, title2):
    worksheet.title = unicode(cap(title1, 31), "utf8")
    worksheet.page_setup.paperSize = worksheet.PAPERSIZE_A4
    worksheet.page_setup.fitToPage = True
    worksheet.page_setup.fitToHeight = 0
    worksheet.page_setup.fitToWidth = 1
    worksheet.print_gridlines = True
    worksheet.add_print_title(1, rows_or_cols='rows')
    worksheet.freeze_panes = 'A2'
    worksheet.header_footer.left_header.text = Site.objects.get_current().name
    worksheet.header_footer.left_footer.text = title2
    worksheet.header_footer.center_footer.text = title1
    worksheet.header_footer.right_footer.text = 'Page &[Page]/&[Pages]'
    orders_responsible = Staff.objects.filter(is_reply_to_order_email=True, is_active=True).order_by().first()
    invoices_responsible = Staff.objects.filter(is_reply_to_invoice_email=True, is_active=True).order_by().first()
    s1 = ""
    if orders_responsible:
        c = orders_responsible.customer_responsible
        if c != None:
            s1 = unicode(_("Orders")) + ": " + c.long_basket_name + ", " + c.phone1
    s2 = ""
    if invoices_responsible:
        c = invoices_responsible.customer_responsible
        if c != None:
            s2 = unicode(_("Invoices")) + ": " + c.long_basket_name + ", " + c.phone1
    separator = chr(10) + " "
    worksheet.header_footer.right_header.text = separator.join((s1, s2))
    return worksheet


def worksheet_setup_portait_a4(worksheet, title1, title2):
    worksheet = worksheet_setup_a4(worksheet, title1, title2)
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    return worksheet.title


def worksheet_setup_landscape_a4(worksheet, title1, title2):
    worksheet = worksheet_setup_a4(worksheet, title1, title2)
    worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
    return worksheet.title


def worksheet_set_header(worksheet, row_num, header):
    for col_num in xrange(len(header)):
        c = worksheet.cell(row=row_num, column=col_num)
        c.value = header[col_num][ROW_TITLE]
        c.style.font.bold = True
        c.style.alignment.wrap_text = False
        worksheet.column_dimensions[get_column_letter(col_num + 1)].width = header[col_num][ROW_WIDTH]
        if header[col_num][ROW_TITLE] == unicode(_("Id")):
            worksheet.column_dimensions[get_column_letter(col_num + 1)].visible = False


def get_validation_formula(wb=None, valid_values=None):
    if valid_values:

        ws_dv_name = cap(unicode(_("data validation")), 31)
        ws_dv = wb.get_sheet_by_name(ws_dv_name)
        if ws_dv == None:
            ws_dv = wb.create_sheet(index=0)
            worksheet_setup_landscape_a4(ws_dv, ws_dv_name, "")
        col_dv = 0
        c = ws_dv.cell(row=0, column=col_dv)
        while (c.value != None ) and (col_dv < 20):
            col_dv += 1
            c = ws_dv.cell(row=0, column=col_dv)

        col_letter_dv = get_column_letter(col_dv + 1)
        row_num = 0
        for v in valid_values:
            c = ws_dv.cell(row=row_num, column=col_dv)
            c.value = v
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            row_num += 1
        return "'%s'!$%s$1:$%s$%s" % (ws_dv_name, col_letter_dv, col_letter_dv, row_num + 1)
    else:
        return ""


def get_id_2_customer_dict():
    id_2_customer_dict = {}
    customer_set = Customer.objects.filter(is_active=True).order_by()
    for customer in customer_set:
        id_2_customer_dict[customer.id] = customer.short_basket_name
    return id_2_customer_dict


def get_id_2_producer_dict():
    id_2_producer_dict = {}
    producer_set = Producer.objects.filter(is_active=True).order_by()
    for producer in producer_set:
        id_2_producer_dict[producer.id] = producer.short_profile_name
    return id_2_producer_dict


def get_id_2_producer_vat_level_dict():
    id_2_producer_vat_level_dict = {}
    producer_set = Producer.objects.filter(is_active=True).order_by()
    for producer in producer_set:
        id_2_producer_vat_level_dict[producer.id] = producer.vat_level
    return id_2_producer_vat_level_dict


def get_id_2_producer_price_list_multiplier_dict():
    id_2_producer_price_list_multiplier_dict = {}
    producer_set = Producer.objects.filter(is_active=True).order_by()
    for producer in producer_set:
        id_2_producer_price_list_multiplier_dict[producer.id] = producer.price_list_multiplier
    return id_2_producer_price_list_multiplier_dict


def get_id_2_department_for_customer_dict():
    id_2_department_for_customer_dict = {}
    department_for_customer_set = LUT_DepartmentForCustomer.objects.filter(is_active=True).order_by()
    for department_for_customer in department_for_customer_set:
        id_2_department_for_customer_dict[department_for_customer.id] = department_for_customer.short_name
    return id_2_department_for_customer_dict


def get_id_2_production_mode_dict():
    id_2_production_mode_dict = {}
    production_mode_set = LUT_ProductionMode.objects.filter(is_active=True).order_by()
    for production_mode in production_mode_set:
        id_2_production_mode_dict[production_mode.id] = production_mode.short_name
    return id_2_production_mode_dict