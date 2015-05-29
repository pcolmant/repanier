# -*- coding: utf-8
from __future__ import unicode_literals
from django.db import transaction
from django.utils.translation import ugettext_lazy as _
from django.http import HttpResponse
from django.utils import timezone
from django.utils import translation
from openpyxl import load_workbook
from openpyxl.datavalidation import DataValidation, ValidationType, ValidationOperator
from openpyxl.style import Border
from openpyxl.style import NumberFormat
from openpyxl.workbook import Workbook

from repanier.tools import recalculate_order_amount
from export_tools import *
from import_tools import *
from repanier.const import *
from repanier.models import LUT_DepartmentForCustomer
from repanier.models import LUT_ProductionMode
from repanier.models import Permanence
from repanier.models import Product
from repanier.tools import cap
from views import import_xslx_view


def export(producer, wb=None):
    if wb is None:
        wb = Workbook()
        ws = wb.get_active_sheet()
    else:
        ws = wb.create_sheet()

    row_num = 0
    product_set = Product.objects.filter(
        producer_id=producer.id, is_active=True
    )
    product_save = None
    now = timezone.localtime(timezone.now())
    worksheet_setup_landscape_a4(ws, producer.short_profile_name, now.strftime('%d-%m-%Y %H:%M'))
    for product in product_set:
        row = [
            (_("Id"), 10, product.id, '#,##0', False),
            (_("department_for_customer"), 15, product.department_for_customer.short_name,
             NumberFormat.FORMAT_TEXT, False),
            (_("is_into_offer"), 7, _("Yes") if product.is_into_offer else None,
             NumberFormat.FORMAT_TEXT, False),
            (_("long_name"), 60, product.long_name, NumberFormat.FORMAT_TEXT, False),
            (_("order unit"), 15, product.get_order_unit_display(), NumberFormat.FORMAT_TEXT, False),
            (_("wrapped"), 7, _("Yes") if product.wrapped else None,
             NumberFormat.FORMAT_TEXT, False),
            (_("order_average_weight"), 10, product.order_average_weight, '#,##0.???', False),
            (_("producer unit price"), 10, product.producer_unit_price,
             '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
            (_("deposit"), 10, product.unit_deposit, '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ',
             False),
            (_("vat or compensation"), 10, product.get_vat_level_display(), NumberFormat.FORMAT_TEXT, False),
            (_("customer_minimum_order_quantity"), 10, product.customer_minimum_order_quantity, '#,##0.???',
             False),
            (
                _("customer_increment_order_quantity"), 10, product.customer_increment_order_quantity,
                '#,##0.???',
                False),
            (
                _("customer_alert_order_quantity"), 10, product.customer_alert_order_quantity, '#,##0.???',
                False),
        ]

        if row_num == 0:
            worksheet_set_header(ws, row)
            row_num += 1

        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.value = "%s" % (row[col_num][ROW_VALUE])
            c.style.number_format.format_code = row[col_num][ROW_FORMAT]
            if row[col_num][ROW_BOX]:
                c.style.borders.top.border_style = Border.BORDER_THIN
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c.style.borders.left.border_style = Border.BORDER_THIN
                c.style.borders.right.border_style = Border.BORDER_THIN
            else:
                c.style.borders.bottom.border_style = Border.BORDER_HAIR
            if product_save != product.department_for_customer.id:
                c.style.borders.top.border_style = Border.BORDER_THIN
        if product_save != product.department_for_customer.id:
            product_save = product.department_for_customer.id
        row_num += 1
        # Now, for helping the user encoding new purchases
    row = [
        (_("Id"), 10, "", NumberFormat.FORMAT_TEXT),
        (_("department_for_customer"), 15, "", NumberFormat.FORMAT_TEXT),
        (_("is_into_offer"), 7, "", NumberFormat.FORMAT_TEXT),
        (_("long_name"), 60, "", NumberFormat.FORMAT_TEXT),
        (_("order unit"), 15, "", NumberFormat.FORMAT_TEXT),
        (_("wrapped"), 7, "", NumberFormat.FORMAT_TEXT),
        (_("order_average_weight"), 10, "", '#,##0.???'),
        (_("producer unit price"), 10, "", '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
        (_("deposit"), 10, "", '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
        (_("vat or compensation"), 10, "", NumberFormat.FORMAT_TEXT),
        (_("customer_minimum_order_quantity"), 10, "", '#,##0.???'),
        (_("customer_increment_order_quantity"), 10, "", '#,##0.???'),
        (_("customer_alert_order_quantity"), 10, "", '#,##0.???'),
    ]

    if row_num == 0:
        # add a header if there is no previous movement.
        worksheet_set_header(ws, row)
        row_num += 1

    # Data validation Id
    dv = DataValidation(ValidationType.WHOLE,
                        ValidationOperator.EQUAL,
                        0)
    ws.add_data_validation(dv)
    dv.ranges.append('A%s:A5000' % (row_num + 1))
    # Data validation Department for customer
    # list of Department for customer
    valid_values = []
    department_for_customer_set = LUT_DepartmentForCustomer.objects.filter(is_active=True).order_by()
    for department_for_customer in department_for_customer_set:
        valid_values.append(department_for_customer.short_name)
    valid_values.sort()
    dv = DataValidation(ValidationType.LIST, formula1=get_validation_formula(wb=wb, valid_values=valid_values),
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.ranges.append('B2:B5000')
    # Data validation Unit
    # List of Unit
    valid_values = []
    for record in LUT_PRODUCT_ORDER_UNIT:
        valid_values.append(record[1])
    dv = DataValidation(ValidationType.LIST, formula1=get_validation_formula(wb=wb, valid_values=valid_values),
                        allow_blank=False)
    ws.add_data_validation(dv)
    dv.ranges.append('E2:E5000')
    # Data validation Yes/
    # List of Yes/
    valid_values = [_('Yes'), ]
    dv = DataValidation(ValidationType.LIST, formula1=get_validation_formula(wb=wb, valid_values=valid_values),
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.ranges.append('C2:C5000')
    # List of Yes/
    valid_values = [_('Yes'), ]
    dv = DataValidation(ValidationType.LIST, formula1=get_validation_formula(wb=wb, valid_values=valid_values),
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.ranges.append('F2:F5000')
    # Data validation qty / weight
    dv = DataValidation(ValidationType.DECIMAL,
                        ValidationOperator.GREATER_THAN_OR_EQUAL,
                        0)
    ws.add_data_validation(dv)
    dv.ranges.append('G2:I5000')
    dv = DataValidation(ValidationType.DECIMAL,
                        ValidationOperator.GREATER_THAN_OR_EQUAL,
                        0)
    ws.add_data_validation(dv)
    dv.ranges.append('K2:M5000')
    # Data validation Vat or Compensation
    # List of Vat or Compensation
    valid_values = []
    for record in LUT_VAT:
        valid_values.append(record[1])
    dv = DataValidation(ValidationType.LIST, formula1=get_validation_formula(wb=wb, valid_values=valid_values),
                        allow_blank=True)
    ws.add_data_validation(dv)
    dv.ranges.append('J2:J5000')
    # End of data validation

    # Add formating for empty cells.
    for row_num in xrange(row_num, row_num + 30):
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.style.number_format.format_code = row[col_num][ROW_FORMAT]

    # if row_num > 0:
    # Let's create a new sheet for the next producer
    wb.worksheets.reverse()
    return wb


def admin_export(request, queryset):
    queryset = queryset.order_by("-short_profile_name")
    wb = None
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = ("%s.xlsx" % (_("products"))).encode('ascii', errors='replace').replace('?', '_')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    for producer in queryset:
        wb = export(producer=producer, wb=wb)
    wb.save(response)
    return response

@transaction.atomic
def import_product_sheet(worksheet, producer=None,
                         department_for_customer_2_id_dict=None
):
    vat_level_dict = dict(LUT_VAT_REVERSE)
    order_unit_dict = dict(LUT_PRODUCT_ORDER_UNIT_REVERSE)
    error = False
    error_msg = None
    header = get_header(worksheet)
    if header:
        row_num = 1
        row = get_row(worksheet, header, row_num)
        while row and not error:
            try:
                row_id = None
                if row[_('Id')] is not None:
                    row_id = Decimal(row[_('Id')])

                if row[_('department_for_customer')] in department_for_customer_2_id_dict:
                    department_for_customer_id = department_for_customer_2_id_dict[row[_('department_for_customer')]]
                else:
                    error = True
                    # transaction.rollback()
                    error_msg = _("Row %(row_num)d : No valid departement for customer") % {'row_num': row_num + 1}
                    break

                producer_unit_price = None if row[_('producer_unit_price')] is None else Decimal(
                    row[_('producer_unit_price')])
                unit_deposit = None if row[_('deposit')] is None else Decimal(row[_('deposit')])
                order_average_weight = None if row[_('order_average_weight')] is None else Decimal(
                    row[_('order_average_weight')])
                customer_minimum_order_quantity = None if row[_(
                    'customer_minimum_order_quantity')] is None else Decimal(row[_('customer_minimum_order_quantity')])
                customer_increment_order_quantity = None if row[_(
                    'customer_increment_order_quantity')] is None else Decimal(
                    row[_('customer_increment_order_quantity')])
                customer_alert_order_quantity = None if row[_('customer_alert_order_quantity')] is None else Decimal(
                    row[_('customer_alert_order_quantity')])

                order_unit = None
                if row[_("order unit")] in order_unit_dict:
                    order_unit = order_unit_dict[row[_("order unit")]]
                if order_unit is None:
                    order_unit = PRODUCT_ORDER_UNIT_PC

                long_name = cap(row[_('long_name')], 100)
                if row_id is not None:
                    product = Product.objects.filter(
                        producer_id=producer.id,
                        id=row_id
                    ).order_by().first()
                else:
                    product = Product.objects.filter(
                        producer_id=producer.id,
                        translations__long_name=long_name,
                        translations__language_code=translation.get_language()
                    ).order_by().first()
                # print(long_name.encode('utf8'))
                if product:
                    if row_id == product.id:
                        # Detect VAT LEVEL. Fall back on product.
                        if row[_("vat or compensation")] in vat_level_dict:
                            vat_level = vat_level_dict[row[_("vat or compensation")]]
                        elif product is not None:
                            vat_level = product.vat_level
                        else:
                            vat_level = producer.vat_level
                        # Let only update if the given id is the same as the product found id
                        product.producer_id = producer.id
                        product.long_name = long_name
                        product.department_for_customer_id = department_for_customer_id
                        product.order_unit = order_unit
                        if order_average_weight is not None:
                            product.order_average_weight = order_average_weight
                        if producer_unit_price is not None:
                            product.producer_unit_price = producer_unit_price
                        if unit_deposit is not None:
                            product.unit_deposit = unit_deposit
                        if customer_minimum_order_quantity is not None:
                            product.customer_minimum_order_quantity = customer_minimum_order_quantity
                        if customer_increment_order_quantity is not None:
                            product.customer_increment_order_quantity = customer_increment_order_quantity
                        if customer_alert_order_quantity is not None:
                            product.customer_alert_order_quantity = customer_alert_order_quantity
                        product.is_into_offer = (row[_('is_into_offer')] is not None)
                        product.wrapped = (row[_('wrapped')] is not None)
                        product.vat_level = vat_level
                        product.is_active = True
                        # product.product_reorder = product_reorder
                        product.save()
                    else:
                        error = True
                        # transaction.rollback()
                        if row[_('Id')] is None:
                            error_msg = _(
                                "Row %(row_num)d : No id given, or the product %(producer)s - %(product)s already exist.") % {
                                            'row_num': row_num + 1, 'producer': producer.short_profile_name,
                                            'product': row[_('long_name')]}
                        else:
                            error_msg = _(
                                "Row %(row_num)d : The given id %(record_id)s is not the id of %(producer)s - %(product)s.") % {
                                            'row_num': row_num + 1, 'record_id': row[_('Id')],
                                            'producer': producer.short_profile_name, 'product': row[_('long_name')]}
                        break
                else:
                    if row_id is None:
                        # Let only create product if non id in the row
                        if order_average_weight is None:
                            order_average_weight = 0
                        if producer_unit_price is None:
                            producer_unit_price = 0
                        if unit_deposit is None:
                            unit_deposit = 0
                        if customer_minimum_order_quantity is None:
                            customer_minimum_order_quantity = 0
                        if customer_increment_order_quantity is None:
                            customer_increment_order_quantity = 0
                        if customer_alert_order_quantity is None:
                            customer_alert_order_quantity = 0
                        Product.objects.create(
                            producer=producer,
                            long_name=long_name,
                            department_for_customer_id=department_for_customer_id,
                            order_unit=order_unit,
                            order_average_weight=order_average_weight,
                            producer_unit_price=producer_unit_price,
                            unit_deposit=unit_deposit,
                            customer_minimum_order_quantity=customer_minimum_order_quantity,
                            customer_increment_order_quantity=customer_increment_order_quantity,
                            customer_alert_order_quantity=customer_alert_order_quantity,
                            is_into_offer=(row[_('is_into_offer')] is not None),
                            wrapped=(row[_('wrapped')] is not None),
                            is_active=True,  # product_reorder = product_reorder
                        )
                    else:
                        error = True
                        # transaction.rollback()
                        error_msg = _(
                            "Row %(row_num)d : The given id %(record_id)s is not the id of %(producer)s - %(product)s.") % {
                                        'row_num': row_num + 1, 'record_id': row[_('Id')],
                                        'producer': producer.short_profile_name, 'product': row[_('long_name')]}
                        break

                row_num += 1
                row = get_row(worksheet, header, row_num)
            except KeyError, e:
                # Missing field
                error = True
                error_msg = _("Row %(row_num)d : A required column is missing.") % {'row_num': row_num + 1}
            except Exception, e:
                error = True
                error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
    return error, error_msg


def handle_uploaded_file(request, queryset, file_to_import):
    error = False
    error_msg = None
    wb = load_workbook(file_to_import)
    # dict for performance optimisation purpose : read the DB only once
    department_for_customer_2_id_dict = get_department_for_customer_2_id_dict()
    for producer in queryset:
        error, error_msg = import_product_sheet(
            wb.get_sheet_by_name(cap(producer.short_profile_name, 31)),
            producer=producer,
            department_for_customer_2_id_dict=department_for_customer_2_id_dict
        )
        if error:
            error_msg = producer.short_profile_name + " > " + error_msg
            break
    return error, error_msg


def admin_import(admin_ui, admin, request, queryset, action):
    return import_xslx_view(admin_ui, admin, request, queryset, handle_uploaded_file, action=action)
