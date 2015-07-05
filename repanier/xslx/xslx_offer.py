# -*- coding: utf-8
from __future__ import unicode_literals
from django.utils.translation import ugettext_lazy as _
from django.utils.translation import ugettext as _not_lazy

from django.http import HttpResponse
from django.utils import translation
from openpyxl.style import Border
from openpyxl.style import NumberFormat
from openpyxl.workbook import Workbook

from export_tools import *
from repanier.models import Producer, OfferItem
from repanier.models import Product
from repanier.tools import *


def export(permanence, wb=None):
    if wb is None:
        wb = Workbook()
        ws = wb.get_active_sheet()
    else:
        ws = wb.create_sheet()
    worksheet_setup_landscape_a4(ws, _("Planned"), permanence)
    row_num = 0

    if permanence.status == PERMANENCE_PLANNED:

        producers_in_this_permanence = Producer.objects.filter(
            permanence=permanence, is_active=True)

        for product in Product.objects.filter(
                producer__in=producers_in_this_permanence, is_active=True, is_into_offer=True,
                translations__language_code=translation.get_language()).order_by(
                "producer__short_profile_name",
                "department_for_customer__tree_id",
                "department_for_customer__lft",
                "translations__long_name",
                "order_average_weight"):
            row = [
                (_("Producer"), 15, product.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
                (_("Department"), 15,
                 product.department_for_customer.short_name if product.department_for_customer is not None else "",
                 NumberFormat.FORMAT_TEXT,
                 False),
                (_("Product"), 60, product.get_long_name(), NumberFormat.FORMAT_TEXT, False),
                (_("Unit Price"), 10, product.producer_unit_price,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
                (_("deposit"), 10, product.unit_deposit,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
            ]

            if row_num == 0:
                worksheet_set_header(ws, row)
                row_num +=1

            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = "%s" % (row[col_num][ROW_VALUE])
                c.style.number_format.format_code = row[col_num][ROW_FORMAT]
                if row[col_num][ROW_BOX]:
                    c.style.borders.top.border_style = Border.BORDER_THIN
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c.style.borders.left.border_style = Border.BORDER_THIN
                    c.style.borders.right.border_style = Border.BORDER_THIN
                else:
                    c.style.borders.bottom.border_style = Border.BORDER_HAIR

            col_num = len(row)
            q_min = product.customer_minimum_order_quantity
            q_alert = product.customer_alert_order_quantity
            q_step = product.customer_increment_order_quantity
            # The q_min cannot be 0. In this case try to replace q_min by q_step.
            # In last ressort by q_alert.
            c = ws.cell(row=row_num, column=col_num)
            c.value = '---'
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 2.3
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            col_num += 1
            q_valid = q_min
            q_counter = 0  # Limit to avoid too long selection list
            while q_valid <= q_alert and q_counter <= 20:
                q_counter += 1
                c = ws.cell(row=row_num, column=col_num)
                qty_display, price_display, base_unit, unit, price = get_display(
                    q_valid,
                    product.order_average_weight,
                    product.order_unit,
                    product.customer_unit_price
                )
                c.value = qty_display + price_display + u" €"
                ws.column_dimensions[get_column_letter(col_num + 1)].width = 20
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                col_num += 1
                if q_valid < q_step:
                    # 1; 2; 4; 6; 8 ... q_min = 1; q_step = 2
                    # 0,5; 1; 2; 3 ... q_min = 0,5; q_step = 1
                    q_valid = q_step
                else:
                    # 1; 2; 3; 4 ... q_min = 1; q_step = 1
                    # 0,125; 0,175; 0,225 ... q_min = 0,125; q_step = 0,50
                    q_valid = q_valid + q_step

            row_num += 1

    if permanence.status == PERMANENCE_OPENED:

        for offer_item in OfferItem.objects.filter(permanence_id=permanence.id, is_active=True,
                product__translations__language_code=translation.get_language()).order_by(
                'producer__short_profile_name',
                'department_for_customer',
                'product__translations__long_name'):
            row = [
                (_("Producer"), 15, offer_item.producer.short_profile_name, NumberFormat.FORMAT_TEXT,
                 False),
                (_("Department"), 15,
                 offer_item.department_for_customer.short_name if offer_item.department_for_customer is not None else "",
                 NumberFormat.FORMAT_TEXT, False),
                (_("Product"), 60, offer_item.get_long_name(), NumberFormat.FORMAT_TEXT, False),
                (_("Unit Price"), 10, offer_item.producer_unit_price,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
                (_("deposit"), 10, offer_item.unit_deposit,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
            ]

            if row_num == 0:
                worksheet_set_header(ws, row)
                row_num += 1

            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = "%s" % (row[col_num][ROW_VALUE])
                c.style.number_format.format_code = row[col_num][ROW_FORMAT]
                if row[col_num][ROW_BOX]:
                    c.style.borders.top.border_style = Border.BORDER_THIN
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c.style.borders.left.border_style = Border.BORDER_THIN
                    c.style.borders.right.border_style = Border.BORDER_THIN
                else:
                    c.style.borders.bottom.border_style = Border.BORDER_HAIR

            col_num = len(row)
            q_min = offer_item.customer_minimum_order_quantity
            q_alert = offer_item.customer_alert_order_quantity
            q_step = offer_item.customer_increment_order_quantity
            c = ws.cell(row=row_num, column=col_num)
            c.value = '---'
            ws.column_dimensions[get_column_letter(col_num + 1)].width = 2.3
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            col_num += 1
            q_valid = q_min
            q_counter = 0  # Limit to avoid too long selection list
            while q_valid <= q_alert and q_counter <= 20:
                q_counter += 1
                c = ws.cell(row=row_num, column=col_num)
                qty_display, price_display, base_unit, unit, price = get_display(
                    q_valid,
                    offer_item.order_average_weight,
                    offer_item.order_unit,
                    offer_item.customer_unit_price
                )
                c.value = qty_display + price_display + u" €"
                ws.column_dimensions[get_column_letter(col_num + 1)].width = 20
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                col_num += 1
                if q_valid < q_step:
                    # 1; 2; 4; 6; 8 ... q_min = 1; q_step = 2
                    # 0,5; 1; 2; 3 ... q_min = 0,5; q_step = 1
                    q_valid = q_step
                else:
                    # 1; 2; 3; 4 ... q_min = 1; q_step = 1
                    # 0,125; 0,175; 0,225 ... q_min = 0,125; q_step = 0,50
                    q_valid = q_valid + q_step

            row_num += 1

    return wb


def admin_export(request, queryset):
    permanence = queryset.first()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = ("%s - %s.xlsx" % (_("Preview report"), permanence)).encode('ascii', errors='replace').replace('?', '_')
    # filename = (unicode(_("Preview")) + u" - " + permanence.__unicode__() + u'.xlsx').encode('latin-1', errors='ignore')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = export(permanence=permanence, wb=None)
    if wb is not None:
        wb.save(response)
    return response