# -*- coding: utf-8
from __future__ import unicode_literals
from django.conf import settings
from django.utils.translation import ugettext_lazy as _
from django.utils import translation
from django.http import HttpResponse
from openpyxl.style import Border
from openpyxl.style import Fill
from openpyxl.style import NumberFormat
from openpyxl.styles import Color
from openpyxl.workbook import Workbook
from django.contrib.sites.models import Site
from repanier.apps import repanier_settings

from export_tools import *
from repanier.const import *
from repanier.models import Customer, PurchaseOpened, OfferItem
from repanier.models import Permanence
from repanier.models import PermanenceBoard
from repanier.models import Producer
from repanier.models import Staff
from repanier.tools import get_display
from xslx_stock import export_stock
from decimal import *


def next_purchase(purchases):
    purchase = next_row(purchases)
    return purchase

def export(permanence, wb=None):
    translation.activate(settings.LANGUAGE_CODE)
    if wb is None:
        wb = Workbook()
        ws = wb.get_active_sheet()
    else:
        ws = wb.create_sheet()

    worksheet_setup_portrait_a4(ws, permanence, '')

    header = [
        (_("Basket"), 20),
        (_('Family'), 35),
        (_('Phone1'), 15),
        (_('Phone2'), 15),
    ]
    row_num = 0
    worksheet_set_header(ws, header)
    row_num += 1
    # Customer info
    customer_set = Customer.objects.filter(
        purchase__permanence_id=permanence.id, represent_this_buyinggroup=False).distinct()
    for customer in customer_set:
        row = [
            customer.short_basket_name,
            customer.long_basket_name,
            customer.phone1,
            customer.phone2
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.value = row[col_num]
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            c.style.alignment.wrap_text = True
        row_num += 1
    c = ws.cell(row=row_num, column=1)
    c.value = "%s" % (_('Permanence Board Member List'))
    c.style.alignment.wrap_text = False
    c.style.font.bold = True
    row_num += 1
    # Permanence board info
    permanence_date_save = None
    next_permanence_set = Permanence.objects.filter(permanence_date__gte=permanence.permanence_date).order_by(
        "permanence_date")[:3]
    for next_permanence in next_permanence_set:
        for permanenceboard in PermanenceBoard.objects.filter(
                permanence_id=next_permanence.id).order_by(
                "permanence_role__tree_id",
                "permanence_role__lft"
        ):
            c = permanenceboard.customer
            if c is not None:
                row = [
                    (next_permanence.permanence_date, NumberFormat.FORMAT_DATE_DMYSLASH,),
                    (c.long_basket_name, NumberFormat.FORMAT_TEXT,),
                    (c.phone1, NumberFormat.FORMAT_TEXT,),
                    (c.phone2, NumberFormat.FORMAT_TEXT,),
                    (permanenceboard.permanence_role.short_name, NumberFormat.FORMAT_TEXT),
                ]
                for col_num in xrange(len(row)):
                    c = ws.cell(row=row_num, column=col_num)
                    c.value = "%s" % (row[col_num][0])
                    c.style.number_format.format_code = row[col_num][1]
                    c.style.alignment.wrap_text = False
                    if permanence_date_save != next_permanence.permanence_date:
                        c.style.font.bold = True
                        permanence_date_save = next_permanence.permanence_date
                row_num += 1
    c = ws.cell(row=row_num, column=1)
    c.value = "%s" % (_('Staff Member List'))
    c.style.alignment.wrap_text = False
    c.style.font.bold = True
    row_num += 1
    for staff in Staff.objects.filter(is_active=True).order_by():
        c = staff.customer_responsible
        if c is not None:
            row = [
                staff.long_name,
                c.long_basket_name,
                c.phone1,
                c.phone2
            ]
            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = row[col_num]
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                c.style.alignment.wrap_text = True
            row_num += 1

    c = ws.cell(row=row_num, column=1)
    c.value = "%s" % (_('producers'))
    c.style.alignment.wrap_text = False
    c.style.font.bold = True
    row_num += 1
    # Producer info
    for producer in Producer.objects.filter(permanence=permanence).order_by("short_profile_name"):
        row = [
            producer.short_profile_name,
            producer.long_profile_name,
            producer.phone1,
            producer.phone2
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True
        row_num += 1

    # Customer label
    ws = wb.create_sheet()
    worksheet_setup_portrait_a4(ws, _('Label'), permanence, add_print_title=False)
    row_num = 0
    customer_set = Customer.objects.filter(
        purchase__permanence_id=permanence.id, represent_this_buyinggroup=False).distinct()
    for customer in customer_set:
        c = ws.cell(row=row_num, column=0)
        c.value = customer.short_basket_name
        c.style.font.size = 36
        c.style.alignment.wrap_text = False
        c.style.borders.top.border_style = Border.BORDER_THIN
        c.style.borders.bottom.border_style = Border.BORDER_THIN
        c.style.borders.left.border_style = Border.BORDER_THIN
        c.style.borders.right.border_style = Border.BORDER_THIN
        c.style.alignment.vertical = 'center'
        c.style.alignment.horizontal = 'center'
        row_num += 1
        ws.row_dimensions[row_num].height = 60
        c = ws.cell(row=row_num, column=0)
        row_num += 1
        ws.row_dimensions[row_num].height = 5
    if row_num > 0:
        ws.column_dimensions[get_column_letter(1)].width = 120

    return wb

def export_preparation(permanence, wb=None):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Id"), 10),
        (_("OfferItem"), 5),
        (_("Date"), 15),
        (_("Placement"), 7),
        (_("producer"), 15),
        (_("product"), 60),
        (_("customer"), 15),
        (_("quantity ordered"), 10),
        (_("Unit"), 10),
        (_("Prepared"), 22),
        (_("To distribute"), 10),
    ]

    producers = Producer.objects.filter(
        purchase__permanence_id=permanence.id
    ).distinct().iterator()
    producer = next_row(producers)
    if producer is not None:
        wb, ws = new_landscape_a4_sheet(
            wb,
            _("Preparation"),
            permanence,
            header
        )
        row_num = 1
        while producer is not None:
            producer_save = producer
            if producer.invoice_by_basket:
                purchases = PurchaseOpened.objects.filter(
                    permanence_id=permanence.id,
                    producer_id=producer.id,
                    offer_item__translations__language_code=translation.get_language()
                ).exclude(
                    quantity_ordered__lte=DECIMAL_ZERO
                ).order_by(
                    "customer__short_basket_name",
                    "offer_item__translations__order_sort_order"
                ).iterator()
                purchase = next_purchase(purchases)

                while purchase is not None:
                    c = ws.cell(row=row_num + 1, column=4)
                    c.value = producer_save.short_profile_name
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.font.bold = True
                    date_save = purchase.permanence_date
                    while purchase is not None and date_save == purchase.permanence_date:
                        customer_save = purchase.customer
                        c = ws.cell(row=row_num + 1, column=6)
                        c.style.font.bold = True
                        count_purchase = 0
                        purchases_price = DECIMAL_ZERO
                        purchases_price_formula = "="
                        while purchase is not None and date_save == purchase.permanence_date \
                                and customer_save == purchase.customer:
                            department_for_customer_save = purchase.offer_item.department_for_customer
                            c = ws.cell(row=row_num, column=5)
                            c.value = department_for_customer_save.short_name \
                                if department_for_customer_save is not None else "---"
                            c.style.font.italic = True
                            c.style.alignment.horizontal = 'right'
                            for col_num in xrange(11):
                                c = ws.cell(row=row_num, column=col_num)
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                            row_num += 1
                            while purchase is not None and date_save == purchase.permanence_date \
                                    and customer_save == purchase.customer \
                                    and department_for_customer_save == purchase.offer_item.department_for_customer:
                                qty_ordered, price_display, base_unit, unit, price = get_display(
                                    purchase.quantity_ordered,
                                    purchase.offer_item.order_average_weight,
                                    purchase.offer_item.order_unit,
                                    0
                                )
                                c = ws.cell(row=row_num, column=0)
                                c.value = purchase.id
                                c = ws.cell(row=row_num, column=1)
                                c.value = purchase.offer_item_id
                                c = ws.cell(row=row_num, column=2)
                                c.value = date_save
                                c.style.number_format.format_code = NumberFormat.FORMAT_DATE_DDMMYYYY
                                c = ws.cell(row=row_num, column=3)
                                c.value = purchase.offer_item.get_placement_display()
                                c = ws.cell(row=row_num, column=4)
                                c.value = producer_save.short_profile_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=5)
                                c.value = purchase.get_long_name()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=6)
                                c.value = customer_save.short_basket_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=7)
                                c.value = purchase.quantity_ordered
                                c.style.number_format.format_code = '#,##0.????'
                                c.style.font.color = Color(Color.BLUE)
                                ws.conditional_formatting.addCellIs(
                                    get_column_letter(8) + str(row_num+1), 'notEqual',
                                    [str(purchase.quantity_ordered)], True, wb,
                                    None, None, yellowFill
                                )
                                c = ws.cell(row=row_num, column=8)
                                c.value = "%s" % base_unit
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                customer_unit_price = purchase.get_customer_unit_price()
                                if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                                    quantity_ordered = purchase.quantity_ordered * purchase.offer_item.order_average_weight
                                    purchases_price_formula += '+ROUND(H%s*%s*%s,2)' % \
                                        (row_num + 1, purchase.offer_item.order_average_weight,
                                         customer_unit_price + purchase.offer_item.unit_deposit )
                                    c = ws.cell(row=row_num, column=9)
                                    if purchase.offer_item.wrapped:
                                        c.value = "%s" % (_('€ :'))
                                    else:
                                        c.value = "%s" % (_('kg :'))
                                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                else:
                                    quantity_ordered = purchase.quantity_ordered
                                    purchases_price_formula += '+ROUND(H%s*%s,2)' % \
                                        (row_num + 1, customer_unit_price + purchase.offer_item.unit_deposit )
                                    if purchase.offer_item.wrapped:
                                        c = ws.cell(row=row_num, column=9)
                                        c.value = "%s" % (_('€ :'))
                                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                purchases_price += (quantity_ordered *
                                    (customer_unit_price + purchase.offer_item.unit_deposit)
                                    ).quantize(TWO_DECIMALS)

                                count_purchase += 1
                                delta = 6
                                for col_num in xrange(4):
                                    c = ws.cell(row=row_num, column=delta+col_num)
                                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                                row_num += 1

                                purchase = next_purchase(purchases)
                        if count_purchase > 1:
                            c = ws.cell(row=row_num, column=10)
                            # c.value = '=SUM(G%s:G%s)' % (row_start_purchase, row_num)
                            c.value = purchases_price_formula
                            c.style.number_format.format_code = \
                                '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                            c.style.font.color = Color(Color.BLUE)
                            ws.conditional_formatting.addCellIs(
                                get_column_letter(11) + str(row_num+1), 'notEqual',
                                [str(purchases_price)], True, wb,
                                None, None, yellowFill
                            )
                            row_num += 1
            else:
                # Using quantity_for_preparation_sort_order the order is by customer__short_basket_name if the product
                # is to be distributed by piece, otherwise by lower qty first.
                purchases = PurchaseOpened.objects.filter(
                    permanence_id=permanence.id,
                    producer_id=producer.id,
                    offer_item__translations__language_code=translation.get_language()
                ).order_by(  # "product__placement",
                    "offer_item__translations__order_sort_order",
                    "quantity_for_preparation_sort_order",
                    "customer__short_basket_name"
                ).iterator()
                purchase = next_purchase(purchases)
                while purchase is not None:
                    date_save = purchase.permanence_date
                    c = ws.cell(row=row_num + 1, column=4)
                    c.value = producer_save.short_profile_name
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.font.bold = True
                    while purchase is not None and date_save == purchase.permanence_date:
                        department_for_customer_save = purchase.offer_item.department_for_customer
                        c = ws.cell(row=row_num, column=5)
                        c.value = department_for_customer_save.short_name \
                            if department_for_customer_save is not None else "---"
                        c.style.font.italic = True
                        c.style.alignment.horizontal = 'right'
                        while purchase is not None and date_save == purchase.permanence_date \
                            and producer_save == purchase.producer \
                            and department_for_customer_save == purchase.offer_item.department_for_customer:
                            offer_item_save = purchase.offer_item
                            row_start_offer_item = row_num + 2
                            count_offer_item = 0
                            for col_num in xrange(11):
                                c = ws.cell(row=row_num, column=col_num)
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                            purchases_quantity = DECIMAL_ZERO
                            row_num += 1
                            while purchase is not None and offer_item_save == purchase.offer_item:
                                # if purchase.quantity_ordered > DECIMAL_ZERO \
                                #         or purchase.offer_item.order_unit >= PRODUCT_ORDER_UNIT_DEPOSIT:
                                qty_ordered, price_display, base_unit, unit, price = get_display(
                                    purchase.quantity_ordered,
                                    purchase.offer_item.order_average_weight,
                                    purchase.offer_item.order_unit,
                                    0
                                )
                                c = ws.cell(row=row_num, column=0)
                                c.value = purchase.id
                                c = ws.cell(row=row_num, column=1)
                                c.value = purchase.offer_item_id
                                c = ws.cell(row=row_num, column=2)
                                c.value = date_save
                                c.style.number_format.format_code = NumberFormat.FORMAT_DATE_DDMMYYYY
                                c = ws.cell(row=row_num, column=3)
                                c.value = purchase.offer_item.get_placement_display()
                                c = ws.cell(row=row_num, column=4)
                                c.value = producer_save.short_profile_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=5)
                                c.value = purchase.get_long_name()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=6)
                                c.value = purchase.customer.short_basket_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=7)
                                c.value = purchase.quantity_ordered
                                c.style.number_format.format_code = '#,##0.????'
                                c.style.font.color = Color(Color.BLUE)
                                ws.conditional_formatting.addCellIs(
                                    get_column_letter(8) + str(row_num+1), 'notEqual',
                                    [str(purchase.quantity_ordered)], True, wb,
                                    None, None, yellowFill
                                )
                                c = ws.cell(row=row_num, column=8)
                                c.value = "%s" % base_unit
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                                    c = ws.cell(row=row_num, column=9)
                                    if purchase.offer_item.wrapped:
                                        c.value = "%s" % (_('€ :'))
                                    else:
                                        c.value = "%s" % (_('kg :'))
                                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                else:
                                    if purchase.offer_item.wrapped:
                                        c = ws.cell(row=row_num, column=9)
                                        c.value = "%s" % (_('€ :'))
                                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                purchases_quantity += purchase.quantity_ordered
                                count_offer_item += 1
                                delta = 6
                                for col_num in xrange(4):
                                    c = ws.cell(row=row_num, column=delta+col_num)
                                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                                row_num += 1

                                purchase = next_purchase(purchases)
                            if count_offer_item > 1:
                                c = ws.cell(row=row_num, column=10)
                                c.value = '=SUM(H%s:H%s)' % (row_start_offer_item, row_num)
                                c.style.number_format.format_code = '#,##0.????'
                                if not offer_item_save.wrapped and offer_item_save.order_unit in [PRODUCT_ORDER_UNIT_KG, PRODUCT_ORDER_UNIT_PC_KG]:
                                    c.style.font.color = Color(Color.BLUE)
                                ws.conditional_formatting.addCellIs(
                                    get_column_letter(11) + str(row_num+1), 'notEqual',
                                    [str(purchases_quantity)], True, wb,
                                    None, None, yellowFill
                                )
                                if purchase is None or department_for_customer_save != purchase.offer_item.department_for_customer:
                                    row_num += 1

            # row_num += 1
            for col_num in xrange(11):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
            row_num += 2
            producer = next_row(producers)
        ws.column_dimensions[get_column_letter(2)].visible = False
        ws.column_dimensions[get_column_letter(3)].visible = False

    return wb


def export_producer_by_product(permanence, producer, wb=None):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    translation.activate(producer.language)

    header = [
        (_("Basket"), 20),
        (_("Quantity"), 10),
        (_("Reference"), 20),
        (_("Product"), 60),
        (_("Unit Price"), 10),
        (_("Deposit"), 10),
        (_("Total Price"), 12),
    ]
    show_column_reference = False
    hide_column_short_basket_name = True
    hide_column_unit_deposit = True
    formula_main_total = []
    offer_items = OfferItem.objects.filter(
        permanence_id=permanence.id,
        producer_id=producer.id,
        translations__language_code=translation.get_language()
    ).exclude(
        quantity_invoiced=DECIMAL_ZERO
    ).order_by(
        "translations__order_sort_order"
    ).iterator()
    offer_item = next_row(offer_items)
    if offer_item:
        wb, ws = new_landscape_a4_sheet(
            wb,
            "%s %s" % (producer.short_profile_name, _("by product")),
            permanence,
            header
        )
        row_num = 1
        producer_purchases_price = DECIMAL_ZERO
        while (offer_item is not None):
            department_for_customer_save = offer_item.department_for_customer
            c = ws.cell(row=row_num, column=1)
            c.value = department_for_customer_save.short_name
            c.style.font.bold = True
            row_num += 1
            row_start_department = row_num
            department_purchases_price = DECIMAL_ZERO
            while (offer_item is not None and department_for_customer_save.id == offer_item.department_for_customer_id):
                if offer_item.unit_deposit != DECIMAL_ZERO:
                    hide_column_unit_deposit = False
                show_column_reference = show_column_reference | len(offer_item.reference) < 36
                if offer_item.wrapped: # or offer_item.limit_order_quantity_to_stock:
                    hide_column_short_basket_name = False
                    first_purchase = True
                    for purchase in PurchaseOpened.objects.filter(
                        offer_item_id=offer_item.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).exclude(
                        quantity_ordered=0
                    ).order_by(
                        "offer_item__translations__order_sort_order",
                        "customer__short_basket_name"
                    ):
                        if offer_item.limit_order_quantity_to_stock:
                            # Don't purchase anything to the producer in this case
                            qty = DECIMAL_ZERO
                        else:
                            qty = purchase.quantity_ordered
                        c = ws.cell(row=row_num, column=0)
                        c.value = purchase.customer.short_basket_name
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=1)
                        c.value = qty
                        c.style.number_format.format_code = '#,##0.???'
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c.style.font.bold = True
                        c.style.font.color = Color(Color.BLUE)
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(2) + str(row_num+1), 'notEqual',
                            [str(qty)], True, wb,
                            None, None, yellowFill
                        )
                        c = ws.cell(row=row_num, column=2)
                        c.value = offer_item.reference if len(offer_item.reference) < 36 else ""
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=3)
                        c.value = offer_item.get_long_name()
                        c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        if first_purchase:
                            c.style.font.bold = True
                            first_purchase = False
                        c = ws.cell(row=row_num, column=4)
                        c.value = purchase.get_producer_unit_price()
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=5)
                        c.value = offer_item.unit_deposit
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        c = ws.cell(row=row_num, column=6)
                        if offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                            c.value = '=ROUND(B%s*%s*(E%s+F%s),2)' % (row_num + 1, offer_item.order_average_weight, row_num + 1, row_num + 1)
                            purchase_price = (qty * offer_item.order_average_weight *
                                (purchase.get_producer_unit_price() +
                                purchase.offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                        else:
                            c.value = '=ROUND(B%s*(E%s+F%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                            purchase_price = (qty *
                                (purchase.get_producer_unit_price() +
                                purchase.offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                        c.style.borders.bottom.border_style = Border.BORDER_THIN
                        department_purchases_price += purchase_price
                        ws.conditional_formatting.addCellIs(
                            get_column_letter(7) + str(row_num+1), 'notEqual',
                            [str(purchase_price)], True, wb,
                            None, None, yellowFill
                        )

                        row_num += 1
                else:
                    qty, stock = offer_item.get_producer_qty_stock_invoiced()
                    if offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG and offer_item.order_average_weight > DECIMAL_ZERO:
                        qty = (qty / offer_item.order_average_weight).quantize(TWO_DECIMALS)
                        stock = DECIMAL_ZERO
                    c = ws.cell(row=row_num, column=0)
                    c.value = repanier_settings['GROUP_NAME']
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=1)
                    c.value = qty
                    c.style.number_format.format_code = '#,##0.???'
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c.style.font.bold = True
                    c.style.font.color = Color(Color.BLUE)
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(2) + str(row_num+1), 'notEqual',
                        [str(qty)], True, wb,
                        None, None, yellowFill
                    )
                    c = ws.cell(row=row_num, column=2)
                    c.value = offer_item.reference if len(offer_item.reference) < 36 else ""
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=3)
                    c.value = offer_item.get_long_name()
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c.style.font.bold = True
                    c = ws.cell(row=row_num, column=4)
                    c.value = offer_item.producer_unit_price
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=5)
                    c.value = offer_item.unit_deposit
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=6)
                    if offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                        c.value = '=ROUND(B%s*%s*(E%s+F%s),2)' % (row_num + 1, offer_item.order_average_weight, row_num + 1, row_num + 1)
                        purchase_price = (qty * offer_item.order_average_weight *
                            (offer_item.producer_unit_price +
                            offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                    else:
                        c.value = '=ROUND(B%s*(E%s+F%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                        purchase_price = (qty *
                            (offer_item.producer_unit_price +
                            offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    department_purchases_price += purchase_price
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(7) + str(row_num+1), 'notEqual',
                        [str(purchase_price)], True, wb,
                        None, None, yellowFill
                    )

                    row_num += 1
                offer_item = next_row(offer_items)
            for col_num in xrange(7):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if col_num == 3:
                    c.value = "%s %s" % (_("Total Price"), department_for_customer_save.short_name)
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                if col_num == 6:
                    formula = 'SUM(G%s:G%s)' % (row_start_department, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.font.bold = True
                    producer_purchases_price += department_purchases_price
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(7) + str(row_num+1), 'notEqual',
                        [str(department_purchases_price)], True, wb,
                        None, None, yellowFill
                    )
                    formula_main_total.append(formula)
            row_num += 1
        for col_num in xrange(7):
            c = ws.cell(row=row_num, column=col_num)
            c.style.borders.bottom.border_style = Border.BORDER_THIN
            if col_num == 1:
                c.value = "%s %s" % (_("Total Price"), repanier_settings['GROUP_NAME'])
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            if col_num == 6:
                c.value = "=" + "+".join(formula_main_total)
                c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                c.style.font.bold = True
                ws.conditional_formatting.addCellIs(
                    get_column_letter(7) + str(row_num+1), 'notEqual',
                    [str(producer_purchases_price)], True, wb,
                    None, None, yellowFill
                )

        if hide_column_short_basket_name:
            ws.column_dimensions[get_column_letter(1)].visible = False
        if not show_column_reference:
            ws.column_dimensions[get_column_letter(3)].visible = False
        if hide_column_unit_deposit:
            ws.column_dimensions[get_column_letter(6)].visible = False

    return wb


def export_producer_by_customer(permanence, producer, wb=None):
    translation.activate(producer.language)

    header = [
        (_("Quantity"), 10),
        (_("Reference"), 20),
        (_("Product"), 60),
        (_("Unit Price"), 10),
        (_("Deposit"), 10),
        (_("Total Price"), 12),
    ]
    show_column_reference = False
    hide_column_unit_deposit = True
    formula_main_total = []
    purchases = PurchaseOpened.objects.filter(
        permanence_id=permanence.id,
        producer_id=producer.id,
        offer_item__translations__language_code=translation.get_language()
    ).exclude(quantity_ordered=0
    ).order_by(
        "customer__short_basket_name",
        "offer_item__translations__order_sort_order"
    ).iterator()
    purchase = next_row(purchases)
    if purchase:
        wb, ws = new_landscape_a4_sheet(
            wb,
            "%s %s" % (producer.short_profile_name, _("by basket")),
            permanence,
            header
        )
        row_num = 1
        while (purchase is not None):
            customer_save = purchase.customer
            c = ws.cell(row=row_num, column=0)
            c.value = customer_save.short_basket_name
            c.style.font.bold = True
            row_num += 1
            row_start_customer = row_num
            first_purchase = True
            while (purchase is not None and customer_save.id == purchase.customer_id):
                c = ws.cell(row=row_num, column=0)
                c.value = purchase.quantity_ordered
                c.style.number_format.format_code = '#,##0.???'
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c.style.font.bold = True
                c = ws.cell(row=row_num, column=1)
                c.value = purchase.offer_item.reference if len(purchase.offer_item.reference) < 36 else ""
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=2)
                c.value = purchase.get_long_name()
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if first_purchase:
                    c.style.font.bold = True
                    first_purchase = False
                c = ws.cell(row=row_num, column=3)
                c.value = purchase.get_producer_unit_price()
                c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=4)
                c.value = purchase.offer_item.unit_deposit
                c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c = ws.cell(row=row_num, column=5)
                if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                    c.value = '=ROUND(A%s*%s*(D%s+E%s),2)' % (row_num + 1, purchase.offer_item.order_average_weight, row_num + 1, row_num + 1)
                else:
                    c.value = '=ROUND(A%s*(D%s+E%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                row_num += 1
                purchase = next_row(purchases)
            for col_num in xrange(6):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if col_num == 2:
                    c.value = "%s %s" % (_("Total Price"), customer_save.short_basket_name)
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                if col_num == 5:
                    formula = 'SUM(F%s:F%s)' % (row_start_customer, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.font.bold = True
                    formula_main_total.append(formula)
            row_num += 1
        for col_num in xrange(6):
            c = ws.cell(row=row_num, column=col_num)
            c.style.borders.bottom.border_style = Border.BORDER_THIN
            if col_num == 0:
                c.value = "%s %s" % (_("Total Price"), repanier_settings['GROUP_NAME'])
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            if col_num == 5:
                c.value = "=" + "+".join(formula_main_total)
                c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                c.style.font.bold = True
        if hide_column_unit_deposit:
            ws.column_dimensions[get_column_letter(5)].visible = False
        if not show_column_reference:
            ws.column_dimensions[get_column_letter(2)].visible = False

    return wb


def export_customer(permanence, customer=None, wb=None, ws_preparation_title=None, deposit=False):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Purchase"), 5),
        (_("OfferItem"), 5),
        (_("Placement"), 15),
        (_("Producer"), 15),
        (_("Product"), 60),
        (_("Quantity"), 10),
        (_("Unit Price"), 10),
        (_("Deposit"), 10),
        (_("Total Price"), 12),
        (_("Basket"), 20),
    ]
    if customer is not None:
        translation.activate(customer.language)
        if deposit:
            purchases = PurchaseOpened.objects.filter(
                permanence_id=permanence.id,
                customer_id=customer.id,
                producer__isnull=False,
                offer_item__translations__language_code=translation.get_language(),
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "offer_item__translations__order_sort_order"
            ).iterator()
        else:
            purchases = PurchaseOpened.objects.filter(
                permanence_id=permanence.id,
                customer_id=customer.id,
                producer__isnull=False,
                offer_item__translations__language_code=translation.get_language(),
            ).exclude(
                quantity_ordered__lte=DECIMAL_ZERO,
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "offer_item__translations__order_sort_order"
            ).iterator()
    else:
        if deposit:
            purchases = PurchaseOpened.objects.filter(
                permanence_id=permanence.id,
                producer__isnull=False,
                offer_item__translations__language_code=translation.get_language(),
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "customer__short_basket_name",
                "offer_item__translations__order_sort_order"
            ).iterator()
        else:
            purchases = PurchaseOpened.objects.filter(
                permanence_id=permanence.id,
                producer__isnull=False,
                offer_item__translations__language_code=translation.get_language(),
            ).exclude(
                quantity_ordered__lte=DECIMAL_ZERO,
                offer_item__order_unit=PRODUCT_ORDER_UNIT_DEPOSIT
            ).order_by(
                "customer__short_basket_name",
                "offer_item__translations__order_sort_order"
            ).iterator()
    purchase = next_purchase(purchases)
    if purchase:
        if deposit:
            if repanier_settings['PAGE_BREAK_ON_CUSTOMER_CHECK']:
                # Change the orientation to reduce the number of page breaks, i.e. the number of printed pages
                wb, ws = new_landscape_a4_sheet(
                    wb,
                    _('Deposits'),
                    permanence,
                    header
                )
            else:
                wb, ws = new_landscape_a4_sheet(
                    wb,
                    _('Deposits'),
                    permanence,
                    header
                )
        else:
            if repanier_settings['PAGE_BREAK_ON_CUSTOMER_CHECK']:
                # Change the orientation to reduce the number of page breaks, i.e. the number of printed pages
                wb, ws = new_landscape_a4_sheet(
                    wb,
                    _('Customer check'),
                    permanence,
                    header
                )
            else:
                wb, ws = new_landscape_a4_sheet(
                    wb,
                    _('Customer check'),
                    permanence,
                    header
                )
        hide_column_placement = True
        placement_save = purchase.offer_item.placement
        row_num = 1
        while (purchase is not None):
            customer_save = purchase.customer
            row_start_customer = row_num + 1
            first_purchase = True
            while purchase is not None and customer_save.id == purchase.customer_id:
                department_for_customer_save = purchase.offer_item.department_for_customer
                c = ws.cell(row=row_num, column=4)
                c.value = department_for_customer_save.short_name
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c.style.alignment.horizontal = 'right'
                c.style.font.italic = True
                c = ws.cell(row=row_num, column=9)
                c.value = customer_save.short_basket_name
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                if first_purchase:
                    c.style.font.bold = True
                    first_purchase = False
                c.style.font.italic = True
                row_num += 1
                while (purchase is not None and customer_save.id == purchase.customer_id and department_for_customer_save.id == purchase.offer_item.department_for_customer_id):
                    if placement_save != purchase.offer_item.placement:
                        hide_column_placement = False
                    qty = purchase.quantity_ordered
                    # if qty != 0 or purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_DEPOSIT:
                    c = ws.cell(row=row_num, column=0)
                    c.value = purchase.id
                    c = ws.cell(row=row_num, column=1)
                    c.value = purchase.offer_item.id
                    c = ws.cell(row=row_num, column=2)
                    c.value = purchase.offer_item.get_placement_display()
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=3)
                    c.value = purchase.producer.short_profile_name
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=4)
                    c.value = purchase.get_long_name()
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=5)
                    if ws_preparation_title is None:
                        c.value = qty
                    else:
                        c.value = "=SUMIF('%s'!A2:A5000,A%s,'%s'!H2:H5000)" % \
                            (ws_preparation_title, row_num + 1, ws_preparation_title)
                    c.style.number_format.format_code = '#,##0.???'
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    if not deposit:
                        c.style.font.color = Color(Color.BLUE)
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(6) + str(row_num+1), 'notEqual',
                        [str(qty)], True, wb,
                        None, None, yellowFill
                    )
                    c = ws.cell(row=row_num, column=6)
                    c.value = purchase.get_customer_unit_price()
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=7)
                    c.value = purchase.offer_item.unit_deposit
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c = ws.cell(row=row_num, column=8)
                    if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG:
                        purchases_price = (qty * purchase.offer_item.order_average_weight *
                            (purchase.get_customer_unit_price() +
                            purchase.offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                        c.value = '=ROUND(F%s*%s*(G%s+H%s),2)' % (row_num + 1, purchase.offer_item.order_average_weight, row_num + 1, row_num + 1)
                    else:
                        purchases_price = (qty *
                            (purchase.get_customer_unit_price() +
                            purchase.offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                        c.value = '=ROUND(F%s*(G%s+H%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(9) + str(row_num+1), 'notEqual',
                        [str(purchases_price)], True, wb,
                        None, None, yellowFill
                    )
                    # c = ws.cell(row=row_num, column=10)
                    # c.value = str(purchases_price)

                    c = ws.cell(row=row_num, column=9)
                    c.value = purchase.customer.short_basket_name
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                    c.style.font.italic = True
                    row_num += 1
                    purchase = next_purchase(purchases)
            row_num += 1
            c = ws.cell(row=row_num, column=7)
            c.value = "%s %s" % (_("Total Price"), customer_save.long_basket_name)
            c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
            c.style.font.bold = True
            c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
            c = ws.cell(row=row_num, column=8)
            c.value = '=SUM(I%s:I%s)' % (row_start_customer, row_num)
            c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
            c.style.font.bold = True
            # Display a separator line between customers
            row_num += 1
            for col_num in xrange(10):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
            row_num += 2
            if deposit and repanier_settings['PAGE_BREAK_ON_CUSTOMER_CHECK']:
                ws.page_breaks.append(row_num)

        ws.column_dimensions[get_column_letter(1)].visible = False
        ws.column_dimensions[get_column_letter(2)].visible = False
        if hide_column_placement:
            ws.column_dimensions[get_column_letter(3)].visible = False
    return wb


def admin_customer_export(request, queryset):
    permanence = queryset.filter(status__gte=PERMANENCE_OPENED).first()
    if permanence is not None:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = ("%s - %s.xlsx" % (_("Customer"), permanence)).encode('ascii', errors='replace').replace('?', '_')
        response['Content-Disposition'] = 'attachment; filename=' + filename
        wb = export(permanence=permanence, wb=None)
        wb = export_preparation(permanence=permanence, wb=wb)
        if wb is not None:
            ws_preparation_title = cap("%s" % (_("Preparation")), 31)
            wb = export_customer(permanence=permanence, wb=wb, ws_preparation_title=ws_preparation_title)
            ws_customer_title = cap("%s" % (_('Customer check')), 31)
            wb = export_customer(permanence=permanence, wb=wb, deposit=True)
        else:
            ws_customer_title = None
        wb = export_stock(permanence=permanence, wb=wb, ws_customer_title=ws_customer_title)

        if wb is not None:
            wb.save(response)
        return response

def admin_producer_export(request, queryset):
    permanence = queryset.filter(status__gte=PERMANENCE_OPENED).first()
    if permanence is not None:
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename = ("%s - %s.xlsx" % (_("Producer"), permanence)).encode('ascii', errors='replace').replace('?', '_')
        response['Content-Disposition'] = 'attachment; filename=' + filename
        wb=None
        producer_set = Producer.objects.filter(permanence=permanence).order_by("short_profile_name")
        for producer in producer_set:
            wb = export_producer_by_product(permanence=permanence, producer=producer, wb=wb)
            if not producer.manage_stock:
                wb = export_producer_by_customer(permanence=permanence, producer=producer, wb=wb)
        if wb is not None:
            wb.save(response)
        return response
