# -*- coding: utf-8
from __future__ import unicode_literals
from django.db import transaction
from django.utils.translation import ugettext_lazy as _
from django.http import HttpResponse
from django.utils import translation
from openpyxl import load_workbook
from openpyxl.style import Border
from openpyxl.style import Fill
from openpyxl.style import NumberFormat
from openpyxl.styles import Color
from openpyxl.workbook import Workbook
from openpyxl.cell import Cell
from django.contrib.sites.models import Site

from export_tools import *
from import_tools import *
from repanier.const import *
from repanier.models import PurchaseSend, PurchaseSendForUpdate
from repanier.models import Producer
from repanier.tools import cap
from views import import_xslx_view


def next_purchase(purchases):
    purchase = next_row(purchases)
    while purchase is not None \
            and purchase.quantity_invoiced <= DECIMAL_ZERO \
            and purchase.offer_item.order_unit < PRODUCT_ORDER_UNIT_DEPOSIT:
        purchase = next_row(purchases)
    return purchase


def export_purchase(permanence=None, year=None, queryset=None, wb=None):
    yellowFill = Fill()
    yellowFill.start_color.index = 'FFEEEE11'
    yellowFill.end_color.index = 'FFEEEE11'
    yellowFill.fill_type = Fill.FILL_SOLID

    header = [
        (_("Format"), 5),
        (_("Id"), 10),
        (_("Date"), 15),
        (_("producer"), 15),
        (_("product"), 60),
        (_("customer"), 15),
        (_("quantity invoiced"), 10),
        (_("producer unit price"), 10),
        (_("deposit"), 10),
        (_("purchase price"), 10),
        (_("tax"), 10),
        (_("rule of 3"), 10),
        (_("comment"), 30),
        (_("vat level"), 10),
    ]

    if queryset is None:
        producers = Producer.objects.filter(
            purchase__permanence_id=permanence.id
        ).distinct().iterator()
    else:
        producers = queryset.iterator()
    producer = next_row(producers)
    if producer is not None:
        wb, ws = new_landscape_a4_sheet(
            wb,
            permanence if year is None else str(year),
            _('invoices'),
            header
        )
        row_num = 1
        while producer is not None:
            producer_save = producer
            count_producer_purchase = 0
            if producer.invoice_by_basket:
                if year is None:
                    purchases = PurchaseSend.objects.filter(
                        permanence_id=permanence.id,
                        producer_id=producer.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).order_by(
                        "customer__short_basket_name",
                        "offer_item__translations__order_sort_order"
                    ).iterator()
                else:
                    purchases = PurchaseSend.objects.filter(
                        permanence__status__gte=PERMANENCE_DONE,
                        permanence__permanence_date__year=year,
                        producer_id=producer.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).order_by(
                        "permanence__permanence_date",
                        "customer__short_basket_name",
                        "offer_item__translations__order_sort_order"
                    ).iterator()
                purchase = next_purchase(purchases)
                while purchase is not None:
                    c = ws.cell(row=row_num + 1, column=3)
                    c.value = producer_save.short_profile_name
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.font.bold = True
                    row_start_producer = row_num + 1
                    producer_price = DECIMAL_ZERO
                    date_save = purchase.permanence_date
                    while purchase is not None and date_save == purchase.permanence_date:
                        c = ws.cell(row=row_num + 1, column=0)
                        c.value = "A"
                        customer_save = purchase.customer
                        c = ws.cell(row=row_num + 1, column=5)
                        c.style.font.bold = True
                        row_start_purchase = row_num + 2
                        count_purchase = 0
                        purchases_price = DECIMAL_ZERO
                        while purchase is not None and date_save == purchase.permanence_date \
                            and customer_save == purchase.customer:
                            offer_item_save = purchase.offer_item
                            department_for_customer_save = offer_item_save.department_for_customer
                            c = ws.cell(row=row_num, column=4)
                            c.value = department_for_customer_save.short_name \
                                if department_for_customer_save is not None else "---"
                            c.style.font.italic = True
                            c.style.alignment.horizontal = 'right'
                            for col_num in xrange(14):
                                c = ws.cell(row=row_num, column=col_num)
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                            row_num += 1
                            while purchase is not None and date_save == purchase.permanence_date \
                                    and customer_save == purchase.customer \
                                    and department_for_customer_save == purchase.offer_item.department_for_customer:
                                # if purchase.quantity_invoiced > DECIMAL_ZERO \
                                #         or purchase.offer_item.order_unit >= PRODUCT_ORDER_UNIT_DEPOSIT:
                                c = ws.cell(row=row_num, column=1)
                                c.value = purchase.id
                                c = ws.cell(row=row_num, column=2)
                                c.value = date_save
                                c.style.number_format.format_code = NumberFormat.FORMAT_DATE_DDMMYYYY
                                c = ws.cell(row=row_num, column=3)
                                c.value = producer_save.short_profile_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=4)
                                c.value = purchase.get_long_name()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=5)
                                c.value = customer_save.short_basket_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=6)
                                c.value = purchase.quantity_invoiced
                                c.style.number_format.format_code = '#,##0.????'
                                if year is None:
                                    c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(7) + str(row_num+1), 'notEqual',
                                        [str(purchase.quantity_invoiced)], True, wb,
                                        None, None, yellowFill
                                    )
                                c = ws.cell(row=row_num, column=7)
                                c.value = purchase.get_producer_unit_price()
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                                c = ws.cell(row=row_num, column=8)
                                c.value = purchase.offer_item.unit_deposit
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                                c = ws.cell(row=row_num, column=9)
                                c.value = '=ROUND(G%s*(H%s+I%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                                if year is None:
                                    purchase_price = (purchase.quantity_invoiced *
                                        (purchase.get_producer_unit_price() +
                                         purchase.offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                                    purchases_price += purchase_price
                                    if offer_item_save.order_unit not in [
                                        PRODUCT_ORDER_UNIT_PC, PRODUCT_ORDER_UNIT_PC_PRICE_KG,
                                        PRODUCT_ORDER_UNIT_PC_PRICE_LT, PRODUCT_ORDER_UNIT_PC_PRICE_PC]:
                                        c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(10) + str(row_num+1), 'notEqual',
                                        [str(purchase_price)], True, wb,
                                        None, None, yellowFill
                                    )
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                                c = ws.cell(row=row_num, column=10)
                                if purchase.invoiced_price_with_compensation:
                                    c.value = '=G%s*%s' % (row_num + 1, purchase.offer_item.compensation)
                                else:
                                    c.value = '=G%s*%s' % (row_num + 1, purchase.offer_item.customer_vat)
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.0000_ ;_ € * -#,##0.0000_ ;_ € * "-"??_ ;_ @_ '
                                c = ws.cell(row=row_num, column=12)
                                c.value = cap(purchase.comment, 100)
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=13)
                                c.value = purchase.offer_item.get_vat_level_display()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                count_purchase += 1
                                delta = 5
                                for col_num in xrange(5):
                                    c = ws.cell(row=row_num, column=delta+col_num)
                                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                                if count_purchase > 1:
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = "B"
                                row_num += 1

                                purchase = next_purchase(purchases)
                        count_producer_purchase += count_purchase
                        if count_purchase > 1:
                            c = ws.cell(row=row_num, column=11)
                            c.value = '=SUM(J%s:J%s)' % (row_start_purchase, row_num)
                            c.style.number_format.format_code = \
                                '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                            if year is None:
                                c.style.font.color = Color(Color.BLUE)
                                ws.conditional_formatting.addCellIs(
                                    get_column_letter(12) + str(row_num+1), 'notEqual',
                                    [str(purchases_price)], True, wb,
                                    None, None, yellowFill
                                )
                                c = ws.cell(row=row_num, column=0)
                                c.value = "C"
                                c = ws.cell(row=row_num, column=1)
                                c.value = permanence.id
                                c = ws.cell(row=row_num, column=2)
                                c.value = customer_save.id
                            row_num += 1

                        producer_price += purchases_price
            else:
                if year is None:
                    # Using quantity_for_preparation_sort_order the order is by customer__short_basket_name if the product
                    # is to be distributed by piece, otherwise by lower qty first.
                    purchases = PurchaseSend.objects.filter(
                        permanence_id=permanence.id,
                        producer_id=producer.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).order_by(  # "product__placement",
                        "offer_item__translations__order_sort_order",
                        "quantity_for_preparation_sort_order",
                        "customer__short_basket_name"
                    ).iterator()
                else:
                    purchases = PurchaseSend.objects.filter(
                        permanence__status__gte=PERMANENCE_DONE,
                        permanence__permanence_date__year=year,
                        producer_id=producer.id,
                        offer_item__translations__language_code=translation.get_language()
                    ).order_by(
                        "permanence__permanence_date",
                        "offer_item__translations__order_sort_order",
                        "quantity_for_preparation_sort_order",
                        "customer__short_basket_name"
                    ).iterator()
                purchase = next_purchase(purchases)
                while purchase is not None:
                    date_save = purchase.permanence_date
                    c = ws.cell(row=row_num + 1, column=3)
                    c.value = producer_save.short_profile_name
                    c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                    c.style.font.bold = True
                    row_start_producer = row_num + 1
                    producer_price = DECIMAL_ZERO
                    while purchase is not None and date_save == purchase.permanence_date:
                        department_for_customer_save = purchase.offer_item.department_for_customer
                        c = ws.cell(row=row_num, column=4)
                        c.value = department_for_customer_save.short_name \
                            if department_for_customer_save is not None else "---"
                        c.style.font.italic = True
                        c.style.alignment.horizontal = 'right'
                        while purchase is not None and date_save == purchase.permanence_date \
                                and producer_save == purchase.producer \
                                and department_for_customer_save == purchase.offer_item.department_for_customer:
                            c = ws.cell(row=row_num + 1, column=0)
                            c.value = "A"
                            offer_item_save = purchase.offer_item
                            row_start_offer_item = row_num + 1
                            count_offer_item = 0
                            offer_items_price = DECIMAL_ZERO
                            for col_num in xrange(14):
                                c = ws.cell(row=row_num, column=col_num)
                                c.style.borders.bottom.border_style = Border.BORDER_THIN
                            row_num += 1
                            while purchase is not None and offer_item_save == purchase.offer_item:
                                # if purchase.quantity_invoiced > DECIMAL_ZERO \
                                #         or purchase.offer_item.order_unit >= PRODUCT_ORDER_UNIT_DEPOSIT:
                                c = ws.cell(row=row_num, column=1)
                                c.value = purchase.id
                                c = ws.cell(row=row_num, column=2)
                                c.value = date_save
                                c.style.number_format.format_code = NumberFormat.FORMAT_DATE_DDMMYYYY
                                c = ws.cell(row=row_num, column=3)
                                c.value = producer_save.short_profile_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=4)
                                c.value = purchase.get_long_name()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=5)
                                c.value = purchase.customer.short_basket_name
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=6)
                                c.value = purchase.quantity_invoiced
                                c.style.number_format.format_code = '#,##0.????'
                                if year is None:
                                    c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(7) + str(row_num+1), 'notEqual',
                                        [str(purchase.quantity_invoiced)], True, wb,
                                        None, None, yellowFill
                                    )
                                c = ws.cell(row=row_num, column=7)
                                c.value = purchase.get_producer_unit_price()
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                                c = ws.cell(row=row_num, column=8)
                                c.value = purchase.offer_item.unit_deposit
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                                c = ws.cell(row=row_num, column=9)
                                c.value = '=ROUND(G%s*(H%s+I%s),2)' % (row_num + 1, row_num + 1, row_num + 1)
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                                if year is None:
                                    offer_item_price = (purchase.quantity_invoiced *
                                                        (purchase.get_producer_unit_price() +
                                                         purchase.offer_item.unit_deposit)).quantize(TWO_DECIMALS)
                                    offer_items_price += offer_item_price
                                    if offer_item_save.order_unit not in [
                                        PRODUCT_ORDER_UNIT_PC, PRODUCT_ORDER_UNIT_PC_PRICE_KG,
                                        PRODUCT_ORDER_UNIT_PC_PRICE_LT, PRODUCT_ORDER_UNIT_PC_PRICE_PC]:
                                        c.style.font.color = Color(Color.BLUE)
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(10) + str(row_num+1), 'notEqual',
                                        [str(offer_item_price)], True, wb,
                                        None, None, yellowFill
                                    )
                                c = ws.cell(row=row_num, column=10)
                                if purchase.invoiced_price_with_compensation:
                                    c.value = '=G%s*%s' % (row_num + 1, purchase.offer_item.compensation)
                                else:
                                    c.value = '=G%s*%s' % (row_num + 1, purchase.offer_item.customer_vat)
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.0000_ ;_ € * -#,##0.0000_ ;_ € * "-"??_ ;_ @_ '
                                c = ws.cell(row=row_num, column=12)
                                c.value = cap(purchase.comment, 100)
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                c = ws.cell(row=row_num, column=13)
                                c.value = purchase.offer_item.get_vat_level_display()
                                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                                count_offer_item += 1
                                delta = 5
                                for col_num in xrange(5):
                                    c = ws.cell(row=row_num, column=delta+col_num)
                                    c.style.borders.bottom.border_style = Border.BORDER_THIN
                                if count_offer_item > 1:
                                    c = ws.cell(row=row_num, column=0)
                                    c.value = "B"
                                row_num += 1

                                purchase = next_purchase(purchases)
                            count_producer_purchase += count_offer_item
                            if count_offer_item > 1:
                                c = ws.cell(row=row_num, column=11)
                                c.value = '=SUM(J%s:J%s)' % (row_start_offer_item, row_num)
                                c.style.number_format.format_code = \
                                    '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                                if year is None:
                                    if not offer_item_save.wrapped and offer_item_save.order_unit in [PRODUCT_ORDER_UNIT_KG, PRODUCT_ORDER_UNIT_PC_KG]:
                                        c.style.font.color = Color(Color.BLUE)
                                        c = ws.cell(row=row_num, column=0)
                                        c.value = "D"
                                        c = ws.cell(row=row_num, column=1)
                                        c.value = permanence.id
                                        c = ws.cell(row=row_num, column=2)
                                        c.value = offer_item_save.id
                                    ws.conditional_formatting.addCellIs(
                                        get_column_letter(12) + str(row_num+1), 'notEqual',
                                        [str(offer_items_price)], True, wb,
                                        None, None, yellowFill
                                    )
                                if purchase is None or department_for_customer_save != purchase.offer_item.department_for_customer:
                                    row_num += 1

                            producer_price += offer_items_price
            if count_producer_purchase > 0:
                row_num += 1
                c = ws.cell(row=row_num, column=8)
                c.value = "%s %s %s" % (_("Total Price"), producer_save.short_profile_name, date_save.strftime('%d-%m-%Y'))
                c.style.number_format.format_code = NumberFormat.FORMAT_TEXT
                c.style.font.bold = True
                c.style.alignment.horizontal = c.style.alignment.HORIZONTAL_RIGHT
                c = ws.cell(row=row_num, column=9)
                c.value = '=ROUND(SUM(J%s:J%s),2)' % (row_start_producer, row_num)
                c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                c.style.font.bold = True
                if year is None:
                    ws.conditional_formatting.addCellIs(
                        get_column_letter(10) + str(row_num+1), 'notEqual',
                        [str(producer_price)], True, wb,
                        None, None, yellowFill
                    )
                c = ws.cell(row=row_num, column=10)
                c.value = '=SUM(K%s:K%s)' % (row_start_producer, row_num)
                c.style.number_format.format_code = '_ € * #,##0.0000_ ;_ € * -#,##0.0000_ ;_ € * "-"??_ ;_ @_ '
                row_num += 1
                for col_num in xrange(14):
                    c = ws.cell(row=row_num, column=col_num)
                    c.style.borders.bottom.border_style = Border.BORDER_MEDIUMDASHED
                row_num += 2
            producer = next_row(producers)
    if year is None:
        ws.column_dimensions[get_column_letter(3)].visible = False
        ws.column_dimensions[get_column_letter(11)].visible = False
    else:
        ws.column_dimensions[get_column_letter(12)].visible = False
    ws.column_dimensions[get_column_letter(1)].visible = False

    return wb


def admin_export(request, queryset):
    permanence = queryset.first()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = ("%s - %s.xlsx" % (_("invoices"), permanence)).encode('ascii', errors='replace').replace('?', '_')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = export_purchase(permanence=permanence, wb=None)
    if wb is not None:
        wb.save(response)
    return response

def admin_export_in(year, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = ("%s - %s.xlsx" % (_("invoices"), str(year))).encode('ascii', errors='replace').replace('?', '_')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = export_purchase(year=year, queryset=queryset, wb=None)
    if wb is not None:
        wb.save(response)
    return response


@transaction.atomic
def import_purchase_sheet(worksheet, permanence=None,
                          customer_2_id_dict=None,
                          producer_2_id_dict=None
    ):
    error = False
    error_msg = None
    header = get_header(worksheet)
    if header:
        row_num = 1
        array_purchase = []
        rule_of_3_source = DECIMAL_ZERO
        row = get_row(worksheet, header, row_num)
        while row and not error:
            try:
                row_format = row[_('Format')]
                if row_format == "A":
                    array_purchase = []
                    rule_of_3_source = DECIMAL_ZERO
                if row_format in ["A", "B"]:
                    if row[_('Id')] is None:
                        error = True
                        error_msg = _("Row %(row_num)d : No purchase id given.") % {'row_num': row_num + 1}
                        break
                    row_id = Decimal(row[_('Id')])

                    purchase = PurchaseSendForUpdate.objects.filter(id=row_id).order_by().first()
                    if purchase is None:
                        error = True
                        error_msg = _("Row %(row_num)d : No purchase corresponding to the given purchase id.") % {
                            'row_num': row_num + 1}
                        break
                    if purchase.permanence_id != permanence.id:
                        error = True
                        error_msg = _("Row %(row_num)d : The given permanence doesn't own the given purchase id.") % {
                            'row_num': row_num + 1}
                        break
                    producer_id = None
                    if row[_('producer')] in producer_2_id_dict:
                        producer_id = producer_2_id_dict[row[_('producer')]]
                    if producer_id != purchase.producer_id:
                        error = True
                        error_msg = _("Row %(row_num)d : No valid producer.") % {'row_num': row_num + 1}
                        break
                    customer_id = None
                    customer_name = "%s" % row[_('customer')]
                    if customer_name in customer_2_id_dict:
                        customer_id = customer_2_id_dict[customer_name]
                    if customer_id != purchase.customer_id:
                        error = True
                        error_msg = _("Row %(row_num)d : No valid customer") % {'row_num': row_num + 1}
                        break
                    comment = cap(row[_('comment')], 100)
                    quantity_invoiced = DECIMAL_ZERO if row[_('quantity invoiced')] is None \
                        else Decimal(row[_('quantity invoiced')]).quantize(FOUR_DECIMALS)
                    producer_row_price = row[_('purchase price')]
                    if producer_row_price is not None:
                        producer_row_price = Decimal(producer_row_price).quantize(TWO_DECIMALS)
                        if purchase.purchase_price != producer_row_price:
                            if purchase.offer_item.order_unit not in [
                                    PRODUCT_ORDER_UNIT_PC, PRODUCT_ORDER_UNIT_PC_PRICE_KG,
                                    PRODUCT_ORDER_UNIT_PC_PRICE_LT, PRODUCT_ORDER_UNIT_PC_PRICE_PC]:
                                producer_unit_price = (purchase.offer_item.producer_unit_price +
                                                       purchase.offer_item.unit_deposit).quantize(TWO_DECIMALS)
                                if producer_unit_price != DECIMAL_ZERO:
                                    purchase.quantity_invoiced = (producer_row_price /
                                                                  producer_unit_price).quantize(FOUR_DECIMALS)
                                else:
                                    purchase.quantity_invoiced = DECIMAL_ZERO
                    elif purchase.quantity_invoiced != quantity_invoiced:
                        purchase.quantity_invoiced = quantity_invoiced
                    purchase.comment = comment
                    purchase.save()
                    rule_of_3_source += purchase.purchase_price
                    array_purchase.append(purchase)
                if row_format in ["C", "D"]:
                    rule_of_3_target = row[_('rule of 3')]
                    if rule_of_3_target is not None:
                        rule_of_3_target = Decimal(rule_of_3_target).quantize(TWO_DECIMALS)
                        if rule_of_3_target != rule_of_3_source:
                            max_purchase_counter = len(array_purchase)
                            if rule_of_3_source != DECIMAL_ZERO:
                                ratio = rule_of_3_target / rule_of_3_source
                            else:
                                if rule_of_3_target == DECIMAL_ZERO:
                                    ratio = DECIMAL_ZERO
                                else:
                                    ratio = DECIMAL_ONE
                            # print "Ratio", ratio, "new_invoiced", new_invoiced, "actual_invoice", actual_invoice
                            # Rule of 3
                            if ratio != DECIMAL_ONE:
                                adjusted_invoice = DECIMAL_ZERO
                                for i, purchase in enumerate(array_purchase, start=1):
                                    producer_unit_price = (purchase.offer_item.producer_unit_price +
                                                           purchase.offer_item.unit_deposit
                                    ).quantize(TWO_DECIMALS)
                                    if i == max_purchase_counter:
                                        delta = rule_of_3_target - adjusted_invoice
                                        if producer_unit_price != DECIMAL_ZERO:
                                            purchase.quantity_invoiced = (delta / producer_unit_price).quantize(FOUR_DECIMALS)
                                        else:
                                            purchase.quantity_invoiced = DECIMAL_ZERO
                                    else:
                                        purchase.quantity_invoiced = (purchase.quantity_invoiced * ratio).quantize(FOUR_DECIMALS)
                                        adjusted_invoice += purchase.quantity_invoiced * producer_unit_price
                                    purchase.save()


                row_num += 1
                row = get_row(worksheet, header, row_num)

            except KeyError, e:
                # Missing field
                error = True
                error_msg = _("Row %(row_num)d : A required column is missing %(error_msg)s.") % {
                    'row_num': row_num + 1, 'error_msg': str(e)}
            except Exception, e:
                error = True
                error_msg = _("Row %(row_num)d : %(error_msg)s.") % {'row_num': row_num + 1, 'error_msg': str(e)}
    return error, error_msg


def handle_uploaded_file(request, queryset, file_to_import):
    error = False
    error_msg = None
    wb = load_workbook(file_to_import)
    # dict for performance optimisation purpose : read the DB only once
    customer_buyinggroup_id, customer_2_id_dict = get_customer_2_id_dict()
    producer_buyinggroup_id, producer_2_id_dict = get_producer_2_id_dict()
    if customer_buyinggroup_id is None:
        error = True
        error_msg = _("At least one customer must represent the buying group.")
    else:
        if producer_buyinggroup_id is None:
            error = True
            error_msg = _("At least one producer must represent the buying group.")

    if not error:
        permanence = queryset.first()
        if permanence.status == PERMANENCE_SEND:
            ws = wb.get_sheet_by_name(cap("%s" % (permanence), 31))
            error, error_msg = import_purchase_sheet(
                ws, permanence=permanence,
                customer_2_id_dict=customer_2_id_dict,
                producer_2_id_dict=producer_2_id_dict
            )
            if error:
                error_msg = cap("%s" % (permanence), 31) + " > " + error_msg
        else:
            error = True
            error_msg = _("The permanence has already been invoiced.")
    return error, error_msg


def admin_import(admin_ui, admin, request, queryset, action):
    return import_xslx_view(admin_ui, admin, request, queryset, handle_uploaded_file, action=action)
