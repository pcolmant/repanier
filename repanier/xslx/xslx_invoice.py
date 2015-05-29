# -*- coding: utf-8
from __future__ import unicode_literals
from repanier.models import repanier_settings
from django.utils import translation
from django.contrib.sites.models import get_current_site
from django.http import HttpResponse
from django.utils.translation import ugettext_lazy as _
from export_tools import *
from openpyxl.style import Border
from openpyxl.style import NumberFormat
from openpyxl.workbook import Workbook
from repanier.const import *
from repanier.models import BankAccount
from repanier.models import Customer
from repanier.models import CustomerInvoice
from repanier.models import Producer
from repanier.models import ProducerInvoice
from repanier.models import PurchaseClosed
from repanier.tools import get_invoice_unit
from repanier.xslx.xslx_stock import export_stock


def export(permanence, customer=None, producer=None, wb=None, sheet_name=""):
    ws = None
    # Detail of what has been prepared
    if customer is None and producer is None:
        if ws is None:
            if wb is None:
                wb = Workbook()
                ws = wb.get_active_sheet()
            else:
                ws = wb.create_sheet()
            worksheet_setup_landscape_a4(ws, "%s %s" % (_('Account summary'), sheet_name),
                                         permanence)

        row_num = 0

        bank_account = BankAccount.objects.filter(
            permanence_id=permanence.id,
            customer__isnull=True,
            producer__isnull=True
        ).order_by().first()
        if bank_account is None:
            # Permanence not invoiced yet : Nothing to do
            return wb
        customer_set = Customer.objects.filter(is_active=True)
        for customer in customer_set:
            payment = 0
            prepared = 0
            last_customer_invoice = CustomerInvoice.objects.filter(
                customer_id=customer.id,
                invoice_sort_order__lte=bank_account.id
            ).order_by('-invoice_sort_order').first()
            if last_customer_invoice is not None:
                if last_customer_invoice.permanence_id == permanence.id:
                    balance_before = last_customer_invoice.previous_balance
                    payment = last_customer_invoice.bank_amount_in - last_customer_invoice.bank_amount_out
                    prepared = last_customer_invoice.total_price_with_tax
                    balance_after = last_customer_invoice.balance
                else:
                    balance_before = last_customer_invoice.balance
                    balance_after = last_customer_invoice.balance
            else:
                # No invoice yet.
                balance_before = customer.initial_balance
                balance_after = customer.initial_balance

            row = [
                (_('Name'), 40, customer.long_basket_name, NumberFormat.FORMAT_TEXT),
                (_('Balance before'), 15, balance_before,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Payment'), 10, payment, '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Prepared'), 10, prepared, '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Balance after'), 15, balance_after,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Name'), 20, customer.short_basket_name, NumberFormat.FORMAT_TEXT),
            ]

            if row_num == 0:
                worksheet_set_header(ws, row)
                row_num += 1

            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = row[col_num][ROW_VALUE]
                c.style.number_format.format_code = row[col_num][ROW_FORMAT]

            row_num += 1

        customer = None
        row_break = row_num
        row_num += 1

        producer_set = Producer.objects.filter(is_active=True)
        for producer in producer_set:
            payment = 0
            prepared = 0
            last_producer_invoice = ProducerInvoice.objects.filter(
                producer_id=producer.id,
                invoice_sort_order__lte=bank_account.id
            ).order_by('-invoice_sort_order').first()
            if last_producer_invoice is not None:
                if last_producer_invoice.permanence_id == permanence.id:
                    balance_before = -last_producer_invoice.previous_balance
                    payment = last_producer_invoice.bank_amount_out - last_producer_invoice.bank_amount_in
                    prepared = last_producer_invoice.total_price_with_tax
                    balance_after = -last_producer_invoice.balance
                else:
                    balance_before = -last_producer_invoice.balance
                    balance_after = -last_producer_invoice.balance
            else:
                # No invoice yet.
                balance_before = -producer.initial_balance
                balance_after = -producer.initial_balance

            row = [
                (_('Name'), 40, producer.long_profile_name, NumberFormat.FORMAT_TEXT),
                (_('Balance before'), 15, balance_before,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Payment'), 10, payment, '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Prepared'), 10, prepared, '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Balance after'), 15, balance_after,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '),
                (_('Name'), 20, producer.short_profile_name, NumberFormat.FORMAT_TEXT),
            ]

            if row_num == 0:
                worksheet_set_header(ws, row)
                row_num += 1

            for col_num in xrange(len(row)):
                c = ws.cell(row=row_num, column=col_num)
                c.value = row[col_num][ROW_VALUE]
                c.style.number_format.format_code = row[col_num][ROW_FORMAT]

            row_num += 1

        producer = None
        final_bank_amount = bank_account.bank_amount_in - bank_account.bank_amount_out
        bank_account = BankAccount.objects.filter(
            id__lt=bank_account.id,
            customer__isnull=True,
            producer__isnull=True
        ).order_by("-id").first()
        if bank_account is not None:
            initial_bank_amount = bank_account.bank_amount_in - bank_account.bank_amount_out
        else:
            # This shouldn't occur because an initial balance is automatically generated
            # if not present
            # when invoicing the very first permanence.
            initial_bank_amount = 0
        row_num += 1
        c = ws.cell(row=row_num, column=1)
        c.value = initial_bank_amount
        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
        c = ws.cell(row=row_num, column=4)
        formula = 'B%s+SUM(C%s:C%s)-SUM(C%s:C%s)' % (row_num + 1, 2, row_break, row_break + 2, row_num - 1)
        c.value = '=' + formula
        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '

        row_num += 1
        c = ws.cell(row=row_num, column=4)
        formula = 'SUM(E%s:E%s)-SUM(E%s:E%s)' % (2, row_break, row_break + 2, row_num - 2)
        c.value = '=' + formula
        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '

        row_num += 1
        c = ws.cell(row=row_num, column=4)
        c.value = final_bank_amount
        c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '

        ws = None

        purchase_set = PurchaseClosed.objects.filter(
            permanence_id=permanence.id,
            offer_item__translations__language_code=translation.get_language()
        ).order_by(
            "offer_item__translations__order_sort_order",
            "customer__short_basket_name"
        ).distinct()
        hide_producer_prices = False
        hide_customer_prices = False
    elif customer is not None:
        purchase_set = PurchaseClosed.objects.filter(
            permanence_id=permanence.id, customer=customer,
            offer_item__translations__language_code=translation.get_language()
        ).order_by(
            "offer_item__translations__order_sort_order",
        ).distinct()
        hide_producer_prices = True
        hide_customer_prices = False
    else:
        purchase_set = PurchaseClosed.objects.filter(
            permanence_id=permanence.id, producer=producer,
            offer_item__translations__language_code=translation.get_language()
        ).order_by(
            "offer_item__translations__order_sort_order",
        ).distinct()
        hide_producer_prices = False
        hide_customer_prices = True

    row_num = 0

    hide_column_deposit = True
    hide_column_compensation = True

    for purchase in purchase_set:

        if ws is None:
            if wb is None:
                wb = Workbook()
                ws = wb.get_active_sheet()
            else:
                ws = wb.create_sheet()
            if producer is not None:
                # To the producer we speak of "payment".
                # This is the detail of the payment to the producer, i.e. received products
                worksheet_setup_landscape_a4(ws, "%s %s" % (_('Payment'), sheet_name),
                                             permanence)
            else:
                # To the customer we speak of "invoice".
                # This is the detail of the invoice, i.e. sold products
                worksheet_setup_landscape_a4(ws, "%s %s" % (_('Invoice'), sheet_name),
                                             permanence)

        qty = purchase.quantity_invoiced

        if purchase.offer_item.compensation != 0:
            hide_column_compensation = False
        if purchase.offer_item.unit_deposit != 0:
            hide_column_deposit = False

        unit = get_invoice_unit(order_unit=purchase.offer_item.order_unit, qty=qty)
        row = [
            (_("Producer"), 15, purchase.producer.short_profile_name, NumberFormat.FORMAT_TEXT, False),
            (_("Basket"), 20, purchase.customer.short_basket_name, NumberFormat.FORMAT_TEXT, False),
            (_("Department"), 15,
             purchase.offer_item.department_for_customer.short_name,
             NumberFormat.FORMAT_TEXT, False),
            (_("Product"), 60, purchase.offer_item.get_long_name(), NumberFormat.FORMAT_TEXT, False),
            (_("Quantity"), 10, qty, '#,##0.????',
             True if purchase.offer_item.order_unit == PRODUCT_ORDER_UNIT_PC_KG else False),
            (_("Unit"), 10, unit, NumberFormat.FORMAT_TEXT, False),
            (_("deposit"), 10, purchase.offer_item.unit_deposit,
             '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False)]
        if hide_producer_prices:
            row += [
                ('', 10, '', NumberFormat.FORMAT_TEXT, False),
                ('', 10, '', NumberFormat.FORMAT_TEXT, False),
                ('', 10, '', NumberFormat.FORMAT_TEXT, False)
            ]
        else:
            row += [
                (_("producer unit price"), 10, purchase.get_producer_unit_price(),
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
                (_("purchase price"), 10, purchase.purchase_price,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
                (_("Vat"), 10, (purchase.offer_item.producer_vat * purchase.get_quantity()).quantize(FOUR_DECIMALS),
                 '_ € * #,##0.000_ ;_ € * -#,##0.000_ ;_ € * "-"??_ ;_ @_ ', False)
            ]

        if hide_customer_prices:
            row += [
                ('', 10, '', NumberFormat.FORMAT_TEXT, False),
                ('', 10, '', NumberFormat.FORMAT_TEXT, False),
                ('', 10, '', NumberFormat.FORMAT_TEXT, False)
            ]
        else:
            row += [
                (_("customer unit price"), 10, purchase.get_customer_unit_price(),
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
                (_("selling price"), 10, purchase.selling_price,
                 '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ ', False),
            ]
            if repanier_settings['DISPLAY_VAT']:
                row += [
                    (_("Vat"), 10, DECIMAL_ZERO if purchase.invoiced_price_with_compensation else
                    (purchase.offer_item.customer_vat * purchase.get_quantity()).quantize(FOUR_DECIMALS),
                     '_ € * #,##0.000_ ;_ € * -#,##0.000_ ;_ € * "-"??_ ;_ @_ ', False),
                ]
            else:
                row += [
                    (_("Vat"), 10, '', NumberFormat.FORMAT_TEXT, False),
                ]
        if repanier_settings['DISPLAY_VAT'] or not hide_producer_prices:
            row += [
                (_("Compensation"), 10, (purchase.offer_item.compensation * purchase.get_quantity()).quantize(FOUR_DECIMALS) if
                purchase.invoiced_price_with_compensation else  DECIMAL_ZERO,
                 '_ € * #,##0.000_ ;_ € * -#,##0.000_ ;_ € * "-"??_ ;_ @_ ', False),
            ]
        else:
            row += [
                (_("Compensation"), 10, '', NumberFormat.FORMAT_TEXT, False),
            ]
        row += [
            (_("comment"), 30, "" if purchase.comment is None else purchase.comment, NumberFormat.FORMAT_TEXT, False),
        ]

        if row_num == 0:
            worksheet_set_header(ws, row)
            row_num += 1

        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num, column=col_num)
            c.value = "%s" % (row[col_num][ROW_VALUE])
            c.style.number_format.format_code = row[col_num][ROW_FORMAT]
            if row[col_num][ROW_BOX]:
                c.style.borders.top.border_style = Border.BORDER_THIN
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                c.style.borders.left.border_style = Border.BORDER_THIN
                c.style.borders.right.border_style = Border.BORDER_THIN
            else:
                c.style.borders.bottom.border_style = Border.BORDER_HAIR
            if col_num == 7:
                c.style.font.bold = True

        row_num += 1

    if wb is not None:
        if hide_column_deposit:
            ws.column_dimensions[get_column_letter(7)].visible = False
        if hide_column_compensation:
            ws.column_dimensions[get_column_letter(14)].visible = False
        if hide_producer_prices:
            ws.column_dimensions[get_column_letter(8)].visible = False
            ws.column_dimensions[get_column_letter(9)].visible = False
            ws.column_dimensions[get_column_letter(10)].visible = False
            if not repanier_settings['DISPLAY_VAT']:
                ws.column_dimensions[get_column_letter(13)].visible = False
        if hide_customer_prices:
            ws.column_dimensions[get_column_letter(11)].visible = False
            ws.column_dimensions[get_column_letter(12)].visible = False
            ws.column_dimensions[get_column_letter(13)].visible = False
        if customer is not None or producer is not None:
            for col_num in xrange(14):
                c = ws.cell(row=row_num, column=col_num)
                c.style.borders.top.border_style = Border.BORDER_THIN
                c.style.borders.bottom.border_style = Border.BORDER_THIN
                if col_num == 1:
                    c.value = _("Total Price") + " " + sheet_name
                    c.style.font.bold = True
                if col_num == 8:
                    formula = 'SUM(I%s:I%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.font.bold = True
                if col_num == 9:
                    formula = 'SUM(J%s:J%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.font.bold = True
                if col_num == 11:
                    formula = 'SUM(L%s:L%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.font.bold = True
                if col_num == 12:
                    formula = 'SUM(M%s:M%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.font.bold = True
                if col_num == 13:
                    formula = 'SUM(N%s:N%s)' % (2, row_num)
                    c.value = '=' + formula
                    c.style.number_format.format_code = '_ € * #,##0.00_ ;_ € * -#,##0.00_ ;_ € * "-"??_ ;_ @_ '
                    c.style.font.bold = True

    return wb


def admin_export(request, queryset):
    current_site = get_current_site(request)
    permanence = queryset.first()
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = ("%s - %s.xlsx" % (_("Accounting report"), permanence)).encode('ascii', errors='replace').replace('?', '_')
    response['Content-Disposition'] = 'attachment; filename=' + filename
    wb = export(permanence=permanence, sheet_name=current_site.name)
    wb = export_stock(permanence=permanence, wb=wb, ws_customer_title=None)
    if wb is not None:
        wb.save(response)
    return response
